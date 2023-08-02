# import dependencies for flask
from flask import Flask, render_template, request

# import dependencies for sqlalchemy
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.automap import automap_base
from sqlalchemy.orm import Session
from sqlalchemy import create_engine, and_, or_

# import dependencies for openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.protection import WorkbookProtection

# import general dependencies
import json
import datetime
from enum import Enum
from inspect import stack
from os.path import join, exists
from os import startfile

# import confidential information
from sys import path
path.insert(0, "..")
from config import pg_key, pg_db, pg_host, pg_port, pg_user, data_entry_loc

# create the sqlalchemy engine
engine = create_engine(f"postgresql://{pg_user}:{pg_key}@{pg_host}:{pg_port}/{pg_db}", pool_pre_ping = True, echo = False)

# reflect the database
base = automap_base()
base.prepare(engine, reflect = True)

# instantiate the database tables
inspections_lot_numbers = base.classes.inspections_lot_numbers
inspections_purchase_orders = base.classes.inspections_purchase_orders
inspections_job_numbers = base.classes.inspections_job_numbers
employee_projects = base.classes.employee_projects
features = base.classes.features
print_features = base.classes.print_features
inspections = base.classes.inspections
gauges = base.classes.gauges
parts = base.classes.parts
inspection_records = base.classes.inspection_records
machines = base.classes.machines
projects = base.classes.projects
receiver_numbers = base.classes.receiver_numbers
purchase_orders = base.classes.purchase_orders
job_numbers = base.classes.job_numbers
employees = base.classes.employees
locations = base.classes.locations
departments = base.classes.departments
suppliers = base.classes.suppliers
lot_numbers = base.classes.lot_numbers
inspection_types = base.classes.inspection_types
frequency_types = base.classes.frequency_types
material_types = base.classes.material_types
project_types = base.classes.project_types
specification_types = base.classes.specification_types
dimension_types = base.classes.dimension_types
gauge_types = base.classes.gauge_types
machine_types = base.classes.machine_types
location_types = base.classes.location_types
disposition_types = base.classes.disposition_types
operation_types = base.classes.operation_types

# instantiate the flask app
app = Flask(__name__)
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

# --------------------------------------------------

#region messaging

# message type enumerations
class message_type(Enum):
    records_not_found = 0
    record_already_exists = 1
    record_locked = 2
    records_deleted = 3
    records_updated = 4
    records_not_deleted = 5
    records_not_updated = 6
    no_records_added = 7
    no_association_found = 8
    sql_exception = 9
    generic = 10

def text_response(current_method:str, type:message_type, **kwargs):

    output_str = ""
    if type == message_type.records_not_found:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"no matching records found in the following tables; {tables}"
    elif type == message_type.record_already_exists:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"record already exists in {tables}"
    elif type == message_type.record_locked:
        output_str = f"editing is not allowed; '{kwargs['table'].__table__.name}' is locked"
    elif type == message_type.records_deleted:
        output_str = f"{kwargs['qty']} records deleted from '{kwargs['table']}'"
    elif type == message_type.records_updated:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"{kwargs['qty']} records updated in {tables}"
    elif type == message_type.records_not_deleted:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"no records deleted from the following tables; {tables}"
    elif type == message_type.records_not_updated:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"no records updated in the following tables; {tables}"
    elif type == message_type.no_records_added:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"no records added to the following tables; {tables}"
    elif type == message_type.no_association_found:
        tables = ", ".join([f"'{x.__table__.name}'" for x in kwargs["tables"]])
        output_str = f"no association found between the following tables; {tables}"
    elif type == message_type.sql_exception:
        output_str = str(kwargs["error"])
    elif type == message_type.generic:
        output_str = kwargs["err_msg"]
    else:
        output_str = "response type not given"

    return f"def {current_method}()\n{output_str}"

#endregion

# --------------------------------------------------

#region page navigation

@app.route("/qa1_data_portal/")
def open_qa1_data_portal():
    return render_template("qa1_data_portal.html")

@app.route("/inspection_schemas/")
def open_measurement_set_schemas():
    return render_template("inspection_schemas.html")

@app.route("/inspection_records/")
def open_inspection_reports():
    return render_template("inspection_records.html")

#endregion

# --------------------------------------------------

#region get enumerations

@app.route("/get_all_dimension_types/")
def get_all_dimension_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(dimension_types.id, dimension_types.name).order_by(dimension_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }

        else:
            return {
                "status": "log",
                "response": "no records found in 'dimension_types'"
            }

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

@app.route("/get_all_gauges/")
def get_all_gauges():

    # define the requested columns
    columns = [
        gauges.id,
        gauges.name,
        gauges.last_calibrated,
        gauges.gauge_type_id,
        gauges.employee_id,
        gauges.location_id
    ]

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(*columns).order_by(gauges.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name, last_calibrated, gauge_type_id, employee_id, location_id in results:
                output_arr.append({
                    "id": id,
                    "name": name,
                    "last_calibrated": last_calibrated,
                    "gauge_type_id": gauge_type_id,
                    "employee_id": employee_id,
                    "location_id": location_id
                })

            return {
                "status": "ok",
                "response": output_arr
            }

        else:
            return {
                "status": "log",
                "response": "no records found in 'gauges'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_gauge_types/")
def get_all_gauge_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(gauge_types.id, gauge_types.name).order_by(gauge_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'gauge_types'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_specification_types/")
def get_all_specification_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(specification_types.id, specification_types.name).order_by(specification_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'specification_types'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_frequency_types/")
def get_all_frequency_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(frequency_types.id, frequency_types.name).order_by(frequency_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'frequency_types'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_inspection_types/")
def get_all_inspection_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(inspection_types.id, inspection_types.name).order_by(inspection_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'inspection_types'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_employees/")
def get_all_employees():

    # define the requested columns
    columns = [
        employees.id,
        employees.first_name,
        employees.last_name,
        employees.department_id,
        employees.location_id
    ]

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(*columns).order_by(employees.last_name.asc(), employees.first_name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, first_name, last_name, department_id, location_id in results:
                output_arr.append({
                    "id": id,
                    "name": f"{last_name}, {first_name}",
                    "department_id": department_id,
                    "location_id": location_id
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'employees'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_filtered_employees/", methods = ["POST"])
def get_filtered_employees():

    # get the posted data
    form_data = json.loads(request.data)

    # get the required parameters
    search_term = str(form_data["search_term"])

    # define the requested columns
    columns = [
        employees.id,
        employees.first_name,
        employees.last_name,
        employees.department_id,
        employees.location_id
    ]

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(*columns)\
            .filter(or_(employees.first_name.ilike(f"%{search_term}%"), employees.last_name.ilike(f"%{search_term}%")))\
            .order_by(employees.last_name.asc(), employees.first_name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, first_name, last_name, department_id, location_id in results:
                output_arr.append({
                    "id": id,
                    "name": f"{last_name}, {first_name}",
                    "department_id": department_id,
                    "location_id": location_id
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'employees'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_disposition_types/")
def get_all_disposition_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(disposition_types.id, disposition_types.name).order_by(disposition_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'disposition_types'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_parts/")
def get_all_parts():

    # define the requested columns
    columns = [
        parts.id,
        parts.drawing,
        parts.revision,
        parts.item
    ]

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(*columns).order_by(parts.item.asc(), parts.drawing.asc(), parts.revision.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, drawing, revision, item in results:
                output_arr.append({
                    "id": id,
                    "drawing": drawing,
                    "revision": revision.upper(),
                    "item": item,
                    "name": f"{item}, {drawing}, {revision.upper()}"
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'parts'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_job_numbers/")
def get_all_job_numbers():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(job_numbers.id, job_numbers.name).order_by(job_numbers.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'job_numbers'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_material_types/")
def get_all_material_types():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(material_types.id, material_types.name).order_by(material_types.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'material_types'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_suppliers/")
def get_all_suppliers():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(suppliers.id, suppliers.name).order_by(suppliers.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'suppliers'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_receiver_numbers/")
def get_all_receiver_numbers():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(receiver_numbers.id, receiver_numbers.name).order_by(receiver_numbers.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no record found in 'receiver_numbers'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_purchase_orders/")
def get_all_purchase_orders():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(purchase_orders.id, purchase_orders.name, purchase_orders.supplier_id).order_by(purchase_orders.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name, supplier_id in results:
                output_arr.append({
                    "id": id,
                    "name": name,
                    "supplier_id": supplier_id
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'purchase_orders'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

@app.route("/get_all_lot_numbers/")
def get_all_lot_numbers():

    try:

        # open the database session
        session = Session(engine)

        # query the database
        results = session.query(lot_numbers.id, lot_numbers.name).order_by(lot_numbers.name.asc()).all()

        # close the session
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, name in results:
                output_arr.append({
                    "id": id,
                    "name": name
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": "no records found in 'purchase_orders'"
            }

    except SQLAlchemyError as e:
        error_msg = str(e.__dict__["orig"])
        return {
            "status": "log",
            "response": error_msg
        }

#endregion

# --------------------------------------------------

# routes

@app.route("/qa1_data_portal/generate_excel_document/", methods = ["POST"])
def qa1_data_portal_generate_document():

    # interpret the posted data
    form_data = json.loads(request.data)

    # get the required parameters
    part_id = int(form_data["part_id"])

    # get the drawing and item
    try:
        # open the database connection
        session = Session(engine)

        # query the database
        results = session.query(parts.drawing, parts.item)\
            .filter(parts.id == part_id)\
            .first()
        if results is None:
            return {
                "status": "alert",
                "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts])
            }
        drawing, item = results

        # close the database connection
        session.close()

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

    # define the workbook file name
    wb_filename = f"wb_{drawing}_{item}.xlsx"

    # check if the workbook already exists
    if exists(join(data_entry_loc, wb_filename)):

        # access the workbook and populate with existing information

        # open the workbook
        wb = load_workbook(filename=join(data_entry_loc, wb_filename))
    else:

        # create the workbook
        wb = Workbook()
        wb.save(join(data_entry_loc, wb_filename))

        # generate format

        # open the workbook
        wb = load_workbook(filename=join(data_entry_loc, wb_filename))

    return {
        "status": "log",
        "response": text_response(stack()[0][3], message_type.generic, err_msg = "workbook created")
    }

@app.route("/qa1_data_portal/update_part_ids/", methods = ["POST"])
def qa1_data_portal_update_part_ids():

    # interpret posted data
    form_data = json.loads(request.data)

    # get the required parameters
    part_id_filter = str(form_data["part_id_filter"])

    try:

        # open the database connection
        session = Session(engine)

        # query the database
        results = session.query(parts.id, parts.drawing, parts.item)\
            .filter(or_(parts.item.ilike(f"%{part_id_filter}%"), parts.drawing.ilike(f"%{part_id_filter}%")))\
            .order_by(parts.drawing.asc(), parts.item.asc())\
            .distinct(parts.drawing, parts.item)\
            .all()

        # close the database connection
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for id, drawing, item in results:
                output_arr.append({
                    "id": id,
                    "name": f"{drawing}, {item}"
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": None
            }

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

@app.route("/qa1_data_portal/update_part_revisions/", methods = ["POST"])
def qa1_data_portal_update_part_revisions():

    # interpret posted data
    form_data = json.loads(request.data)

    # get the required parameters
    part_id = int(form_data["part_id"])

    try:

        # open the database connection
        session = Session(engine)

        # query the database
        results = session.query(parts.revision)\
            .filter(parts.id == part_id)\
            .order_by(parts.revision.asc())\
            .distinct(parts.revision)\
            .all()

        # close the database connection
        session.close()

        # return the results
        if len(results) > 0:
            output_arr = []
            for revision in results:
                output_arr.append({
                    "revision": revision[0].upper()
                })

            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "log",
                "response": None
            }

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

@app.route("/qa1_data_portal/get_associations/", methods = ["POST"])
def qa1_data_portal_get_associations():

    # interpret posted data
    form_data = json.loads(request.data)

    # get the required parameters
    part_id = int(form_data["part_id"])
    part_revision = str(form_data["part_revision"])

    try:

        # open the database connection
        session = Session(engine)

        # get the item and drawing
        item, drawing = session.query(parts.item, parts.drawing)\
            .filter(parts.id == part_id).first()

        # get the job numbers
        job_numbers_query = session.query(job_numbers.id, job_numbers.name)\
            .join(inspections_job_numbers, (inspections_job_numbers.job_number_id == job_numbers.id))\
            .join(inspections, (inspections.id == inspections_job_numbers.inspection_id))\
            .join(parts, (parts.id == inspections.part_id))\
            .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"%{drawing}%"), parts.revision.ilike(f"%{part_revision}%")))\
            .order_by(job_numbers.name.asc())\
            .distinct(job_numbers.name)\
            .all()

        # get the purchase orders
        purchase_orders_query = session.query(purchase_orders.id, purchase_orders.name)\
            .join(inspections_purchase_orders, (inspections_purchase_orders.purchase_order_id == purchase_orders.id))\
            .join(inspections, (inspections.id == inspections_purchase_orders.inspection_id))\
            .join(parts, (parts.id == inspections.part_id))\
            .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"%{drawing}%"), parts.revision.ilike(f"%{part_revision}%")))\
            .order_by(purchase_orders.name.asc())\
            .distinct(purchase_orders.name)\
            .all()

        # close the database connection
        session.close()

        # build the return object
        jn_output = []
        if len(job_numbers_query) > 0:
            for id, name in job_numbers_query:
                jn_output.append({
                    "id": id,
                    "name": name
                })
        else:
            jn_output = None
        po_output = []
        if len(purchase_orders_query) > 0:
            for id, name in purchase_orders_query:
                po_output.append({
                    "id": id,
                    "name": name
                })
        else:
            po_output = None

        # return the results
        return {
            "status": "ok",
            "response": {
                "job_numbers": jn_output,
                "purchase_orders": po_output
            }
        }

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

@app.route("/qa1_data_portal/get_receiver_numbers/", methods = ["POST"])
def qa1_data_portal_get_receiver_numbers():

    # interpret posted data
    form_data = json.loads(request.data)

    # get the required parameters
    purchase_order = int(form_data["purchase_order"])

    try:

        # open the database session
        session = Session(engine)

        # query the database
        query = session.query(receiver_numbers.id, receiver_numbers.name)\
            .join(purchase_orders, (purchase_orders.id == receiver_numbers.purchase_order_id))\
            .filter(purchase_orders.id == purchase_order)\
            .order_by(receiver_numbers.name.asc())\
            .all()

        # close the database session
        session.close()

        if len(query) > 0:
            output_arr = []
            for id, name in query:
                output_arr.append({
                    "id": id,
                    "name": name
                })
            return {
                "status": "ok",
                "response": output_arr
            }
        else:
            return {
                "status": "ok",
                "response": None
            }

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

@app.route("/qa1_data_portal/get_inspection_record/", methods = ["POST"])
def qa1_data_portal_get_inspection_record():

    # interpret posted data
    form_data = json.loads(request.data)

    # get the required parameters
    part_id = int(form_data["part_id"])
    part_revision = str(form_data["part_revision"])
    job_number = int(form_data["job_number"])
    purchase_order = int(form_data["purchase_order"])

    # get the required data
    data = retrieve_data(part_id, part_revision, job_number, purchase_order)

    if data["status"] == "ok":

        # create and format the workbook
        wb_filename = data["response"]["filename"]
        wb_address = join(data_entry_loc, wb_filename)

        # load the blank template
        wb = load_workbook(join(data_entry_loc, "_template.xlsm"), keep_vba = True)

        # format
        format_workbook(wb, data["response"])

        # save the changes under a unique name
        try:
            wb.save(wb_address)
            wb.close()
        except PermissionError as e:
            return {
                "status": "alert",
                "response": text_response(stack()[0][3], message_type.generic, err_msg = "Workbook is already in use.")
            }

        # open the edited file
        startfile(wb_address)

        return {
            "status": "ok",
            "response": None
        }
    else:
        return data

# recycled methods

def retrieve_data(part_id:int, part_revision:str, job_number:int, purchase_order:int):

    try:

        # open the database connection
        session = Session(engine)

        # get the item and drawing
        item, drawing = session.query(parts.item, parts.drawing)\
            .filter(parts.id == part_id).first()

        # print detail fields
        print_detail_fields = [
            print_features.id,
            print_features.name,
            print_features.nominal,
            print_features.usl,
            print_features.lsl,
            print_features.precision,
            specification_types.name,
            dimension_types.name,
            frequency_types.name,
            operation_types.name,
            gauge_types.name
        ]

        # construct the print details query
        print_details_query = session.query(*print_detail_fields)\
            .join(specification_types, (specification_types.id == print_features.specification_type_id))\
            .join(dimension_types, (dimension_types.id == print_features.dimension_type_id))\
            .join(frequency_types, (frequency_types.id == print_features.frequency_type_id))\
            .join(operation_types, (operation_types.id == print_features.operation_type_id))\
            .join(features, (features.print_feature_id == print_features.id))\
            .join(gauges, (gauges.id == features.gauge_id))\
            .join(gauge_types, (gauge_types.id == gauges.gauge_type_id))\
            .join(inspections, (inspections.id == features.inspection_id))\
            .join(parts, (parts.id == inspections.part_id))\
            .filter(parts.id == part_id)\
            .distinct(print_features.id, print_features.name)\
            .order_by(print_features.id.asc())\
            .all()

        # construct the print details data object
        print_details_list = None
        if len(print_details_query) > 0:
            print_details_list = []
            for id, name, nominal, usl, lsl, precision, spec_type, dim_type, freq_type, op_type, gauge_type in print_details_query:
                print_details_list.append({
                    "id": id,
                    "name": name,
                    "nominal": nominal,
                    "usl": usl,
                    "lsl": lsl,
                    "precision": precision,
                    "specification_type": spec_type,
                    "dimension_type": dim_type,
                    "frequency_type": freq_type,
                    "operation_type": op_type,
                    "gauge_type": gauge_type
                })
        else:
            session.close()
            return {
                "status": "log",
                "response": text_response(stack()[0][3], message_type.records_not_found, tables = [features, print_features, specification_types, dimension_types, frequency_types, operation_types, parts, inspections, gauges])
            }

        # inspections fields
        inspections_fields = [
            inspections.id,
            inspections.part_index,
            inspections.datetime_measured,
            employees.id,
            employees.first_name,
            employees.last_name
        ]

        # construct the inspections query
        inspections_query = session.query(*inspections_fields)\
            .join(employees, (employees.id == inspections.employee_id))\
            .join(parts, (parts.id == inspections.part_id))\
            .filter(parts.id == part_id)\
            .filter(parts.revision.ilike(part_revision))

        # association filter
        metadata_obj = None
        if job_number > -1:

            # refine inspections query
            inspections_query = inspections_query.join(inspections_job_numbers, (inspections_job_numbers.inspection_id == inspections.id))\
                .join(job_numbers, (job_numbers.id == inspections_job_numbers.job_number_id))\
                .filter(job_numbers.id == job_number).all()

            # define metadata fields
            metadata_fields = [
                inspection_records.employee_id,
                disposition_types.name,
                job_numbers.employee_id,
                job_numbers.full_inspect_interval,
                job_numbers.released_qty,
                job_numbers.completed_qty,
                job_numbers.production_rate,
                material_types.name,
                locations.name,
                job_numbers.name
            ]

            # construct & execute metadata query
            metadata_query = session.query(*metadata_fields)\
                .join(inspections, (inspections.inspection_record_id == inspection_records.id))\
                .join(inspections_job_numbers, (inspections_job_numbers.inspection_id == inspections.id))\
                .join(job_numbers, (job_numbers.id == inspections_job_numbers.job_number_id))\
                .join(parts, (parts.id == inspections.part_id))\
                .join(disposition_types, (disposition_types.id == inspection_records.disposition_id))\
                .join(material_types, (material_types.id == job_numbers.material_type_id))\
                .join(locations, (locations.id == job_numbers.location_id))\
                .filter(inspections.part_id == part_id)\
                .filter(parts.revision.ilike(part_revision))\
                .filter(job_numbers.id == job_number)\
                .first()

            inspector_fname, inspector_lname = session.query(employees.first_name, employees.last_name)\
                .filter(employees.id == metadata_query[0])\
                .first()
            operator_fname, operator_lname = session.query(employees.first_name, employees.last_name)\
                .filter(employees.id == metadata_query[2])\
                .first()

            metadata_obj = {
                "type": "internal",
                "date": datetime.datetime.now().date(),
                "inspector": f"{inspector_fname} {inspector_lname}",
                "disposition": metadata_query[1],
                "job_number": metadata_query[9],
                "operator": f"{operator_fname} {operator_lname}",
                "qc_full_inspect_interval": metadata_query[3],
                "released_quantity": metadata_query[4],
                "completed_quantity": metadata_query[5],
                "production_rate": metadata_query[6],
                "material_type": metadata_query[7],
                "workcenter": metadata_query[8]
            }

        elif purchase_order > -1:
            inspections_query = inspections_query.join(inspections_purchase_orders, (inspections_purchase_orders.inspection_id == inspections.id))\
                .join(purchase_orders, (purchase_orders.id == inspections_purchase_orders.purchase_order_id))\
                .filter(purchase_orders.id == purchase_order).all()

            # define metadata fields
            metadata_fields = [
                inspection_records.employee_id,
                disposition_types.name,
                purchase_orders.name,
                receiver_numbers.name,
                suppliers.name,
                receiver_numbers.received_qty
            ]

            # construct & execute metadata query
            metadata_query = session.query(*metadata_fields)\
                .join(inspections, (inspections.inspection_record_id == inspection_records.id))\
                .join(inspections_purchase_orders, (inspections_purchase_orders.inspection_id == inspections.id))\
                .join(purchase_orders, (purchase_orders.id == inspections_purchase_orders.purchase_order_id))\
                .join(parts, (parts.id == inspections.part_id))\
                .join(disposition_types, (disposition_types.id == inspection_records.disposition_id))\
                .join(suppliers, (suppliers.id == purchase_orders.supplier_id))\
                .join(receiver_numbers, (receiver_numbers.purchase_order_id == purchase_orders.id))\
                .filter(inspections.part_id == part_id)\
                .filter(parts.revision.ilike(part_revision))\
                .filter(purchase_orders.id == purchase_order)\
                .first()

            inspector_fname, inspector_lname = session.query(employees.first_name, employees.last_name)\
                .filter(employees.id == metadata_query[0])\
                .first()

            metadata_obj = {
                "type": "freight",
                "date": datetime.datetime.now().date(),
                "inspector": f"{inspector_fname} {inspector_lname}",
                "disposition": metadata_query[1],
                "purchase_order": metadata_query[2],
                "receiver_number": metadata_query[3],
                "supplier": metadata_query[4],
                "received_quantity": metadata_query[5]
            }

        # construct the inspections data object
        inspections_list = None
        if len(inspections_query) > 0:
            inspections_list = []
            for inspection_id, part_index, datetime_measured, employee_id, first_name, last_name in inspections_query:

                # get the inspection features
                features_query = session.query(features.measured, features.print_feature_id, gauges.name)\
                    .join(gauges, (gauges.id == features.gauge_id))\
                    .filter(features.inspection_id == inspection_id)\
                    .all()

                # get all the features
                features_list = None
                if len(features_query) > 0:
                    features_list = []
                    for measured, print_feature_id, gauge in features_query:
                        features_list.append({
                            "measured": measured,
                            "print_feature_id": print_feature_id,
                            "gauge": gauge
                        })

                inspections_list.append({
                    "inspection_id": inspection_id,
                    "part_index": part_index,
                    "datetime_measured": datetime_measured,
                    "employee_id": employee_id,
                    "employee_name": f"{first_name} {last_name}",
                    "features": features_list
                })
        else:
            session.close()
            return {
                "status": "log",
                "response": text_response(stack()[0][3], message_type.records_not_found, tables = [inspections, employees, parts])
            }

        # list of employees
        employees_query = session.query(employees.first_name, employees.last_name)\
            .order_by(employees.last_name.asc(), employees.first_name.asc())\
            .all()
        employees_list = []
        for first_name, last_name in employees_query:
            employees_list.append(f"{first_name} {last_name}")
        employees_str = '"' + ','.join(employees_list) + '"'

        # list of dispositions
        dispositions_query = session.query(disposition_types.id, disposition_types.name)\
            .order_by(disposition_types.name.asc())\
            .all()
        dispositions_list = []
        for id, name in dispositions_query:
            dispositions_list.append(name)
        dispositions_str = '"' + ','.join(dispositions_list) + '"'

        # list of material types
        material_types_query = session.query(material_types.id, material_types.name)\
            .order_by(material_types.name.asc())\
            .all()
        material_types_list = []
        for id, name in material_types_query:
            material_types_list.append(name)
        material_types_str = '"' + ','.join(material_types_list) + '"'

        # list of work centers
        locations_query = session.query(locations.id, locations.name)\
            .order_by(locations.name.asc())\
            .all()
        locations_list = []
        for id, name in locations_query:
            locations_list.append(name)
        locations_str = '"' + ','.join(locations_list) + '"'

        # list of suppliers
        suppliers_query = session.query(suppliers.id, suppliers.name)\
            .order_by(suppliers.name.asc())\
            .all()
        suppliers_list = []
        for id, name in suppliers_query:
            suppliers_list.append(name)
        suppliers_str = '"' + ','.join(suppliers_list) + '"'

        # list of gauges
        gauges_query = session.query(gauges.id, gauges.name)\
            .order_by(gauges.name.asc())\
            .all()
        gauges_list = []
        for id, name in gauges_query:
            gauges_list.append(name)
        gauges_str = '"' + ','.join(gauges_list) + '"'

        # close the database connection
        session.close()

        # return the results
        return {
            "status": "ok",
            "response": {
                "data_validation": {
                    "employees": employees_str,
                    "dispositions": dispositions_str,
                    "material_types": material_types_str,
                    "locations": locations_str,
                    "suppliers": suppliers_str,
                    "gauges": gauges_str
                },
                "item": item,
                "drawing": drawing,
                "revision": part_revision.upper(),
                "print": print_details_list,
                "inspections": inspections_list,
                "metadata": metadata_obj,
                "filename": f"{item}, {drawing}, {part_revision.upper()}.xlsm"
            }
        }

    except SQLAlchemyError as e:
        return {
            "status": "log",
            "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
        }

def format_workbook(wb:Workbook, data):

    #region define formatting objects
    font_header = Font(
        name = "Calibri",
        size = 11,
        bold = True,
        italic = True,
        vertAlign = "baseline",
        underline = "none",
        color = "FF000000"
    )
    font_data = Font(
        name = "Calibri",
        size = 11,
        bold = False,
        italic = False,
        vertAlign = "baseline",
        underline = "none",
        color = "FF000000"
    )
    fill_header = PatternFill(
        fill_type = "solid",
        start_color = "FFD9D9D9",
        end_color = "FFD9D9D9"
    )
    fill_data = PatternFill(
        fill_type = "solid",
        start_color = "FFF2F2F2",
        end_color = "FFF2F2F2"
    )
    border_header = Border(
        left = Side(border_style = "thin", color = "FF808080"),
        top = Side(border_style = "thin", color = "FF808080"),
        right = Side(border_style = "thin", color = "FF808080"),
        bottom = Side(border_style = "thin", color = "FF808080"),
        diagonal = Side(border_style = None, color = "FF000000"),
        diagonal_direction = Side(border_style = None, color = "FF000000"),
        outline = Side(border_style = None, color = "FF000000"),
        vertical = Side(border_style = None, color = "FF000000"),
        horizontal = Side(border_style = None, color = "FF000000")
    )
    border_data = Border(
        left = Side(border_style = "thin", color = "FFA6A6A6"),
        top = Side(border_style = "thin", color = "FFA6A6A6"),
        right = Side(border_style = "thin", color = "FFA6A6A6"),
        bottom = Side(border_style = "thin", color = "FFA6A6A6"),
        diagonal = Side(border_style = None, color = "FF000000"),
        diagonal_direction = Side(border_style = None, color = "FF000000"),
        outline = Side(border_style = None, color = "FF000000"),
        vertical = Side(border_style = None, color = "FF000000"),
        horizontal = Side(border_style = None, color = "FF000000")
    )
    alignment_header = Alignment(
        horizontal = "left",
        vertical = "bottom",
        text_rotation = 0,
        wrap_text = False,
        shrink_to_fit = False,
        indent = 0
    )
    alignment_data = Alignment(
        horizontal = "right",
        vertical = "bottom",
        text_rotation = 0,
        wrap_text = False,
        shrink_to_fit = False,
        indent = 0
    )
    fill_pass = PatternFill(
        fill_type = "solid",
        start_color = "FF20df20",
        end_color = "FF20df20"
    )
    fill_fail = PatternFill(
        fill_type = "solid",
        start_color = "FFdf2020",
        end_color = "FFdf2020"
    )
    #endregion

    # assemble format object
    format_obj = {
        "fill": {
            "data": fill_data,
            "header": fill_header
        },
        "font": {
            "data": font_data,
            "header": font_header
        },
        "border": {
            "data": border_data,
            "header": border_header
        },
        "alignment": {
            "data": alignment_data,
            "header": alignment_header
        },
        "pass_fail": {
            "pass": fill_pass,
            "fail": fill_fail
        },
        "number_format": {
            "text": "General",
            "date": "yyyy-mm-dd",
            "datetime": "yyyy-mm-dd hh:mm",
            "number": get_decimal_format
        }
    }

    #region create data validation rules
    dv_employees = DataValidation(
        type = "list",
        showDropDown = False,
        formula1 = data["data_validation"]["employees"],
        allow_blank = True
    )
    dv_employees.error = "Your entry is not in the list."
    dv_employees.errorTitle = "Invalid Entry"
    dv_dispositions = DataValidation(
        type = "list",
        showDropDown = False,
        formula1 = data["data_validation"]["dispositions"],
        allow_blank = True
    )
    dv_dispositions.error = "Your entry is not in the list."
    dv_dispositions.errorTitle = "Invalid Entry"
    dv_material_types = DataValidation(
        type = "list",
        showDropDown = False,
        formula1 = data["data_validation"]["material_types"],
        allow_blank = True
    )
    dv_material_types.error = "Your entry is not in the list."
    dv_material_types.errorTitle = "Invalid Entry"
    dv_locations = DataValidation(
        type = "list",
        showDropDown = False,
        formula1 = data["data_validation"]["locations"],
        allow_blank = True
    )
    dv_locations.error = "Your entry is not in the list."
    dv_locations.errorTitle = "Invalid Entry"
    dv_suppliers = DataValidation(
        type = "list",
        showDropDown = False,
        formula1 = data["data_validation"]["suppliers"],
        allow_blank = True
    )
    dv_suppliers.error = "Your entry is not in the list."
    dv_suppliers.errorTitle = "Invalid Entry"
    dv_integer = DataValidation(
        type = "whole",
        operator = "greaterThanOrEqual",
        formula1 = 0,
        allow_blank = True
    )
    dv_integer.error = "Your entry must be a whole number."
    dv_integer.errorTitle = "Invalid Entry"
    dv_decimal = DataValidation(
        type = "decimal",
        allow_blank = True
    )
    dv_decimal.error = "Your entry must be a valid number."
    dv_decimal.errorTitle = "Invalid Entry"
    dv_gauges = DataValidation(
        type = "list",
        showDropDown = False,
        formula1 = data["data_validation"]["gauges"],
        allow_blank = True
    )
    dv_gauges.error = "Your entry is not in the list."
    dv_gauges.errorTitle = "Invalid Entry"
    #endregion

    # assemble data validation object
    data_validation_obj = {
        "employees": dv_employees,
        "dispositions": dv_dispositions,
        "locations": dv_locations,
        "material_types": dv_material_types,
        "suppliers": dv_suppliers,
        "integer": dv_integer,
        "decimal": dv_decimal,
        "gauges": dv_gauges
    }

    # create the correctly named worksheet
    ws = wb.active
    ws.title = data["metadata"]["type"]

    # add data validation attributes to the worksheet
    ws.add_data_validation(dv_employees)
    ws.add_data_validation(dv_dispositions)
    ws.add_data_validation(dv_integer)
    ws.add_data_validation(dv_locations)
    ws.add_data_validation(dv_material_types)
    ws.add_data_validation(dv_suppliers)
    ws.add_data_validation(dv_decimal)
    ws.add_data_validation(dv_gauges)

    # apply formats
    if data["metadata"]["type"] == "internal":
        format_metadata_block_internal(ws, data, format_obj, data_validation_obj)
        format_print_block(ws, data, format_obj)
        format_inspections_block_internal(ws, data, format_obj, data_validation_obj)
    elif data["metadata"]["type"] == "freight":
        format_metadata_block_freight(ws, data, format_obj, data_validation_obj)
        format_print_block(ws, data, format_obj)
        format_inspections_block_freight(ws, data, format_obj, data_validation_obj)
    else:
        wb.close()
        return {
            "status": "alert",
            "response": "A purchase order and receiver number or a job number must be selected."
        }

    # set protections
    wb.security = WorkbookProtection(workbookPassword="quality", lockStructure=True)
    ws.protection.sheet = True
    ws.protection.enable()

    return {
        "status": "ok",
        "response": None
    }

def format_metadata_block_internal(ws, data, format_obj, data_validation_obj):

    # define editable attribute
    is_editable = Protection(False, False)

    # extract data validation attributes
    dv_employees = data_validation_obj["employees"]
    dv_dispositions = data_validation_obj["dispositions"]
    dv_locations = data_validation_obj["locations"]
    dv_material_types = data_validation_obj["material_types"]
    dv_integer = data_validation_obj["integer"]

    #region merge cells
    ws.merge_cells(range_string = "B4:C4")
    ws.merge_cells(range_string = "B5:C5")
    ws.merge_cells(range_string = "B6:C6")
    ws.merge_cells(range_string = "D4:E4")
    ws.merge_cells(range_string = "D5:E5")
    ws.merge_cells(range_string = "D6:E6")
    ws.merge_cells(range_string = "G4:H4")
    ws.merge_cells(range_string = "G5:H5")
    ws.merge_cells(range_string = "G6:H6")
    ws.merge_cells(range_string = "I4:J4")
    ws.merge_cells(range_string = "I5:J5")
    ws.merge_cells(range_string = "I6:J6")
    ws.merge_cells(range_string = "L4:M4")
    ws.merge_cells(range_string = "L5:M5")
    ws.merge_cells(range_string = "N4:O4")
    ws.merge_cells(range_string = "N5:O5")
    ws.merge_cells(range_string = "Q4:R4")
    ws.merge_cells(range_string = "Q5:R5")
    ws.merge_cells(range_string = "Q6:R6")
    ws.merge_cells(range_string = "S4:T4")
    ws.merge_cells(range_string = "S5:T5")
    ws.merge_cells(range_string = "S6:T6")
    ws.merge_cells(range_string = "V4:W4")
    ws.merge_cells(range_string = "V5:W5")
    ws.merge_cells(range_string = "V6:W6")
    ws.merge_cells(range_string = "X4:Y4")
    ws.merge_cells(range_string = "X5:Y5")
    ws.merge_cells(range_string = "X6:Y6")
    #endregion

    #region set the data
    ws["D4"].value = data["item"]
    ws["D5"].value = data["drawing"]
    ws["D6"].value = data["revision"]
    ws["I4"].value = data["metadata"]["date"]
    ws["I5"].value = data["metadata"]["inspector"]
    ws["I6"].value = data["metadata"]["disposition"]
    ws["N4"].value = data["metadata"]["job_number"]
    ws["N5"].value = data["metadata"]["operator"]
    ws["S4"].value = data["metadata"]["qc_full_inspect_interval"]
    ws["S5"].value = data["metadata"]["released_quantity"]
    ws["S6"].value = data["metadata"]["completed_quantity"]
    ws["X4"].value = data["metadata"]["material_type"]
    ws["X5"].value = data["metadata"]["workcenter"]
    ws["X6"].value = data["metadata"]["production_rate"]
    #endregion

    #region set the headers
    ws["B4"].value = "Item Number"
    ws["B5"].value = "Drawing Number"
    ws["B6"].value = "Drawing Revision"
    ws["G4"].value = "Date"
    ws["G5"].value = "Inspector"
    ws["G6"].value = "Disposition"
    ws["L4"].value = "Job Number"
    ws["L5"].value = "Operator"
    ws["Q4"].value = "QC Full Inspect Interval"
    ws["Q5"].value = "Released Quantity"
    ws["Q6"].value = "Completed Quantity"
    ws["V4"].value = "Material Type"
    ws["V5"].value = "Work Center"
    ws["V6"].value = "Parts / Hour"
    #endregion

    #region format the data
    ws["D4"].fill = format_obj["fill"]["data"]
    ws["E4"].fill = format_obj["fill"]["data"]
    ws["D5"].fill = format_obj["fill"]["data"]
    ws["E5"].fill = format_obj["fill"]["data"]
    ws["D6"].fill = format_obj["fill"]["data"]
    ws["E6"].fill = format_obj["fill"]["data"]
    ws["I4"].fill = format_obj["fill"]["data"]
    ws["J4"].fill = format_obj["fill"]["data"]
    ws["I5"].fill = format_obj["fill"]["data"]
    ws["J5"].fill = format_obj["fill"]["data"]
    ws["I6"].fill = format_obj["fill"]["data"]
    ws["J6"].fill = format_obj["fill"]["data"]
    ws["N4"].fill = format_obj["fill"]["data"]
    ws["O4"].fill = format_obj["fill"]["data"]
    ws["N5"].fill = format_obj["fill"]["data"]
    ws["O5"].fill = format_obj["fill"]["data"]
    ws["S4"].fill = format_obj["fill"]["data"]
    ws["T4"].fill = format_obj["fill"]["data"]
    ws["S5"].fill = format_obj["fill"]["data"]
    ws["T5"].fill = format_obj["fill"]["data"]
    ws["S6"].fill = format_obj["fill"]["data"]
    ws["T6"].fill = format_obj["fill"]["data"]
    ws["X4"].fill = format_obj["fill"]["data"]
    ws["Y4"].fill = format_obj["fill"]["data"]
    ws["X5"].fill = format_obj["fill"]["data"]
    ws["Y5"].fill = format_obj["fill"]["data"]
    ws["X6"].fill = format_obj["fill"]["data"]
    ws["Y6"].fill = format_obj["fill"]["data"]
    ws["D4"].font = format_obj["font"]["data"]
    ws["E4"].font = format_obj["font"]["data"]
    ws["D5"].font = format_obj["font"]["data"]
    ws["E5"].font = format_obj["font"]["data"]
    ws["D6"].font = format_obj["font"]["data"]
    ws["E6"].font = format_obj["font"]["data"]
    ws["I4"].font = format_obj["font"]["data"]
    ws["J4"].font = format_obj["font"]["data"]
    ws["I5"].font = format_obj["font"]["data"]
    ws["J5"].font = format_obj["font"]["data"]
    ws["I6"].font = format_obj["font"]["data"]
    ws["J6"].font = format_obj["font"]["data"]
    ws["N4"].font = format_obj["font"]["data"]
    ws["O4"].font = format_obj["font"]["data"]
    ws["N5"].font = format_obj["font"]["data"]
    ws["O5"].font = format_obj["font"]["data"]
    ws["S4"].font = format_obj["font"]["data"]
    ws["T4"].font = format_obj["font"]["data"]
    ws["S5"].font = format_obj["font"]["data"]
    ws["T5"].font = format_obj["font"]["data"]
    ws["S6"].font = format_obj["font"]["data"]
    ws["T6"].font = format_obj["font"]["data"]
    ws["X4"].font = format_obj["font"]["data"]
    ws["Y4"].font = format_obj["font"]["data"]
    ws["X5"].font = format_obj["font"]["data"]
    ws["Y5"].font = format_obj["font"]["data"]
    ws["X6"].font = format_obj["font"]["data"]
    ws["Y6"].font = format_obj["font"]["data"]
    ws["D4"].border = format_obj["border"]["data"]
    ws["E4"].border = format_obj["border"]["data"]
    ws["D5"].border = format_obj["border"]["data"]
    ws["E5"].border = format_obj["border"]["data"]
    ws["D6"].border = format_obj["border"]["data"]
    ws["E6"].border = format_obj["border"]["data"]
    ws["I4"].border = format_obj["border"]["data"]
    ws["J4"].border = format_obj["border"]["data"]
    ws["I5"].border = format_obj["border"]["data"]
    ws["J5"].border = format_obj["border"]["data"]
    ws["I6"].border = format_obj["border"]["data"]
    ws["J6"].border = format_obj["border"]["data"]
    ws["N4"].border = format_obj["border"]["data"]
    ws["O4"].border = format_obj["border"]["data"]
    ws["N5"].border = format_obj["border"]["data"]
    ws["O5"].border = format_obj["border"]["data"]
    ws["S4"].border = format_obj["border"]["data"]
    ws["T4"].border = format_obj["border"]["data"]
    ws["S5"].border = format_obj["border"]["data"]
    ws["T5"].border = format_obj["border"]["data"]
    ws["S6"].border = format_obj["border"]["data"]
    ws["T6"].border = format_obj["border"]["data"]
    ws["X4"].border = format_obj["border"]["data"]
    ws["Y4"].border = format_obj["border"]["data"]
    ws["X5"].border = format_obj["border"]["data"]
    ws["Y5"].border = format_obj["border"]["data"]
    ws["X6"].border = format_obj["border"]["data"]
    ws["Y6"].border = format_obj["border"]["data"]
    ws["D4"].alignment = format_obj["alignment"]["data"]
    ws["E4"].alignment = format_obj["alignment"]["data"]
    ws["D5"].alignment = format_obj["alignment"]["data"]
    ws["E5"].alignment = format_obj["alignment"]["data"]
    ws["D6"].alignment = format_obj["alignment"]["data"]
    ws["E6"].alignment = format_obj["alignment"]["data"]
    ws["I4"].alignment = format_obj["alignment"]["data"]
    ws["J4"].alignment = format_obj["alignment"]["data"]
    ws["I5"].alignment = format_obj["alignment"]["data"]
    ws["J5"].alignment = format_obj["alignment"]["data"]
    ws["I6"].alignment = format_obj["alignment"]["data"]
    ws["J6"].alignment = format_obj["alignment"]["data"]
    ws["N4"].alignment = format_obj["alignment"]["data"]
    ws["O4"].alignment = format_obj["alignment"]["data"]
    ws["N5"].alignment = format_obj["alignment"]["data"]
    ws["O5"].alignment = format_obj["alignment"]["data"]
    ws["S4"].alignment = format_obj["alignment"]["data"]
    ws["T4"].alignment = format_obj["alignment"]["data"]
    ws["S5"].alignment = format_obj["alignment"]["data"]
    ws["T5"].alignment = format_obj["alignment"]["data"]
    ws["S6"].alignment = format_obj["alignment"]["data"]
    ws["T6"].alignment = format_obj["alignment"]["data"]
    ws["X4"].alignment = format_obj["alignment"]["data"]
    ws["Y4"].alignment = format_obj["alignment"]["data"]
    ws["X5"].alignment = format_obj["alignment"]["data"]
    ws["Y5"].alignment = format_obj["alignment"]["data"]
    ws["X6"].alignment = format_obj["alignment"]["data"]
    ws["Y6"].alignment = format_obj["alignment"]["data"]
    ws["N4"].number_format = format_obj["number_format"]["text"]
    ws["O4"].number_format = format_obj["number_format"]["text"]
    ws["N5"].number_format = format_obj["number_format"]["text"]
    ws["O5"].number_format = format_obj["number_format"]["text"]
    ws["S4"].number_format = format_obj["number_format"]["number"](0)
    ws["T4"].number_format = format_obj["number_format"]["number"](0)
    ws["S5"].number_format = format_obj["number_format"]["number"](0)
    ws["T5"].number_format = format_obj["number_format"]["number"](0)
    ws["S6"].number_format = format_obj["number_format"]["number"](0)
    ws["T6"].number_format = format_obj["number_format"]["number"](0)
    ws["X4"].number_format = format_obj["number_format"]["text"]
    ws["Y4"].number_format = format_obj["number_format"]["text"]
    ws["X5"].number_format = format_obj["number_format"]["text"]
    ws["Y5"].number_format = format_obj["number_format"]["text"]
    ws["X6"].number_format = format_obj["number_format"]["number"](2)
    ws["Y6"].number_format = format_obj["number_format"]["number"](2)
    #endregion

    #region format the headers
    ws["B4"].fill = format_obj["fill"]["header"]
    ws["C4"].fill = format_obj["fill"]["header"]
    ws["B5"].fill = format_obj["fill"]["header"]
    ws["C5"].fill = format_obj["fill"]["header"]
    ws["B6"].fill = format_obj["fill"]["header"]
    ws["C6"].fill = format_obj["fill"]["header"]
    ws["G4"].fill = format_obj["fill"]["header"]
    ws["H4"].fill = format_obj["fill"]["header"]
    ws["G5"].fill = format_obj["fill"]["header"]
    ws["H5"].fill = format_obj["fill"]["header"]
    ws["G6"].fill = format_obj["fill"]["header"]
    ws["H6"].fill = format_obj["fill"]["header"]
    ws["L4"].fill = format_obj["fill"]["header"]
    ws["M4"].fill = format_obj["fill"]["header"]
    ws["L5"].fill = format_obj["fill"]["header"]
    ws["M5"].fill = format_obj["fill"]["header"]
    ws["Q4"].fill = format_obj["fill"]["header"]
    ws["R4"].fill = format_obj["fill"]["header"]
    ws["Q5"].fill = format_obj["fill"]["header"]
    ws["R5"].fill = format_obj["fill"]["header"]
    ws["Q6"].fill = format_obj["fill"]["header"]
    ws["R6"].fill = format_obj["fill"]["header"]
    ws["V4"].fill = format_obj["fill"]["header"]
    ws["W4"].fill = format_obj["fill"]["header"]
    ws["V5"].fill = format_obj["fill"]["header"]
    ws["W5"].fill = format_obj["fill"]["header"]
    ws["V6"].fill = format_obj["fill"]["header"]
    ws["W6"].fill = format_obj["fill"]["header"]
    ws["B4"].font = format_obj["font"]["header"]
    ws["C4"].font = format_obj["font"]["header"]
    ws["B5"].font = format_obj["font"]["header"]
    ws["C5"].font = format_obj["font"]["header"]
    ws["B6"].font = format_obj["font"]["header"]
    ws["C6"].font = format_obj["font"]["header"]
    ws["G4"].font = format_obj["font"]["header"]
    ws["H4"].font = format_obj["font"]["header"]
    ws["G5"].font = format_obj["font"]["header"]
    ws["H5"].font = format_obj["font"]["header"]
    ws["G6"].font = format_obj["font"]["header"]
    ws["H6"].font = format_obj["font"]["header"]
    ws["L4"].font = format_obj["font"]["header"]
    ws["M4"].font = format_obj["font"]["header"]
    ws["L5"].font = format_obj["font"]["header"]
    ws["M5"].font = format_obj["font"]["header"]
    ws["Q4"].font = format_obj["font"]["header"]
    ws["R4"].font = format_obj["font"]["header"]
    ws["Q5"].font = format_obj["font"]["header"]
    ws["R5"].font = format_obj["font"]["header"]
    ws["Q6"].font = format_obj["font"]["header"]
    ws["R6"].font = format_obj["font"]["header"]
    ws["V4"].font = format_obj["font"]["header"]
    ws["W4"].font = format_obj["font"]["header"]
    ws["V5"].font = format_obj["font"]["header"]
    ws["W5"].font = format_obj["font"]["header"]
    ws["V6"].font = format_obj["font"]["header"]
    ws["W6"].font = format_obj["font"]["header"]
    ws["B4"].border = format_obj["border"]["header"]
    ws["C4"].border = format_obj["border"]["header"]
    ws["B5"].border = format_obj["border"]["header"]
    ws["C5"].border = format_obj["border"]["header"]
    ws["B6"].border = format_obj["border"]["header"]
    ws["C6"].border = format_obj["border"]["header"]
    ws["G4"].border = format_obj["border"]["header"]
    ws["H4"].border = format_obj["border"]["header"]
    ws["G5"].border = format_obj["border"]["header"]
    ws["H5"].border = format_obj["border"]["header"]
    ws["G6"].border = format_obj["border"]["header"]
    ws["H6"].border = format_obj["border"]["header"]
    ws["L4"].border = format_obj["border"]["header"]
    ws["M4"].border = format_obj["border"]["header"]
    ws["L5"].border = format_obj["border"]["header"]
    ws["M5"].border = format_obj["border"]["header"]
    ws["Q4"].border = format_obj["border"]["header"]
    ws["R4"].border = format_obj["border"]["header"]
    ws["Q5"].border = format_obj["border"]["header"]
    ws["R5"].border = format_obj["border"]["header"]
    ws["Q6"].border = format_obj["border"]["header"]
    ws["R6"].border = format_obj["border"]["header"]
    ws["V4"].border = format_obj["border"]["header"]
    ws["W4"].border = format_obj["border"]["header"]
    ws["V5"].border = format_obj["border"]["header"]
    ws["W5"].border = format_obj["border"]["header"]
    ws["V6"].border = format_obj["border"]["header"]
    ws["W6"].border = format_obj["border"]["header"]
    ws["B4"].alignment = format_obj["alignment"]["header"]
    ws["C4"].alignment = format_obj["alignment"]["header"]
    ws["B5"].alignment = format_obj["alignment"]["header"]
    ws["C5"].alignment = format_obj["alignment"]["header"]
    ws["B6"].alignment = format_obj["alignment"]["header"]
    ws["C6"].alignment = format_obj["alignment"]["header"]
    ws["G4"].alignment = format_obj["alignment"]["header"]
    ws["H4"].alignment = format_obj["alignment"]["header"]
    ws["G5"].alignment = format_obj["alignment"]["header"]
    ws["H5"].alignment = format_obj["alignment"]["header"]
    ws["G6"].alignment = format_obj["alignment"]["header"]
    ws["H6"].alignment = format_obj["alignment"]["header"]
    ws["L4"].alignment = format_obj["alignment"]["header"]
    ws["M4"].alignment = format_obj["alignment"]["header"]
    ws["L5"].alignment = format_obj["alignment"]["header"]
    ws["M5"].alignment = format_obj["alignment"]["header"]
    ws["Q4"].alignment = format_obj["alignment"]["header"]
    ws["R4"].alignment = format_obj["alignment"]["header"]
    ws["Q5"].alignment = format_obj["alignment"]["header"]
    ws["R5"].alignment = format_obj["alignment"]["header"]
    ws["Q6"].alignment = format_obj["alignment"]["header"]
    ws["R6"].alignment = format_obj["alignment"]["header"]
    ws["V4"].alignment = format_obj["alignment"]["header"]
    ws["W4"].alignment = format_obj["alignment"]["header"]
    ws["V5"].alignment = format_obj["alignment"]["header"]
    ws["W5"].alignment = format_obj["alignment"]["header"]
    ws["V6"].alignment = format_obj["alignment"]["header"]
    ws["W6"].alignment = format_obj["alignment"]["header"]
    #endregion

    #region set editable states
    ws["I5"].protection = is_editable
    ws["I6"].protection = is_editable
    ws["N5"].protection = is_editable
    ws["S4"].protection = is_editable
    ws["S5"].protection = is_editable
    ws["S6"].protection = is_editable
    ws["X4"].protection = is_editable
    ws["X5"].protection = is_editable
    ws["X6"].protection = is_editable
    #endregion

    #region apply data validation
    dv_employees.add("I5:J5")
    dv_dispositions.add("I6:I6")
    dv_employees.add("N5:O5")
    dv_integer.add("S4:T6")
    dv_integer.add("X6:Y6")
    dv_material_types.add("X4:Y4")
    dv_locations.add("X5:Y5")
    #endregion

def format_metadata_block_freight(ws, data, format_obj, data_validation_obj):

    # define editable attribute
    is_editable = Protection(False, False)

    # extract data validation attributes
    dv_employees = data_validation_obj["employees"]
    dv_dispositions = data_validation_obj["dispositions"]
    dv_suppliers = data_validation_obj["suppliers"]
    dv_integer = data_validation_obj["integer"]

    #region merge cells
    ws.merge_cells(range_string = "B4:C4")
    ws.merge_cells(range_string = "B5:C5")
    ws.merge_cells(range_string = "B6:C6")
    ws.merge_cells(range_string = "D4:E4")
    ws.merge_cells(range_string = "D5:E5")
    ws.merge_cells(range_string = "D6:E6")
    ws.merge_cells(range_string = "G4:H4")
    ws.merge_cells(range_string = "G5:H5")
    ws.merge_cells(range_string = "G6:H6")
    ws.merge_cells(range_string = "I4:J4")
    ws.merge_cells(range_string = "I5:J5")
    ws.merge_cells(range_string = "I6:J6")
    ws.merge_cells(range_string = "L4:M4")
    ws.merge_cells(range_string = "L5:M5")
    ws.merge_cells(range_string = "L6:M6")
    ws.merge_cells(range_string = "N4:O4")
    ws.merge_cells(range_string = "N5:O5")
    ws.merge_cells(range_string = "N6:O6")
    ws.merge_cells(range_string = "Q4:R4")
    ws.merge_cells(range_string = "S4:T4")
    #endregion

    #region set the data
    ws["D4"].value = data["item"]
    ws["D5"].value = data["drawing"]
    ws["D6"].value = data["revision"]
    ws["I4"].value = data["metadata"]["date"]
    ws["I5"].value = data["metadata"]["inspector"]
    ws["I6"].value = data["metadata"]["disposition"]
    ws["N4"].value = data["metadata"]["purchase_order"]
    ws["N5"].value = data["metadata"]["receiver_number"]
    ws["N6"].value = data["metadata"]["supplier"]
    ws["S4"].value = data["metadata"]["received_quantity"]
    #endregion

    #region set the headers
    ws["B4"].value = "Item Number"
    ws["B5"].value = "Drawing Number"
    ws["B6"].value = "Drawing Revision"
    ws["G4"].value = "Date"
    ws["G5"].value = "Inspector"
    ws["G6"].value = "Disposition"
    ws["L4"].value = "Purchase Order"
    ws["L5"].value = "Receiver Number"
    ws["L6"].value = "Supplier"
    ws["Q4"].value = "Received Quantity"
    #endregion

    #region format the data
    ws["D4"].fill = format_obj["fill"]["data"]
    ws["E4"].fill = format_obj["fill"]["data"]
    ws["D5"].fill = format_obj["fill"]["data"]
    ws["E5"].fill = format_obj["fill"]["data"]
    ws["D6"].fill = format_obj["fill"]["data"]
    ws["E6"].fill = format_obj["fill"]["data"]
    ws["I4"].fill = format_obj["fill"]["data"]
    ws["J4"].fill = format_obj["fill"]["data"]
    ws["I5"].fill = format_obj["fill"]["data"]
    ws["J5"].fill = format_obj["fill"]["data"]
    ws["I6"].fill = format_obj["fill"]["data"]
    ws["J6"].fill = format_obj["fill"]["data"]
    ws["N4"].fill = format_obj["fill"]["data"]
    ws["O4"].fill = format_obj["fill"]["data"]
    ws["N5"].fill = format_obj["fill"]["data"]
    ws["O5"].fill = format_obj["fill"]["data"]
    ws["N6"].fill = format_obj["fill"]["data"]
    ws["O6"].fill = format_obj["fill"]["data"]
    ws["S4"].fill = format_obj["fill"]["data"]
    ws["T4"].fill = format_obj["fill"]["data"]
    ws["D4"].font = format_obj["font"]["data"]
    ws["E4"].font = format_obj["font"]["data"]
    ws["D5"].font = format_obj["font"]["data"]
    ws["E5"].font = format_obj["font"]["data"]
    ws["D6"].font = format_obj["font"]["data"]
    ws["E6"].font = format_obj["font"]["data"]
    ws["I4"].font = format_obj["font"]["data"]
    ws["J4"].font = format_obj["font"]["data"]
    ws["I5"].font = format_obj["font"]["data"]
    ws["J5"].font = format_obj["font"]["data"]
    ws["I6"].font = format_obj["font"]["data"]
    ws["J6"].font = format_obj["font"]["data"]
    ws["N4"].font = format_obj["font"]["data"]
    ws["O4"].font = format_obj["font"]["data"]
    ws["N5"].font = format_obj["font"]["data"]
    ws["O5"].font = format_obj["font"]["data"]
    ws["N6"].font = format_obj["font"]["data"]
    ws["O6"].font = format_obj["font"]["data"]
    ws["S4"].font = format_obj["font"]["data"]
    ws["T4"].font = format_obj["font"]["data"]
    ws["D4"].border = format_obj["border"]["data"]
    ws["E4"].border = format_obj["border"]["data"]
    ws["D5"].border = format_obj["border"]["data"]
    ws["E5"].border = format_obj["border"]["data"]
    ws["D6"].border = format_obj["border"]["data"]
    ws["E6"].border = format_obj["border"]["data"]
    ws["I4"].border = format_obj["border"]["data"]
    ws["J4"].border = format_obj["border"]["data"]
    ws["I5"].border = format_obj["border"]["data"]
    ws["J5"].border = format_obj["border"]["data"]
    ws["I6"].border = format_obj["border"]["data"]
    ws["J6"].border = format_obj["border"]["data"]
    ws["N4"].border = format_obj["border"]["data"]
    ws["O4"].border = format_obj["border"]["data"]
    ws["N5"].border = format_obj["border"]["data"]
    ws["O5"].border = format_obj["border"]["data"]
    ws["N6"].border = format_obj["border"]["data"]
    ws["O6"].border = format_obj["border"]["data"]
    ws["S4"].border = format_obj["border"]["data"]
    ws["T4"].border = format_obj["border"]["data"]
    ws["D4"].alignment = format_obj["alignment"]["data"]
    ws["E4"].alignment = format_obj["alignment"]["data"]
    ws["D5"].alignment = format_obj["alignment"]["data"]
    ws["E5"].alignment = format_obj["alignment"]["data"]
    ws["D6"].alignment = format_obj["alignment"]["data"]
    ws["E6"].alignment = format_obj["alignment"]["data"]
    ws["I4"].alignment = format_obj["alignment"]["data"]
    ws["J4"].alignment = format_obj["alignment"]["data"]
    ws["I5"].alignment = format_obj["alignment"]["data"]
    ws["J5"].alignment = format_obj["alignment"]["data"]
    ws["I6"].alignment = format_obj["alignment"]["data"]
    ws["J6"].alignment = format_obj["alignment"]["data"]
    ws["N4"].alignment = format_obj["alignment"]["data"]
    ws["O4"].alignment = format_obj["alignment"]["data"]
    ws["N5"].alignment = format_obj["alignment"]["data"]
    ws["O5"].alignment = format_obj["alignment"]["data"]
    ws["N6"].alignment = format_obj["alignment"]["data"]
    ws["O6"].alignment = format_obj["alignment"]["data"]
    ws["S4"].alignment = format_obj["alignment"]["data"]
    ws["T4"].alignment = format_obj["alignment"]["data"]
    ws["N4"].number_format = format_obj["number_format"]["text"]
    ws["O4"].number_format = format_obj["number_format"]["text"]
    ws["N5"].number_format = format_obj["number_format"]["text"]
    ws["O5"].number_format = format_obj["number_format"]["text"]
    ws["N6"].number_format = format_obj["number_format"]["text"]
    ws["O6"].number_format = format_obj["number_format"]["text"]
    ws["S4"].number_format = format_obj["number_format"]["number"](0)
    ws["T4"].number_format = format_obj["number_format"]["number"](0)
    #endregion

    #region format the headers
    ws["B4"].fill = format_obj["fill"]["header"]
    ws["C4"].fill = format_obj["fill"]["header"]
    ws["B5"].fill = format_obj["fill"]["header"]
    ws["C5"].fill = format_obj["fill"]["header"]
    ws["B6"].fill = format_obj["fill"]["header"]
    ws["C6"].fill = format_obj["fill"]["header"]
    ws["G4"].fill = format_obj["fill"]["header"]
    ws["H4"].fill = format_obj["fill"]["header"]
    ws["G5"].fill = format_obj["fill"]["header"]
    ws["H5"].fill = format_obj["fill"]["header"]
    ws["G6"].fill = format_obj["fill"]["header"]
    ws["H6"].fill = format_obj["fill"]["header"]
    ws["L4"].fill = format_obj["fill"]["header"]
    ws["M4"].fill = format_obj["fill"]["header"]
    ws["L5"].fill = format_obj["fill"]["header"]
    ws["M5"].fill = format_obj["fill"]["header"]
    ws["L6"].fill = format_obj["fill"]["header"]
    ws["M6"].fill = format_obj["fill"]["header"]
    ws["Q4"].fill = format_obj["fill"]["header"]
    ws["R4"].fill = format_obj["fill"]["header"]
    ws["B4"].font = format_obj["font"]["header"]
    ws["C4"].font = format_obj["font"]["header"]
    ws["B5"].font = format_obj["font"]["header"]
    ws["C5"].font = format_obj["font"]["header"]
    ws["B6"].font = format_obj["font"]["header"]
    ws["C6"].font = format_obj["font"]["header"]
    ws["G4"].font = format_obj["font"]["header"]
    ws["H4"].font = format_obj["font"]["header"]
    ws["G5"].font = format_obj["font"]["header"]
    ws["H5"].font = format_obj["font"]["header"]
    ws["G6"].font = format_obj["font"]["header"]
    ws["H6"].font = format_obj["font"]["header"]
    ws["L4"].font = format_obj["font"]["header"]
    ws["M4"].font = format_obj["font"]["header"]
    ws["L5"].font = format_obj["font"]["header"]
    ws["M5"].font = format_obj["font"]["header"]
    ws["L6"].font = format_obj["font"]["header"]
    ws["M6"].font = format_obj["font"]["header"]
    ws["Q4"].font = format_obj["font"]["header"]
    ws["R4"].font = format_obj["font"]["header"]
    ws["B4"].border = format_obj["border"]["header"]
    ws["C4"].border = format_obj["border"]["header"]
    ws["B5"].border = format_obj["border"]["header"]
    ws["C5"].border = format_obj["border"]["header"]
    ws["B6"].border = format_obj["border"]["header"]
    ws["C6"].border = format_obj["border"]["header"]
    ws["G4"].border = format_obj["border"]["header"]
    ws["H4"].border = format_obj["border"]["header"]
    ws["G5"].border = format_obj["border"]["header"]
    ws["H5"].border = format_obj["border"]["header"]
    ws["G6"].border = format_obj["border"]["header"]
    ws["H6"].border = format_obj["border"]["header"]
    ws["L4"].border = format_obj["border"]["header"]
    ws["M4"].border = format_obj["border"]["header"]
    ws["L5"].border = format_obj["border"]["header"]
    ws["M5"].border = format_obj["border"]["header"]
    ws["L6"].border = format_obj["border"]["header"]
    ws["M6"].border = format_obj["border"]["header"]
    ws["Q4"].border = format_obj["border"]["header"]
    ws["R4"].border = format_obj["border"]["header"]
    ws["B4"].alignment = format_obj["alignment"]["header"]
    ws["C4"].alignment = format_obj["alignment"]["header"]
    ws["B5"].alignment = format_obj["alignment"]["header"]
    ws["C5"].alignment = format_obj["alignment"]["header"]
    ws["B6"].alignment = format_obj["alignment"]["header"]
    ws["C6"].alignment = format_obj["alignment"]["header"]
    ws["G4"].alignment = format_obj["alignment"]["header"]
    ws["H4"].alignment = format_obj["alignment"]["header"]
    ws["G5"].alignment = format_obj["alignment"]["header"]
    ws["H5"].alignment = format_obj["alignment"]["header"]
    ws["G6"].alignment = format_obj["alignment"]["header"]
    ws["H6"].alignment = format_obj["alignment"]["header"]
    ws["L4"].alignment = format_obj["alignment"]["header"]
    ws["M4"].alignment = format_obj["alignment"]["header"]
    ws["L5"].alignment = format_obj["alignment"]["header"]
    ws["M5"].alignment = format_obj["alignment"]["header"]
    ws["L6"].alignment = format_obj["alignment"]["header"]
    ws["M6"].alignment = format_obj["alignment"]["header"]
    ws["Q4"].alignment = format_obj["alignment"]["header"]
    ws["R4"].alignment = format_obj["alignment"]["header"]
    #endregion

    #region set editable states
    ws["I5"].protection = is_editable
    ws["I6"].protection = is_editable
    ws["N6"].protection = is_editable
    ws["S4"].protection = is_editable
    #endregion

    #region apply the data validation
    dv_employees.add("I5:J5")
    dv_dispositions.add("I6:I6")
    dv_suppliers.add("N6:O6")
    dv_integer.add("S4:T4")
    #endregion

def format_print_block(ws, data, format_obj):

    # set the data
    feature_count = len(data["print"])
    for i in range(feature_count):

        # merge and format
        ws.cell(8, 7 + i * 2).fill = format_obj["fill"]["data"]
        ws.cell(8, 8 + i * 2).fill = format_obj["fill"]["data"]
        ws.cell(8, 7 + i * 2).font = format_obj["font"]["data"]
        ws.cell(8, 8 + i * 2).font = format_obj["font"]["data"]
        ws.cell(8, 7 + i * 2).border = format_obj["border"]["data"]
        ws.cell(8, 8 + i * 2).border = format_obj["border"]["data"]
        ws.cell(8, 7 + i * 2).alignment = format_obj["alignment"]["data"]
        ws.cell(8, 8 + i * 2).alignment = format_obj["alignment"]["data"]
        for j in range(9, 15):
            ws.merge_cells(start_row = j, end_row = j, start_column = 7 + i * 2, end_column = 8 + i * 2)
            ws.cell(j, 7 + i * 2).fill = format_obj["fill"]["data"]
            ws.cell(j, 8 + i * 2).fill = format_obj["fill"]["data"]
            ws.cell(j, 7 + i * 2).font = format_obj["font"]["data"]
            ws.cell(j, 8 + i * 2).font = format_obj["font"]["data"]
            ws.cell(j, 7 + i * 2).border = format_obj["border"]["data"]
            ws.cell(j, 8 + i * 2).border = format_obj["border"]["data"]
            ws.cell(j, 7 + i * 2).alignment = format_obj["alignment"]["data"]
            ws.cell(j, 8 + i * 2).alignment = format_obj["alignment"]["data"]

        # set the data
        ws.cell(8, 7 + i * 2).value = data["print"][i]["id"]
        ws.cell(8, 8 + i * 2).value = data["print"][i]["name"]
        ws.cell(9, 7 + i * 2).value = data["print"][i]["operation_type"]
        ws.cell(10, 7 + i * 2).value = data["print"][i]["dimension_type"]
        ws.cell(11, 7 + i * 2).value = data["print"][i]["usl"]
        ws.cell(12, 7 + i * 2).value = data["print"][i]["lsl"]
        ws.cell(13, 7 + i * 2).value = data["print"][i]["gauge_type"]
        ws.cell(14, 7 + i * 2).value = data["print"][i]["frequency_type"]

        # apply number formats
        ws.cell(8, 7 + i * 2).number_format = format_obj["number_format"]["text"]
        ws.cell(9, 7 + i * 2).number_format = format_obj["number_format"]["text"]
        ws.cell(10, 7 + i * 2).number_format = format_obj["number_format"]["text"]
        ws.cell(11, 7 + i * 2).number_format = format_obj["number_format"]["number"](data["print"][i]["precision"])
        ws.cell(12, 7 + i * 2).number_format = format_obj["number_format"]["number"](data["print"][i]["precision"])
        ws.cell(13, 7 + i * 2).number_format = format_obj["number_format"]["text"]
        ws.cell(14, 7 + i * 2).number_format = format_obj["number_format"]["text"]

    # set the header
    ws.merge_cells(range_string = "E8:F8")
    ws.merge_cells(range_string = "E9:F9")
    ws.merge_cells(range_string = "E10:F10")
    ws.merge_cells(range_string = "E11:F11")
    ws.merge_cells(range_string = "E12:F12")
    ws.merge_cells(range_string = "E13:F13")
    ws.merge_cells(range_string = "E14:F14")
    ws["E8"].value = "Feature"
    ws["E9"].value = "Operation"
    ws["E10"].value = "Dimension"
    ws["E11"].value = "Upper"
    ws["E12"].value = "Lower"
    ws["E13"].value = "Gauge Type"
    ws["E14"].value = "Check Frequency"
    for i in range(8, 15):
        ws.cell(i, 5).fill = format_obj["fill"]["header"]
        ws.cell(i, 6).fill = format_obj["fill"]["header"]
        ws.cell(i, 5).font = format_obj["font"]["header"]
        ws.cell(i, 6).font = format_obj["font"]["header"]
        ws.cell(i, 5).border = format_obj["border"]["header"]
        ws.cell(i, 6).border = format_obj["border"]["header"]
        ws.cell(i, 5).alignment = format_obj["alignment"]["header"]
        ws.cell(i, 6).alignment = format_obj["alignment"]["header"]

def format_inspections_block_internal(ws, data, format_obj, data_validation_obj):

    # define editable attribute
    is_editable = Protection(False, False)

    # extract data validation attributes
    dv_employees = data_validation_obj["employees"]
    dv_integer = data_validation_obj["integer"]
    dv_gauges = data_validation_obj["gauges"]

    # get the correct feature indexing
    feature_index = []
    for j in range(len(data["inspections"][0]["features"])):
        print_feature_id = ws.cell(8, 7 + j * 2).value
        for i in range(len(data["inspections"][0]["features"])):
            if data["inspections"][i]["features"][i]["print_feature_id"] == print_feature_id:
                feature_index.append(i)

    # create the inspections list
    for i in range(len(data["inspections"])):

        # data
        for j in range(len(data["inspections"][i]["features"])):

            # set data
            ws.cell(17 + i, 7 + j * 2).value = data["inspections"][i]["features"][feature_index[j]]["measured"]
            ws.cell(17 + i, 8 + j * 2).value = data["inspections"][i]["features"][feature_index[j]]["gauge"]

            # set formatting
            ws.cell(17 + i, 7 + j * 2).fill = format_obj["fill"]["data"]
            ws.cell(17 + i, 8 + j * 2).fill = format_obj["fill"]["data"]
            ws.cell(17 + i, 7 + j * 2).font = format_obj["font"]["data"]
            ws.cell(17 + i, 8 + j * 2).font = format_obj["font"]["data"]
            ws.cell(17 + i, 7 + j * 2).border = format_obj["border"]["data"]
            ws.cell(17 + i, 8 + j * 2).border = format_obj["border"]["data"]
            ws.cell(17 + i, 7 + j * 2).alignment = format_obj["alignment"]["data"]
            ws.cell(17 + i, 8 + j * 2).alignment = format_obj["alignment"]["data"]
            ws.cell(17 + i, 7 + j * 2).number_format = format_obj["number_format"]["number"](data["print"][j]["precision"])
            ws.cell(17 + i, 8 + j * 2).number_format = format_obj["number_format"]["text"]

            # set editable state
            ws.cell(17 + i, 7 + j * 2).protection = is_editable
            ws.cell(17 + i, 8 + j * 2).protection = is_editable

            # set data validation
            dv_gauges.add(ws.cell(17 + i, 8 + j * 2).coordinate)

            # set conditional formatting
            column = ws.cell(17 + i, 7 + j * 2).column_letter
            usl_address = f"${column}${11}"
            lsl_address = f"${column}${12}"
            measure_address = f"{column}{17 + i}"
            blnk_formula = f"=ISBLANK({measure_address})"
            pass_formula = f"=AND({measure_address}<={usl_address},{measure_address}>={lsl_address})"
            fail_formula = f"=OR({measure_address}>{usl_address},{measure_address}<{lsl_address})"
            blnk_rule = FormulaRule(
                formula = [blnk_formula],
                fill = format_obj["fill"]["data"],
                stopIfTrue = True
            )
            pass_rule = FormulaRule(
                formula = [pass_formula],
                fill = format_obj["pass_fail"]["pass"],
                stopIfTrue = True
            )
            fail_rule = FormulaRule(
                formula = [fail_formula],
                fill = format_obj["pass_fail"]["fail"],
                stopIfTrue = True
            )
            ws.conditional_formatting.add(
                measure_address,
                blnk_rule
            )
            ws.conditional_formatting.add(
                measure_address,
                pass_rule
            )
            ws.conditional_formatting.add(
                measure_address,
                fail_rule
            )

    row = 17 + len(data["inspections"])
    for j in range(len(data["inspections"][0]["features"])):

        # set data
        ws.cell(row, 7 + j * 2).value = ""
        ws.cell(row, 8 + j * 2).value = ""

        # set format
        ws.cell(row, 7 + j * 2).fill = format_obj["fill"]["data"]
        ws.cell(row, 8 + j * 2).fill = format_obj["fill"]["data"]
        ws.cell(row, 7 + j * 2).font = format_obj["font"]["data"]
        ws.cell(row, 8 + j * 2).font = format_obj["font"]["data"]
        ws.cell(row, 7 + j * 2).border = format_obj["border"]["data"]
        ws.cell(row, 8 + j * 2).border = format_obj["border"]["data"]
        ws.cell(row, 7 + j * 2).alignment = format_obj["alignment"]["data"]
        ws.cell(row, 8 + j * 2).alignment = format_obj["alignment"]["data"]
        ws.cell(row, 7 + j * 2).number_format = format_obj["number_format"]["number"](data["print"][j]["precision"])
        ws.cell(row, 8 + j * 2).number_format = format_obj["number_format"]["text"]

        # set editable state
        ws.cell(row, 7 + j * 2).protection = is_editable
        ws.cell(row, 8 + j * 2).protection = is_editable

        # set data validation
        dv_gauges.add(ws.cell(row, 8 + j * 2).coordinate)

        # set conditional formatting
        column = ws.cell(row, 7 + j * 2).column_letter
        usl_address = f"${column}${11}"
        lsl_address = f"${column}${12}"
        measure_address = f"{column}{row}"
        blnk_formula = f"=ISBLANK({measure_address})"
        pass_formula = f"=AND({measure_address}<={usl_address},{measure_address}>={lsl_address})"
        fail_formula = f"=OR({measure_address}>{usl_address},{measure_address}<{lsl_address})"
        blnk_rule = FormulaRule(
            formula = [blnk_formula],
            fill = format_obj["fill"]["data"],
            stopIfTrue = True
        )
        pass_rule = FormulaRule(
            formula = [pass_formula],
            fill = format_obj["pass_fail"]["pass"],
            stopIfTrue = True
        )
        fail_rule = FormulaRule(
            formula = [fail_formula],
            fill = format_obj["pass_fail"]["fail"],
            stopIfTrue = True
        )
        ws.conditional_formatting.add(
            measure_address,
            blnk_rule
        )
        ws.conditional_formatting.add(
            measure_address,
            pass_rule
        )
        ws.conditional_formatting.add(
            measure_address,
            fail_rule
        )
    ws.merge_cells(start_row = row, end_row = row, start_column = 2, end_column = 3)
    ws.merge_cells(start_row = row, end_row = row, start_column = 4, end_column = 5)
    ws.cell(row, 2).fill = format_obj["fill"]["data"]
    ws.cell(row, 3).fill = format_obj["fill"]["data"]
    ws.cell(row, 4).fill = format_obj["fill"]["data"]
    ws.cell(row, 5).fill = format_obj["fill"]["data"]
    ws.cell(row, 6).fill = format_obj["fill"]["data"]
    ws.cell(row, 2).font = format_obj["font"]["data"]
    ws.cell(row, 3).font = format_obj["font"]["data"]
    ws.cell(row, 4).font = format_obj["font"]["data"]
    ws.cell(row, 5).font = format_obj["font"]["data"]
    ws.cell(row, 6).font = format_obj["font"]["data"]
    ws.cell(row, 2).border = format_obj["border"]["data"]
    ws.cell(row, 3).border = format_obj["border"]["data"]
    ws.cell(row, 4).border = format_obj["border"]["data"]
    ws.cell(row, 5).border = format_obj["border"]["data"]
    ws.cell(row, 6).border = format_obj["border"]["data"]
    ws.cell(row, 2).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 3).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 4).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 5).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 6).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 2).number_format = format_obj["number_format"]["text"]
    ws.cell(row, 3).number_format = format_obj["number_format"]["text"]
    ws.cell(row, 4).number_format = format_obj["number_format"]["datetime"]
    ws.cell(row, 5).number_format = format_obj["number_format"]["datetime"]
    ws.cell(row, 6).number_format = format_obj["number_format"]["number"](0)
    ws.cell(row, 2).protection = is_editable
    ws.cell(row, 3).protection = is_editable
    ws.cell(row, 6).protection = is_editable
    dv_employees.add(f"{ws.cell(row, 2).coordinate}:{ws.cell(row, 3).coordinate}")

    # merge
    ws.merge_cells(range_string = "B16:C16")
    ws.merge_cells(range_string = "D16:E16")

    # set contents
    ws["B16"].value = "Inspector"
    ws["D16"].value = "Date/Time"
    ws["F16"].value = "Part Count"

    # set formatting
    ws["B16"].fill = format_obj["fill"]["header"]
    ws["C16"].fill = format_obj["fill"]["header"]
    ws["D16"].fill = format_obj["fill"]["header"]
    ws["E16"].fill = format_obj["fill"]["header"]
    ws["F16"].fill = format_obj["fill"]["header"]
    ws["B16"].font = format_obj["font"]["header"]
    ws["C16"].font = format_obj["font"]["header"]
    ws["D16"].font = format_obj["font"]["header"]
    ws["E16"].font = format_obj["font"]["header"]
    ws["F16"].font = format_obj["font"]["header"]
    ws["B16"].border = format_obj["border"]["header"]
    ws["C16"].border = format_obj["border"]["header"]
    ws["D16"].border = format_obj["border"]["header"]
    ws["E16"].border = format_obj["border"]["header"]
    ws["f16"].border = format_obj["border"]["header"]
    ws["B16"].alignment = format_obj["alignment"]["header"]
    ws["C16"].alignment = format_obj["alignment"]["header"]
    ws["D16"].alignment = format_obj["alignment"]["header"]
    ws["E16"].alignment = format_obj["alignment"]["header"]
    ws["F16"].alignment = format_obj["alignment"]["header"]

    # dynamic headers
    for i in range(len(data["inspections"])):

        # merge
        ws.merge_cells(start_row = 17 + i, end_row = 17 + i, start_column = 2, end_column = 3)
        ws.merge_cells(start_row = 17 + i, end_row = 17 + i, start_column = 4, end_column = 5)

        # set contents
        ws.cell(17 + i, 2).value = data["inspections"][i]["employee_name"]
        ws.cell(17 + i, 4).value = data["inspections"][i]["datetime_measured"]
        ws.cell(17 + i, 6).value = data["inspections"][i]["part_index"]

        # set format
        ws.cell(17 + i, 2).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 3).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 4).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 5).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 6).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 2).font = format_obj["font"]["data"]
        ws.cell(17 + i, 3).font = format_obj["font"]["data"]
        ws.cell(17 + i, 4).font = format_obj["font"]["data"]
        ws.cell(17 + i, 5).font = format_obj["font"]["data"]
        ws.cell(17 + i, 6).font = format_obj["font"]["data"]
        ws.cell(17 + i, 2).border = format_obj["border"]["data"]
        ws.cell(17 + i, 3).border = format_obj["border"]["data"]
        ws.cell(17 + i, 4).border = format_obj["border"]["data"]
        ws.cell(17 + i, 5).border = format_obj["border"]["data"]
        ws.cell(17 + i, 6).border = format_obj["border"]["data"]
        ws.cell(17 + i, 2).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 3).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 4).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 5).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 6).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 2).number_format = format_obj["number_format"]["text"]
        ws.cell(17 + i, 3).number_format = format_obj["number_format"]["text"]
        ws.cell(17 + i, 4).number_format = format_obj["number_format"]["datetime"]
        ws.cell(17 + i, 5).number_format = format_obj["number_format"]["datetime"]
        ws.cell(17 + i, 6).number_format = format_obj["number_format"]["number"](0)

        # set editable state
        ws.cell(17 + i, 2).protection = is_editable
        ws.cell(17 + i, 3).protection = is_editable
        ws.cell(17 + i, 6).protection = is_editable

        # set data validation
        dv_employees.add(f"{ws.cell(17 + i, 2).coordinate}:{ws.cell(17 + i, 3).coordinate}")
        dv_integer.add(ws.cell(17 + i, 6).coordinate)

    for j in range(len(data["inspections"][0]["features"])):

        # merge
        ws.merge_cells(start_row = 16, end_row = 16, start_column = 7 + j * 2, end_column = 8 + j * 2)

        # set contents
        ws.cell(16, 7 + j * 2).value = j + 1

        # set format
        ws.cell(16, 7 + j * 2).fill = format_obj["fill"]["header"]
        ws.cell(16, 8 + j * 2).fill = format_obj["fill"]["header"]
        ws.cell(16, 7 + j * 2).font = format_obj["font"]["header"]
        ws.cell(16, 8 + j * 2).font = format_obj["font"]["header"]
        ws.cell(16, 7 + j * 2).border = format_obj["border"]["header"]
        ws.cell(16, 8 + j * 2).border = format_obj["border"]["header"]
        ws.cell(16, 7 + j * 2).alignment = Alignment(horizontal = "center", vertical = "center")
        ws.cell(16, 8 + j * 2).alignment = Alignment(horizontal = "center", vertical = "center")
        ws.cell(16, 7 + j * 2).number_format = format_obj["number_format"]["number"](0)
        ws.cell(16, 8 + j * 2).number_format = format_obj["number_format"]["number"](0)

    return {
        "status": "ok",
        "response": None
    }

def format_inspections_block_freight(ws, data, format_obj, data_validation_obj):

    # define editable attribute
    is_editable = Protection(False, False)

    # extract data validation attributes
    dv_employees = data_validation_obj["employees"]
    dv_gauges = data_validation_obj["gauges"]

    # get the correct feature indexing
    feature_index = []
    for j in range(len(data["inspections"][0]["features"])):
        print_feature_id = ws.cell(8, 7 + j * 2).value
        for i in range(len(data["inspections"][0]["features"])):
            if data["inspections"][i]["features"][i]["print_feature_id"] == print_feature_id:
                feature_index.append(i)

    # create the inspections list
    for i in range(len(data["inspections"])):

        # data
        for j in range(len(data["inspections"][i]["features"])):

            # set data
            ws.cell(17 + i, 7 + j * 2).value = data["inspections"][i]["features"][feature_index[j]]["measured"]
            ws.cell(17 + i, 8 + j * 2).value = data["inspections"][i]["features"][feature_index[j]]["gauge"]

            # set formatting
            ws.cell(17 + i, 7 + j * 2).fill = format_obj["fill"]["data"]
            ws.cell(17 + i, 8 + j * 2).fill = format_obj["fill"]["data"]
            ws.cell(17 + i, 7 + j * 2).font = format_obj["font"]["data"]
            ws.cell(17 + i, 8 + j * 2).font = format_obj["font"]["data"]
            ws.cell(17 + i, 7 + j * 2).border = format_obj["border"]["data"]
            ws.cell(17 + i, 8 + j * 2).border = format_obj["border"]["data"]
            ws.cell(17 + i, 7 + j * 2).alignment = format_obj["alignment"]["data"]
            ws.cell(17 + i, 8 + j * 2).alignment = format_obj["alignment"]["data"]
            ws.cell(17 + i, 7 + j * 2).number_format = format_obj["number_format"]["number"](data["print"][j]["precision"])
            ws.cell(17 + i, 8 + j * 2).number_format = format_obj["number_format"]["text"]

            # set editable state
            ws.cell(17 + i, 7 + j * 2).protection = is_editable
            ws.cell(17 + i, 8 + j * 2).protection = is_editable

            # set data validation
            dv_gauges.add(ws.cell(17 + i, 8 + j * 2).coordinate)

            # set conditional formatting
            column = ws.cell(17 + i, 7 + j * 2).column_letter
            usl_address = f"${column}${11}"
            lsl_address = f"${column}${12}"
            measure_address = f"{column}{17 + i}"
            blnk_formula = f"=ISBLANK({measure_address})"
            pass_formula = f"=AND({measure_address}<={usl_address},{measure_address}>={lsl_address})"
            fail_formula = f"=OR({measure_address}>{usl_address},{measure_address}<{lsl_address})"
            blnk_rule = FormulaRule(
                formula = [blnk_formula],
                fill = format_obj["fill"]["data"],
                stopIfTrue = True
            )
            pass_rule = FormulaRule(
                formula = [pass_formula],
                fill = format_obj["pass_fail"]["pass"],
                stopIfTrue = True
            )
            fail_rule = FormulaRule(
                formula = [fail_formula],
                fill = format_obj["pass_fail"]["fail"],
                stopIfTrue = True
            )
            ws.conditional_formatting.add(
                measure_address,
                blnk_rule
            )
            ws.conditional_formatting.add(
                measure_address,
                pass_rule
            )
            ws.conditional_formatting.add(
                measure_address,
                fail_rule
            )

    row = 17 + len(data["inspections"])
    for j in range(len(data["inspections"][0]["features"])):

        # set data
        ws.cell(row, 7 + j * 2).value = ""
        ws.cell(row, 8 + j * 2).value = ""

        # set format
        ws.cell(row, 7 + j * 2).fill = format_obj["fill"]["data"]
        ws.cell(row, 8 + j * 2).fill = format_obj["fill"]["data"]
        ws.cell(row, 7 + j * 2).font = format_obj["font"]["data"]
        ws.cell(row, 8 + j * 2).font = format_obj["font"]["data"]
        ws.cell(row, 7 + j * 2).border = format_obj["border"]["data"]
        ws.cell(row, 8 + j * 2).border = format_obj["border"]["data"]
        ws.cell(row, 7 + j * 2).alignment = format_obj["alignment"]["data"]
        ws.cell(row, 8 + j * 2).alignment = format_obj["alignment"]["data"]
        ws.cell(row, 7 + j * 2).number_format = format_obj["number_format"]["number"](data["print"][j]["precision"])
        ws.cell(row, 8 + j * 2).number_format = format_obj["number_format"]["text"]

        # set editable state
        ws.cell(row, 7 + j * 2).protection = is_editable
        ws.cell(row, 8 + j * 2).protection = is_editable

        # set data validation
        dv_gauges.add(ws.cell(row, 8 + j * 2).coordinate)

        # set conditional formatting
        column = ws.cell(row, 7 + j * 2).column_letter
        usl_address = f"${column}${11}"
        lsl_address = f"${column}${12}"
        measure_address = f"{column}{row}"
        blnk_formula = f"=ISBLANK({measure_address})"
        pass_formula = f"=AND({measure_address}<={usl_address},{measure_address}>={lsl_address})"
        fail_formula = f"=OR({measure_address}>{usl_address},{measure_address}<{lsl_address})"
        blnk_rule = FormulaRule(
            formula = [blnk_formula],
            fill = format_obj["fill"]["data"],
            stopIfTrue = True
        )
        pass_rule = FormulaRule(
            formula = [pass_formula],
            fill = format_obj["pass_fail"]["pass"],
            stopIfTrue = True
        )
        fail_rule = FormulaRule(
            formula = [fail_formula],
            fill = format_obj["pass_fail"]["fail"],
            stopIfTrue = True
        )
        ws.conditional_formatting.add(
            measure_address,
            blnk_rule
        )
        ws.conditional_formatting.add(
            measure_address,
            pass_rule
        )
        ws.conditional_formatting.add(
            measure_address,
            fail_rule
        )
    ws.merge_cells(start_row = row, end_row = row, start_column = 2, end_column = 3)
    ws.merge_cells(start_row = row, end_row = row, start_column = 4, end_column = 5)
    ws.cell(row, 2).fill = format_obj["fill"]["data"]
    ws.cell(row, 3).fill = format_obj["fill"]["data"]
    ws.cell(row, 4).fill = format_obj["fill"]["data"]
    ws.cell(row, 5).fill = format_obj["fill"]["data"]
    ws.cell(row, 6).fill = format_obj["fill"]["data"]
    ws.cell(row, 2).font = format_obj["font"]["data"]
    ws.cell(row, 3).font = format_obj["font"]["data"]
    ws.cell(row, 4).font = format_obj["font"]["data"]
    ws.cell(row, 5).font = format_obj["font"]["data"]
    ws.cell(row, 6).font = format_obj["font"]["data"]
    ws.cell(row, 2).border = format_obj["border"]["data"]
    ws.cell(row, 3).border = format_obj["border"]["data"]
    ws.cell(row, 4).border = format_obj["border"]["data"]
    ws.cell(row, 5).border = format_obj["border"]["data"]
    ws.cell(row, 6).border = format_obj["border"]["data"]
    ws.cell(row, 2).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 3).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 4).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 5).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 6).alignment = format_obj["alignment"]["data"]
    ws.cell(row, 2).number_format = format_obj["number_format"]["text"]
    ws.cell(row, 3).number_format = format_obj["number_format"]["text"]
    ws.cell(row, 4).number_format = format_obj["number_format"]["datetime"]
    ws.cell(row, 5).number_format = format_obj["number_format"]["datetime"]
    ws.cell(row, 6).number_format = format_obj["number_format"]["number"](0)
    ws.cell(row, 2).protection = is_editable
    ws.cell(row, 3).protection = is_editable
    ws.cell(row, 6).protection = is_editable
    dv_employees.add(f"{ws.cell(row, 2).coordinate}:{ws.cell(row, 3).coordinate}")

    # merge
    ws.merge_cells(range_string = "B16:C16")
    ws.merge_cells(range_string = "D16:E16")

    # set contents
    ws["B16"].value = "Inspector"
    ws["D16"].value = "Date/Time"
    ws["F16"].value = "Check"

    # set formatting
    ws["B16"].fill = format_obj["fill"]["header"]
    ws["C16"].fill = format_obj["fill"]["header"]
    ws["D16"].fill = format_obj["fill"]["header"]
    ws["E16"].fill = format_obj["fill"]["header"]
    ws["F16"].fill = format_obj["fill"]["header"]
    ws["B16"].font = format_obj["font"]["header"]
    ws["C16"].font = format_obj["font"]["header"]
    ws["D16"].font = format_obj["font"]["header"]
    ws["E16"].font = format_obj["font"]["header"]
    ws["F16"].font = format_obj["font"]["header"]
    ws["B16"].border = format_obj["border"]["header"]
    ws["C16"].border = format_obj["border"]["header"]
    ws["D16"].border = format_obj["border"]["header"]
    ws["E16"].border = format_obj["border"]["header"]
    ws["f16"].border = format_obj["border"]["header"]
    ws["B16"].alignment = format_obj["alignment"]["header"]
    ws["C16"].alignment = format_obj["alignment"]["header"]
    ws["D16"].alignment = format_obj["alignment"]["header"]
    ws["E16"].alignment = format_obj["alignment"]["header"]
    ws["F16"].alignment = format_obj["alignment"]["header"]

    # dynamic headers
    for i in range(len(data["inspections"])):

        # merge
        ws.merge_cells(start_row = 17 + i, end_row = 17 + i, start_column = 2, end_column = 3)
        ws.merge_cells(start_row = 17 + i, end_row = 17 + i, start_column = 4, end_column = 5)

        # set contents
        ws.cell(17 + i, 2).value = data["inspections"][i]["employee_name"]
        ws.cell(17 + i, 4).value = data["inspections"][i]["datetime_measured"]
        ws.cell(17 + i, 6).value = data["inspections"][i]["part_index"]

        # set format
        ws.cell(17 + i, 2).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 3).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 4).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 5).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 6).fill = format_obj["fill"]["data"]
        ws.cell(17 + i, 2).font = format_obj["font"]["data"]
        ws.cell(17 + i, 3).font = format_obj["font"]["data"]
        ws.cell(17 + i, 4).font = format_obj["font"]["data"]
        ws.cell(17 + i, 5).font = format_obj["font"]["data"]
        ws.cell(17 + i, 6).font = format_obj["font"]["data"]
        ws.cell(17 + i, 2).border = format_obj["border"]["data"]
        ws.cell(17 + i, 3).border = format_obj["border"]["data"]
        ws.cell(17 + i, 4).border = format_obj["border"]["data"]
        ws.cell(17 + i, 5).border = format_obj["border"]["data"]
        ws.cell(17 + i, 6).border = format_obj["border"]["data"]
        ws.cell(17 + i, 2).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 3).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 4).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 5).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 6).alignment = format_obj["alignment"]["data"]
        ws.cell(17 + i, 2).number_format = format_obj["number_format"]["text"]
        ws.cell(17 + i, 3).number_format = format_obj["number_format"]["text"]
        ws.cell(17 + i, 4).number_format = format_obj["number_format"]["datetime"]
        ws.cell(17 + i, 5).number_format = format_obj["number_format"]["datetime"]
        ws.cell(17 + i, 6).number_format = format_obj["number_format"]["number"](0)

        # set editable state
        ws.cell(17 + i, 2).protection = is_editable
        ws.cell(17 + i, 3).protection = is_editable

        # set data validation
        dv_employees.add(f"{ws.cell(17 + i, 2).coordinate}:{ws.cell(17 + i, 3).coordinate}")

    # apply check coloring
    received_qty = int(data["metadata"]["received_quantity"])
    array_size = 0
    reduced_1 = 0
    reduced_2 = 0
    reduced_3 = 0

    if received_qty < 2:
        array_size = 1
        reduced_1 = 1
        reduced_2 = 1
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 25:
        array_size = 3
        reduced_1 = 3
        reduced_2 = 3
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 50:
        array_size = 5
        reduced_1 = 4
        reduced_2 = 3
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 90:
        array_size = 6
        reduced_1 = 4
        reduced_2 = 3
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 150:
        array_size = 7
        reduced_1 = 4
        reduced_2 = 3
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 280:
        array_size = 10
        reduced_1 = 5
        reduced_2 = 4
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 500:
        array_size = 11
        reduced_1 = 6
        reduced_2 = 4
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 1200:
        array_size = 15
        reduced_1 = 8
        reduced_2 = 5
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 3200:
        array_size = 18
        reduced_1 = 9
        reduced_2 = 6
        reduced_3 = 1
    elif received_qty >= 2 and received_qty <= 10000:
        array_size = 22
        reduced_1 = 11
        reduced_2 = 8
        reduced_3 = 1
    else:
        array_size = 29
        reduced_1 = 15
        reduced_2 = 10
        reduced_3 = 1

    # blue
    reduced_1_color = PatternFill(
        fill_type = "solid",
        start_color = "ff4d4dff",
        end_color = "ff4d4dff"
    )
    reduced_2_color = PatternFill(
        fill_type = "solid",
        start_color = "ffffa64d",
        end_color = "ffffa64d"
    )
    reduced_3_color = PatternFill(
        fill_type = "solid",
        start_color = "ff4dff4d",
        end_color = "ff4dff4d"
    )

    for j in range(reduced_1):
        ws.cell(17 + j, 6).fill = reduced_1_color

    for j in range(reduced_2):
        ws.cell(17 + j, 6).fill = reduced_2_color

    for j in range(reduced_3):
        ws.cell(17 + j, 6).fill = reduced_3_color

    for j in range(len(data["inspections"][0]["features"])):

        # merge
        ws.merge_cells(start_row = 16, end_row = 16, start_column = 7 + j * 2, end_column = 8 + j * 2)

        # set contents
        ws.cell(16, 7 + j * 2).value = j + 1

        # set format
        ws.cell(16, 7 + j * 2).fill = format_obj["fill"]["header"]
        ws.cell(16, 8 + j * 2).fill = format_obj["fill"]["header"]
        ws.cell(16, 7 + j * 2).font = format_obj["font"]["header"]
        ws.cell(16, 8 + j * 2).font = format_obj["font"]["header"]
        ws.cell(16, 7 + j * 2).border = format_obj["border"]["header"]
        ws.cell(16, 8 + j * 2).border = format_obj["border"]["header"]
        ws.cell(16, 7 + j * 2).alignment = Alignment(horizontal = "center", vertical = "center")
        ws.cell(16, 8 + j * 2).alignment = Alignment(horizontal = "center", vertical = "center")
        ws.cell(16, 7 + j * 2).number_format = format_obj["number_format"]["number"](0)
        ws.cell(16, 8 + j * 2).number_format = format_obj["number_format"]["number"](0)

    return {
        "status": "ok",
        "response": None
    }

def get_decimal_format(decimal_places:int) -> str:
    if decimal_places > 0:
        return "#,##0." + "0" * decimal_places
    else:
        return "#,##0"

# --------------------------------------------------

# #region inspection schemas - schemas

# # routes

# @app.route("/inspection_schemas/schemas/create_new_schema/", methods = ["POST"])
# def inspection_schemas_schemas_create_new_schema():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     part_id = form_data["part_id"]
#     search_term = form_data["search_term"]
#     is_locked = int(form_data["is_locked"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # make sure the measurement schema doesn't already exist
#         exists = session.query(inspection_schemas.id)\
#             .filter(inspection_schemas.part_id == part_id)\
#             .first()
#         if exists is not None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [inspection_schemas])
#             }

#         # create new governing record in the database
#         results = inspection_schemas(
#             is_locked = False,
#             part_id = part_id
#         )
#         session.add(results)
#         session.commit()
#         schema_id = results.id
#         if schema_id is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.no_records_added, tables = [inspection_schemas])
#             }

#         # close the database session
#         session.close()

#         # create the first measurement in the new schema
#         returned_obj0 = func_inspection_schemas_add_detail_record(schema_id)

#         # requery the schemas with the same filter parameters
#         returned_obj1 = func_inspection_schemas_get_filtered_schemas(search_term, is_locked)

#         # return the results
#         if returned_obj0["status"] == "ok" and returned_obj1["status"] == "ok":
#             return {
#                 "status": "ok",
#                 "response": {
#                     "schema_list": returned_obj1["response"],
#                     "schema_measurements": returned_obj0["response"]
#                 }
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.no_records_added, tables = [inspection_schemas, inspection_schema_details])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_schemas/schemas/add_row/", methods = ["POST"])
# def inspection_schemas_schemas_add_row():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     schema_id = form_data["schema_id"]

#     try:

#         # open the database session
#         session = Session(engine)

#         # measurement_set if the schema is locked
#         locked_query = session.query(inspection_schemas.is_locked)\
#             .filter(inspection_schemas.id == schema_id)\
#             .first()[0]
#         if locked_query:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_locked, table = inspection_schemas)
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

#     return func_inspection_schemas_add_detail_record(schema_id)

# @app.route("/inspection_schemas/schemas/remove_row/", methods = ["POST"])
# def inspection_schemas_schemas_remove_row():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     detail_id = form_data["detail_id"]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the schema id
#         schema_id = session.query(inspection_schema_details.schema_id)\
#             .filter(inspection_schema_details.id == detail_id)\
#             .first()[0]

#         # measurement_set if the schema is locked
#         locked_query = session.query(inspection_schemas.is_locked)\
#             .filter(inspection_schemas.id == schema_id)\
#             .first()[0]
#         if locked_query:
#             session.close()
#             return text_response(stack()[0][3], message_type.record_locked, table = inspection_schemas)

#         # make sure there is something to be deleted
#         results = session.query(inspection_schema_details.id)\
#             .filter(inspection_schema_details.id == detail_id)\
#             .first()
#         if results is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [inspection_schema_details])
#             }

#         # remove the matching schema id
#         results = session.query(inspection_schema_details)\
#             .filter(inspection_schema_details.id == detail_id)\
#             .delete()

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return the results
#         if results > 0:
#             return {
#                 "status": "ok",
#                 "response": text_response(stack()[0][3], message_type.records_deleted, qty = results, table = inspection_schema_details)
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_deleted, qty = results, table = inspection_schema_details)
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_schemas/schemas/toggle_lock_schema/", methods = ["POST"])
# def inspection_schemas_schemas_toggle_lock_schema():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = form_data["search_term"]
#     is_locked = int(form_data["is_locked"])
#     schema_id = form_data["schema_id"]

#     try:

#         # open the database session
#         session = Session(engine)

#         # make sure the measurement schema exists
#         exists = session.query(inspection_schemas.id)\
#             .filter(inspection_schemas.id == schema_id)\
#             .first()
#         if exists is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [inspection_schemas])
#             }

#         # get the current locked status
#         locked_query = session.query(inspection_schemas.is_locked)\
#             .filter(inspection_schemas.id == schema_id)\
#             .first()[0]

#         # set the locked status
#         rows_affected = session.query(inspection_schemas)\
#             .filter(inspection_schemas.id == schema_id)\
#             .update({ "is_locked": not locked_query })

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # logic gate
#         if rows_affected == 0:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [inspection_schemas])
#             }

#         # requery the schemas
#         returned_obj = func_inspection_schemas_get_filtered_schemas(search_term, is_locked)

#         # return the results
#         if returned_obj["status"] == "ok":
#             return {
#                 "status": "ok",
#                 "response": {
#                     "is_locked": not locked_query,
#                     "schema_list": returned_obj["response"]
#                 }
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_schemas/schemas/save_inspection_schema/", methods = ["POST"])
# def inspection_schemas_schemas_save_inspection_schema():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     schema_id = form_data["schema_id"]

#     try:

#         # open the database session
#         session = Session(engine)

#         # make sure the measurement schema exists
#         exists = session.query(inspection_schema_details.id)\
#             .filter(inspection_schema_details.schema_id == schema_id)\
#             .first()
#         if exists is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [inspection_schema_details])
#             }

#         # query the database
#         rows_affected = 0
#         for obj in form_data["data"]:
#             results = session.query(inspection_schema_details)\
#                 .filter(inspection_schema_details.id == obj["detail_id"])
#             field_affected = 0
#             for x in obj["contents"]:
#                 field_affected += results.update({ x["key"]: x["value"] })
#             if field_affected > 0:
#                 rows_affected += 1

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return the results
#         if rows_affected > 0:
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [inspection_schema_details])
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [inspection_schema_details])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_schemas/schemas/delete_inspection_schema/", methods = ["POST"])
# def inspection_schemas_schemas_delete_inspection_schema():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     schema_id = form_data["schema_id"]
#     search_term = form_data["search_term"]
#     is_locked = int(form_data["is_locked"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # make sure the referenced schema exists
#         exists = session.query(inspection_schemas.id)\
#             .filter(inspection_schemas.id == schema_id).first()
#         if exists is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [inspection_schemas])
#             }

#         # measurement_set if the schema is already locked
#         schema_is_locked = session.query(inspection_schemas.is_locked)\
#             .filter(inspection_schemas.id == schema_id)\
#             .first()[0]
#         if schema_is_locked:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_locked, table = inspection_schemas)
#             }

#         # delete the referenced schema
#         details_results = session.query(inspection_schema_details)\
#             .filter(inspection_schema_details.schema_id == schema_id)\
#             .delete()
#         schema_results = session.query(inspection_schemas)\
#             .filter(inspection_schemas.id == schema_id)\
#             .delete()

#         # logic gate
#         if details_results == 0 and schema_results == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_deleted, tables = [inspection_schemas, inspection_schema_details])
#             }
#         elif details_results == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_deleted, qty = details_results, table = inspection_schema_details)
#             }
#         elif schema_results == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_deleted, qty = schema_results, table = inspection_schemas)
#             }

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # requery the schemas
#         return func_inspection_schemas_get_filtered_schemas(search_term, is_locked)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_schemas/schemas/get_filtered_inspection_schemas/", methods = ["POST"])
# def inspection_schemas_schema_get_filtered_schemas():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     is_locked = int(form_data["is_locked"])

#     # call the relevant method
#     return func_inspection_schemas_get_filtered_schemas(search_term, is_locked)

# @app.route("/inspection_schemas/schemas/get_filtered_parts/", methods = ["POST"])
# def inspection_schemas_schema_get_filtered_parts():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])

#     try:

#         # start the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(parts.id, parts.item, parts.drawing, parts.revision)\
#             .filter(or_(parts.item.ilike(f"%{search_term}%"), parts.drawing.ilike(f"%{search_term}%"), parts.revision.ilike(f"%{search_term}%")))\
#             .order_by(parts.item.asc(), parts.drawing.asc(), parts.revision.asc()).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, item, drawing, revision in results:
#                 output_arr.append({
#                     "id": id,
#                     "item": item,
#                     "drawing": drawing,
#                     "revision": revision,
#                     "part_name": f"{item}, {drawing}, {revision.upper()}"
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# # recycled methods

# def func_inspection_schemas_add_detail_record(schema_id:int):

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the default specification type id
#         default_specification_type_id = session.query(specification_types.id)\
#             .order_by(specification_types.name.asc())\
#             .first()[0]

#         # get the default dimension type id
#         default_dimension_type_id = session.query(dimension_types.id)\
#             .order_by(dimension_types.name.asc())\
#             .first()[0]

#         # get the default frequency type id
#         default_frequency_type_id = session.query(frequency_types.id)\
#             .order_by(frequency_types.name.asc())\
#             .first()[0]

#         # get the default gauge type id
#         default_gauge_type_id = session.query(gauge_types.id)\
#             .order_by(gauge_types.name.asc())\
#             .first()[0]

#         # define the placeholder values
#         default_name = "DIM X"
#         default_nominal = 0
#         default_usl = 0
#         default_lsl = 0
#         default_precision = 1

#         # set the placeholder data
#         results = inspection_schema_details(
#             name = default_name,
#             nominal = default_nominal,
#             usl = default_usl,
#             lsl = default_lsl,
#             precision = default_precision,
#             specification_type_id = default_specification_type_id,
#             dimension_type_id = default_dimension_type_id,
#             frequency_type_id = default_frequency_type_id,
#             gauge_type_id = default_gauge_type_id,
#             schema_id = schema_id
#         )
#         session.add(results)
#         session.commit()

#         # capture the new serial id
#         detail_id = results.id

#         # close the database session
#         session.close()

#         # return the results
#         if results.id is not None:
#             return {
#                 "status": "ok",
#                 "response": {
#                     "id": detail_id,
#                     "name": default_name,
#                     "nominal": default_nominal,
#                     "usl": default_usl,
#                     "lsl": default_lsl,
#                     "precision": default_precision,
#                     "specification_type_id": default_specification_type_id,
#                     "dimension_type_id": default_dimension_type_id,
#                     "frequency_type_id": default_frequency_type_id,
#                     "gauge_type_id": default_gauge_type_id,
#                     "schema_id": schema_id
#                 }
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.no_records_added, tables = [inspection_schema_details])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# def func_inspection_schemas_get_filtered_schemas(search_term:str, is_locked:int):

#     # define the output columns
#     columns = [
#         inspection_schemas.id,
#         inspection_schemas.is_locked,
#         parts.id,
#         parts.item,
#         parts.drawing,
#         parts.revision
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(*columns)\
#             .join(parts, (parts.id == inspection_schemas.part_id))\
#             .filter(or_(parts.item.ilike(f"%{search_term}%"), parts.drawing.ilike(f"%{search_term}%"), parts.revision.ilike(f"%{search_term}%")))
#         if is_locked >= 0:
#             results = results.filter(inspection_schemas.is_locked == bool(is_locked))
#         results = results.order_by(parts.item.asc(), parts.drawing.asc(), parts.revision.asc()).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for schema_id, is_locked, part_id, item, drawing, revision in results:
#                 output_arr.append({
#                     "schema_id": schema_id,
#                     "is_locked": is_locked,
#                     "part_id": part_id,
#                     "item": item,
#                     "drawing": drawing,
#                     "revision": revision.upper()
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection schemas - schema view

# @app.route("/inspection_schemas/view/get_schema_features/", methods = ["POST"])
# def inspection_schemas_view_get_schema_measurements():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     schema_id = form_data["schema_id"]

#     # define the output columns
#     columns = [
#         inspection_schemas.id,
#         inspection_schemas.is_locked,
#         parts.id,
#         inspection_schema_details.id,
#         inspection_schema_details.name,
#         inspection_schema_details.nominal,
#         inspection_schema_details.usl,
#         inspection_schema_details.lsl,
#         inspection_schema_details.precision,
#         inspection_schema_details.specification_type_id,
#         inspection_schema_details.dimension_type_id,
#         inspection_schema_details.frequency_type_id,
#         inspection_schema_details.gauge_type_id
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the requested schema
#         results = session.query(*columns)\
#             .join(parts, (parts.id == inspection_schemas.part_id))\
#             .join(inspection_schema_details, (inspection_schema_details.schema_id == inspection_schemas.id))\
#             .filter(inspection_schemas.id == schema_id)\
#             .order_by(inspection_schema_details.id.asc(), inspection_schema_details.name.asc()).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for schema_id, is_locked, part_id, detail_id, name, nominal, usl, lsl, precision, specification_type_id, dimension_type_id, frequency_type_id, gauge_type_id in results:

#                 # parse decimal to float
#                 nominal_flt = round(float(nominal), precision)
#                 usl_flt = round(float(usl), precision)
#                 lsl_flt = round(float(lsl), precision)

#                 output_arr.append({
#                     "schema_id": schema_id,
#                     "is_locked": is_locked,
#                     "part_id": part_id,
#                     "detail_id": detail_id,
#                     "name": name,
#                     "nominal": nominal_flt,
#                     "usl": usl_flt,
#                     "lsl": lsl_flt,
#                     "precision": precision,
#                     "specification_type_id": specification_type_id,
#                     "dimension_type_id": dimension_type_id,
#                     "frequency_type_id": frequency_type_id,
#                     "gauge_type_id": gauge_type_id
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspection_schemas, inspection_schema_details])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# --------------------------------------------------

# #region inspection records - inspection records

# # routes

# @app.route("/inspection_records/inspection_records/create_new_record/", methods = ["POST"])
# def inspection_records_inspection_records_create_new_record():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     part_id = int(form_data["part_id"])
#     schema_id = int(form_data["schema_id"])
#     employee_id = int(form_data["employee_id"])
#     part_search_term = str(form_data["part_search_term"])
#     job_number_search_term = str(form_data["job_number_search_term"])
#     started_after_str = str(form_data["started_after"])
#     finished_before_str = str(form_data["finished_before"])
#     material_type_search_term = str(form_data["material_type_search_term"])
#     employee_search_term = str(form_data["employee_search_term"])
#     disposition_search_term = str(form_data["disposition_search_term"])
#     receiver_number_search_term = str(form_data["receiver_number_search_term"])
#     purchase_order_search_term = str(form_data["purchase_order_search_term"])
#     lot_number_search_term = str(form_data["lot_number_search_term"])
#     supplier_search_term = str(form_data["supplier_search_term"])

#     # convert date strings to datetime objects
#     started_after = datetime.date(1970, 1, 1)
#     finished_before = datetime.date(2100, 1, 1)
#     if started_after_str != "":
#         started_after = datetime.datetime.strptime(started_after_str, "%Y-%m-%d")
#     if finished_before_str != "":
#         finished_before = datetime.datetime.strptime(finished_before_str, "%Y-%m-%d")

#     try:

#         # open the database session
#         session = Session(engine)

#         # make sure the part exists
#         exists = session.query(parts.id)\
#             .filter(parts.id == part_id).first()
#         if exists is None:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts])
#             }

#         # make sure this part isn't already associated with an inspection report
#         exists = session.query(parts.id, inspection_records.id, inspections.id)\
#             .join(parts, (parts.id == inspections.part_id))\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .filter(parts.id == part_id).first()
#         if exists is not None:
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [parts, inspection_records])
#             }

#         # close the database session
#         session.close()

#         # create the new records
#         func_inspections_add_inspection(part_id, -1, schema_id, employee_id, 0)

#         # return the filtered records
#         return func_inspection_records_get_filtered_records(
#             part_search_term,
#             started_after,
#             finished_before,
#             material_type_search_term,
#             employee_search_term,
#             disposition_search_term,
#             receiver_number_search_term,
#             purchase_order_search_term,
#             job_number_search_term,
#             lot_number_search_term,
#             supplier_search_term
#         )

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspection_records/delete_record/", methods = ["POST"])
# def inspection_records_inspection_records_delete_record():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = form_data["inspection_record_id"]
#     part_search_term = str(form_data["part_search_term"])
#     job_number_search_term = str(form_data["job_number_search_term"])
#     started_after_str = form_data["started_after"]
#     finished_before_str = form_data["finished_before"]
#     material_type_search_term = str(form_data["material_type_search_term"])
#     employee_search_term = str(form_data["employee_search_term"])
#     disposition_search_term = str(form_data["disposition_search_term"])
#     receiver_number_search_term = str(form_data["receiver_number_search_term"])
#     purchase_order_search_term = str(form_data["purchase_order_search_term"])
#     lot_number_search_term = str(form_data["lot_number_search_term"])
#     supplier_search_term = str(form_data["supplier_search_term"])

#     # convert date strings to datetime objects
#     started_after = datetime.date(1970, 1, 1)
#     finished_before = datetime.date(2100, 1, 1)
#     if started_after_str != "":
#         started_after = datetime.datetime.strptime(started_after_str, "%Y-%m-%d")
#     if finished_before_str != "":
#         finished_before = datetime.datetime.strptime(finished_before_str, "%Y-%m-%d")

#     try:

#         # open the database connection
#         session = Session(engine)

#         # make sure the required records exist
#         inspection_exists = session.query(inspection_records.id)\
#             .filter(inspection_records.id == inspection_record_id).first()
#         if inspection_exists is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [inspection_records])
#             }

#         # get the associated ids
#         inspection_ids_query = session.query(inspections.id)\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .filter(inspection_records.id == inspection_record_id).all()
#         inspection_ids = [x[0] for x in inspection_ids_query]

#         feature_ids_query = session.query(features.id)\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .filter(inspection_records.id == inspection_record_id).all()
#         feature_ids = [x[0] for x in feature_ids_query]

#         deviation_ids_query = session.query(deviations.id)\
#             .join(features, (features.id == deviations.feature_id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .filter(inspection_records.id == inspection_record_id).all()
#         deviation_ids = [x[0] for x in deviation_ids_query]

#         # delete the referenced records
#         session.query(deviations)\
#             .filter(deviations.id.in_(deviation_ids)).delete(synchronize_session = False)

#         session.query(features)\
#             .filter(features.id.in_(feature_ids)).delete(synchronize_session = False)

#         session.query(inspections)\
#             .filter(inspections.id.in_(inspection_ids)).delete(synchronize_session = False)

#         session.query(inspections_purchase_orders)\
#             .filter(inspections_purchase_orders.inspection_record_id == inspection_record_id).delete(synchronize_session = False)

#         session.query(inspections_lot_numbers)\
#             .filter(inspections_lot_numbers.inspection_record_id == inspection_record_id).delete(synchronize_session = False)

#         session.query(inspection_records)\
#             .filter(inspection_records.id == inspection_record_id).delete(synchronize_session = False)

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return the filtered records
#         return func_inspection_records_get_filtered_records(
#             part_search_term,
#             started_after,
#             finished_before,
#             material_type_search_term,
#             employee_search_term,
#             disposition_search_term,
#             receiver_number_search_term,
#             purchase_order_search_term,
#             job_number_search_term,
#             lot_number_search_term,
#             supplier_search_term
#         )

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspection_records/get_filtered_records/", methods = ["POST"])
# def inspection_records_inspection_records_get_filtered_records():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     part_search_term = str(form_data["part"])
#     started_after_str = str(form_data["started_after"])
#     finished_before_str = str(form_data["finished_before"])
#     material_type_search_term = str(form_data["material_type"])
#     employee_search_term = str(form_data["employee"])
#     disposition_search_term = str(form_data["disposition"])
#     receiver_number_search_term = str(form_data["receiver_number"])
#     purchase_order_search_term = str(form_data["purchase_order"])
#     job_number_search_term = str(form_data["job_number"])
#     lot_number_search_term = str(form_data["lot_number"])
#     supplier_search_term = str(form_data["supplier"])

#     # convert date strings to datetime objects
#     started_after = datetime.date(1970, 1, 1)
#     finished_before = datetime.date(2100, 1, 1)
#     if started_after_str != "":
#         started_after = datetime.datetime.strptime(started_after_str, "%Y-%m-%d")
#     if finished_before_str != "":
#         finished_before = datetime.datetime.strptime(finished_before_str, "%Y-%m-%d")

#     return func_inspection_records_get_filtered_records(
#         part_search_term,
#         started_after,
#         finished_before,
#         material_type_search_term,
#         employee_search_term,
#         disposition_search_term,
#         receiver_number_search_term,
#         purchase_order_search_term,
#         job_number_search_term,
#         lot_number_search_term,
#         supplier_search_term
#     )

# @app.route("/inspection_records/inspection_records/get_filtered_parts/", methods = ["POST"])
# def inspection_records_inspection_records_get_filtered_parts():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = form_data["search_term"]

#     try:

#         # start the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(parts.id, parts.item, parts.drawing)\
#             .filter(or_(parts.item.ilike(f"%{search_term}%"), parts.drawing.ilike(f"%{search_term}%")))\
#             .order_by(parts.item.asc(), parts.drawing.asc())\
#             .distinct(parts.item, parts.drawing).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, item, drawing in results:
#                 output_arr.append({
#                     "id": id,
#                     "item": item,
#                     "drawing": drawing,
#                     "part_name": f"{item}, {drawing}"
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspection_records/get_filtered_schemas/", methods = ["POST"])
# def inspection_records_inspection_records_get_filtered_schemas():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     part_id = int(form_data["part_id"])

#     # define the output columns
#     columns = [
#         inspection_schemas.id,
#         parts.item,
#         parts.drawing,
#         parts.revision
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the associated item and drawing from part id
#         part_query = session.query(parts.item, parts.drawing)\
#             .filter(parts.id == part_id).first()
#         if part_query is None:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts])
#             }
#         item, drawing = part_query

#         # query the database
#         results = session.query(*columns)\
#             .join(parts, (parts.id == inspection_schemas.part_id))\
#             .filter(or_(parts.item.ilike(f"%{search_term}%"), parts.drawing.ilike(f"%{search_term}%"), parts.revision.ilike(f"%{search_term}%")))\
#             .filter(and_(parts.item.ilike(item), parts.drawing.ilike(drawing)))\
#             .order_by(parts.item.asc(), parts.drawing.asc(), parts.revision.asc()).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for schema_id, item, drawing, revision in results:
#                 output_arr.append({
#                     "schema_id": schema_id,
#                     "name": f"{item}, {drawing}, {revision.upper()}"
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspection_schemas])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspection_records/get_filtered_employees/", methods = ["POST"])
# def inspection_records_inspection_records_get_filtered_employees():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = form_data["search_term"]

#     # define the required fields
#     columns = [
#         employees.id,
#         employees.first_name,
#         employees.last_name
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(*columns)\
#             .filter(or_(employees.first_name.ilike(f"%{search_term}%"), employees.last_name.ilike(f"%{search_term}%")))\
#             .order_by(employees.last_name.asc(), employees.first_name.asc()).all()
#         if search_term.isnumeric():
#             results = session.query(*columns)\
#                 .filter(employees.id == search_term)\
#                 .order_by(employees.last_name.asc(), employees.first_name.asc()).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, first_name, last_name in results:
#                 output_arr.append({
#                     "id": id,
#                     "first_name": first_name,
#                     "last_name": last_name,
#                     "name": f"{last_name}, {first_name}"
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [employees])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# # recycled methods

# def func_inspection_records_get_filtered_records(part_search_term:str, started_after:datetime, finished_before:datetime, material_type_search_term:str, employee_search_term:str, disposition_search_term:str, receiver_number_search_term:str, purchase_order_search_term:str, job_number_search_term:str, lot_number_search_term:str, supplier_search_term:str):

#     # define the required fields
#     columns = [
#         inspection_records.id,
#         parts.id,
#         parts.item,
#         parts.drawing,
#         inspections.disposition_id,
#         inspection_records.material_type_id,
#         inspection_records.employee_id,
#         disposition_types.name
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(*columns)\
#             .join(inspections, (inspections.inspection_record_id == inspection_records.id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .outerjoin(material_types, (material_types.id == inspection_records.material_type_id))\
#             .outerjoin(employees, (employees.id == inspection_records.employee_id))\
#             .outerjoin(disposition_types, (disposition_types.id == inspections.disposition_id))\
#             .outerjoin(inspections_purchase_orders, (inspections_purchase_orders.inspection_record_id == inspection_records.id))\
#             .outerjoin(purchase_orders, (purchase_orders.id == inspections_purchase_orders.purchase_order_id))\
#             .outerjoin(receiver_numbers, (receiver_numbers.purchase_order_id == purchase_orders.id))\
#             .outerjoin(inspections_job_numbers, (inspections_job_numbers.inspection_id == inspections.id))\
#             .outerjoin(job_numbers, (job_numbers.id == inspections_job_numbers.job_number_id))\
#             .outerjoin(inspections_lot_numbers, (inspections_lot_numbers.inspection_record_id == inspection_records.id))\
#             .outerjoin(lot_numbers, (lot_numbers.id == inspections_lot_numbers.lot_number_id))\
#             .outerjoin(suppliers, (suppliers.id == purchase_orders.supplier_id))\
#             .filter(inspections.datetime_measured >= started_after)\
#             .filter(or_(inspections.datetime_measured <= finished_before, inspections.datetime_measured == None))\
#             .filter(or_(parts.item.ilike(f"%{part_search_term}%"), parts.drawing.ilike(f"%{part_search_term}%"), parts.revision.ilike(f"%{part_search_term}%")))\
#             .filter(material_types.name.ilike(f"%{material_type_search_term}%"))\
#             .filter(or_(employees.first_name.ilike(f"%{employee_search_term}%"), employees.last_name.ilike(f"%{employee_search_term}%")))\
#             .filter(disposition_types.name.ilike(f"%{disposition_search_term}%"))\
#             .filter(or_(receiver_numbers.name == None, receiver_numbers.name.ilike(f"%{receiver_number_search_term}%")))\
#             .filter(or_(purchase_orders.name == None, purchase_orders.name.ilike(f"%{purchase_order_search_term}%")))\
#             .filter(or_(job_numbers.name == None, job_numbers.name.ilike(f"%{job_number_search_term}%")))\
#             .filter(or_(lot_numbers.name == None, lot_numbers.name.ilike(f"%{lot_number_search_term}%")))\
#             .filter(or_(suppliers.name == None, suppliers.name.ilike(f"%{supplier_search_term}%")))\
#             .order_by(parts.drawing.asc(), parts.item.asc())\
#             .distinct(parts.drawing, parts.item)

#         # close the database session
#         session.close()

#         # return the results
#         if len(results.all()) > 0:
#             output_arr = []
#             for inspection_record_id, part_id, item, drawing, disposition_type_id, material_type_id, employee_id, disposition in results.all():
#                 output_arr.append({
#                     "inspection_record_id": inspection_record_id,
#                     "part_id": part_id,
#                     "item": item,
#                     "drawing": drawing,
#                     "disposition_type_id": disposition_type_id,
#                     "material_type_id": material_type_id,
#                     "employee_id": employee_id,
#                     "disposition": disposition
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection records - inspections

# # routes

# @app.route("/inspection_records/inspections/add_inspection/", methods = ["POST"])
# def inspection_records_inspections_add_inspection():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     schema_id = int(form_data["schema_id"])
#     employee_id = int(form_data["employee_id"])
#     started_after = datetime.datetime.strptime(str(form_data["started_after"]), "%Y-%m-%dT%H:%M")
#     finished_before = datetime.datetime.strptime(str(form_data["finished_before"]), "%Y-%m-%dT%H:%M")
#     employee = str(form_data["employee_filter"])
#     part_index_str = str(form_data["part_index_filter"])
#     revision = str(form_data["revision_filter"])
#     inspection_type = int(form_data["inspection_type_filter"])
#     disposition_type = int(form_data["disposition_type_filter"])
#     shown_job_numbers = list(form_data["shown_job_numbers"])
#     shown_purchase_orders = list(form_data["shown_purchase_orders"])
#     jn_active = bool(form_data["job_numbers_active"])
#     po_active = bool(form_data["purchase_orders_active"])

#     if part_index_str == "":
#         part_index = -1
#     else:
#         part_index = int(part_index_str)

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the part id
#         part_id = session.query(inspection_schemas.part_id)\
#             .filter(inspection_schemas.id == schema_id).first()[0]

#         # close the database session
#         session.close()

#         # add the records
#         func_inspections_add_inspection(part_id, inspection_record_id, schema_id, employee_id, -1)

#         # return an updated list of measurement sets
#         return func_inspections_get_filtered_inspections(inspection_record_id, started_after, finished_before, employee, part_index, revision, inspection_type, disposition_type, shown_job_numbers, shown_purchase_orders, jn_active, po_active)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspections/delete_inspection/", methods = ["POST"])
# def inspection_records_inspections_delete_inspection():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     inspection_id = int(form_data["inspection_id"])
#     started_after = datetime.datetime.strptime(str(form_data["started_after"]), "%Y-%m-%dT%H:%M")
#     finished_before = datetime.datetime.strptime(str(form_data["finished_before"]), "%Y-%m-%dT%H:%M")
#     employee = str(form_data["employee_filter"])
#     part_index_str = str(form_data["part_index_filter"])
#     revision = str(form_data["revision_filter"])
#     inspection_type = int(form_data["inspection_type_filter"])
#     disposition_type = int(form_data["disposition_type_filter"])
#     shown_job_numbers = list(form_data["shown_job_numbers"])
#     shown_purchase_orders = list(form_data["shown_purchase_orders"])
#     jn_active = bool(form_data["job_numbers_active"])
#     po_active = bool(form_data["purchase_orders_active"])

#     if part_index_str == "":
#         part_index = -1
#     else:
#         part_index = int(part_index_str)

#     try:

#         # open the database session
#         session = Session(engine)

#         # remove associated measurements
#         measurements_query = session.query(features)\
#             .filter(features.inspection_id == inspection_id)\
#             .delete()

#         # narrow search to measurement set
#         measurement_set_query = session.query(inspections)\
#             .filter(inspections.id == inspection_id)\
#             .delete()

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return an updated list of measurement sets
#         return func_inspections_get_filtered_inspections(inspection_record_id, started_after, finished_before, employee, part_index, revision, inspection_type, disposition_type, shown_job_numbers, shown_purchase_orders, jn_active, po_active)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspections/copy_inspection/", methods = ["POST"])
# def inspection_records_inspections_copy_inspection():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     inspection_id = int(form_data["inspection_id"])
#     started_after = datetime.datetime.strptime(str(form_data["started_after"]), "%Y-%m-%dT%H:%M")
#     finished_before = datetime.datetime.strptime(str(form_data["finished_before"]), "%Y-%m-%dT%H:%M")
#     employee = str(form_data["employee_filter"])
#     part_index_str = str(form_data["part_index_filter"])
#     revision = str(form_data["revision_filter"])
#     inspection_type = int(form_data["inspection_type_filter"])
#     disposition_type = int(form_data["disposition_type_filter"])
#     shown_job_numbers = list(form_data["shown_job_numbers"])
#     shown_purchase_orders = list(form_data["shown_purchase_orders"])
#     jn_active = bool(form_data["job_numbers_active"])
#     po_active = bool(form_data["purchase_orders_active"])

#     if part_index_str == "":
#         part_index = -1
#     else:
#         part_index = int(part_index_str)

#     # define the query columns
#     inspection_columns = [
#         inspections.part_index,
#         inspections.inspection_record_id,
#         inspections.part_id,
#         inspections.employee_id,
#         inspections.inspection_type_id,
#         inspections.disposition_id
#     ]
#     feature_columns = [
#         features.name,
#         features.nominal,
#         features.usl,
#         features.lsl,
#         features.measured,
#         features.precision,
#         features.specification_type_id,
#         features.dimension_type_id,
#         features.frequency_type_id,
#         features.gauge_id
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the current inspection
#         current_inspection = session.query(*inspection_columns).filter(inspections.id == inspection_id).first()
#         if current_inspection is None:
#             session.close()
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#             }
#         current_part_index, current_inspection_record_id, current_part_id, current_employee_id, current_inspection_type_id, current_disposition_id = current_inspection

#         # get the current features
#         current_features = session.query(*feature_columns).filter(features.inspection_id == inspection_id).all()
#         if len(current_features) == 0:
#             session.close()
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#             }
#         features_arr = []
#         for name, nominal, usl, lsl, measured, precision, specification_type_id, dimension_type_id, frequency_type_id, gauge_id in current_features:

#             measured_flt = 0
#             if measured is None:
#                 measured_flt = None
#             else:
#                 measured_flt = float(measured)

#             features_arr.append({
#                 "name": name,
#                 "nominal": float(nominal),
#                 "usl": float(usl),
#                 "lsl": float(lsl),
#                 "measured": measured_flt,
#                 "precision": int(precision),
#                 "specification_type_id": int(specification_type_id),
#                 "dimension_type_id": int(dimension_type_id),
#                 "frequency_type_id": int(frequency_type_id),
#                 "gauge_id": int(gauge_id)
#             })

#         # create the new inspection
#         new_inspection = inspections(
#             part_index = current_part_index,
#             datetime_measured = datetime.datetime.now(),
#             inspection_record_id = current_inspection_record_id,
#             part_id = current_part_id,
#             employee_id = current_employee_id,
#             inspection_type_id = current_inspection_type_id,
#             disposition_id = current_disposition_id
#         )
#         session.add(new_inspection)
#         session.commit()
#         new_inspection_id = new_inspection.id

#         # create the new features
#         for row in features_arr:
#             new_feature = features(
#                 name = row["name"],
#                 nominal = row["nominal"],
#                 usl = row["usl"],
#                 lsl = row["lsl"],
#                 measured = row["measured"],
#                 precision = row["precision"],
#                 inspection_id = new_inspection_id,
#                 specification_type_id = row["specification_type_id"],
#                 dimension_type_id = row["dimension_type_id"],
#                 frequency_type_id = row["frequency_type_id"],
#                 gauge_id = row["gauge_id"]
#             )
#             session.add(new_feature)
#         session.commit()

#         # close the database session
#         session.close()

#         # return an updated list of measurement sets
#         return func_inspections_get_filtered_inspections(inspection_record_id, started_after, finished_before, employee, part_index, revision, inspection_type, disposition_type, shown_job_numbers, shown_purchase_orders, jn_active, po_active)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspections/save_edits/", methods = ["POST"])
# def inspection_records_inspections_save_edits():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     data_list = list(form_data["data"])

#     # parse data object
#     data_object = []
#     for x in data_list:
#         data_object.append({
#             "inspection_id": int(x["inspection_id"]),
#             "data": {
#                 "part_index": int(x["part_index"]),
#                 "inspection_type_id": int(x["inspection_type_id"]),
#                 "datetime_measured": datetime.datetime.strptime(x["timestamp"], "%Y-%m-%dT%H:%M"),
#                 "employee_id": int(x["employee_id"])
#             }
#         })

#     try:

#         # open the database session
#         session = Session(engine)

#         # narrow query object to proper scope
#         rows_affected = 0
#         for x in data_object:
#             inspections_query = session.query(inspections)\
#                 .filter(inspections.id == x["inspection_id"])

#             is_affected = 0
#             for k, v in x["data"].items():
#                 is_affected += inspections_query.update({ k: v })

#             if is_affected > 0:
#                 rows_affected += 1

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return an updated list of measurement sets
#         return {
#             "status": "alert",
#             "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [inspections])
#         }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspections/get_filtered_schemas/", methods = ["POST"])
# def inspection_records_inspections_get_filtered_schemas():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     search_term = str(form_data["search_term"])

#     # define the output columns
#     columns = [
#         inspection_schemas.id,
#         parts.id,
#         parts.item,
#         parts.drawing,
#         parts.revision
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the item and drawing from the provided inspection id
#         part_query = session.query(parts.item, parts.drawing)\
#             .join(inspections, (inspections.part_id == parts.id))\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .filter(inspection_records.id == inspection_record_id)\
#             .order_by(parts.item.asc(), parts.drawing.asc())
#         if part_query.first() is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspections, inspection_records])
#             }
#         item, drawing = part_query.first()

#         # get the list of matching measurement set schemas
#         set_schema_query = session.query(*columns)\
#             .join(parts, (parts.id == inspection_schemas.part_id))\
#             .filter(func.lower(parts.item) == func.lower(item))\
#             .filter(func.lower(parts.drawing) == func.lower(drawing))\
#             .filter(or_(parts.item.ilike(f"%{search_term}%"), parts.drawing.ilike(f"%{search_term}%"), parts.revision.ilike(f"%{search_term}%")))\
#             .order_by(parts.item.asc(), parts.drawing.asc(), parts.revision.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(set_schema_query) > 0:
#             output_arr = []
#             for schema_id, part_id, item, drawing, revision in set_schema_query:
#                 output_arr.append({
#                     "schema_id": schema_id,
#                     "part_id": part_id,
#                     "item": item,
#                     "drawing": drawing,
#                     "revision": revision.upper(),
#                     "name": f"{item}, {drawing}, {revision.upper()}"
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspection_schemas])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspections/get_filtered_inspections/", methods = ["POST"])
# def inspection_records_inspections_get_filtered_inspections():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     started_after = datetime.datetime.strptime(str(form_data["started_after"]), "%Y-%m-%dT%H:%M")
#     finished_before = datetime.datetime.strptime(str(form_data["finished_before"]), "%Y-%m-%dT%H:%M")
#     employee = str(form_data["employee_filter"])
#     part_index_str = str(form_data["part_index_filter"])
#     revision = str(form_data["revision_filter"])
#     inspection_type = int(form_data["inspection_type_filter"])
#     disposition_type = int(form_data["disposition_type_filter"])
#     shown_job_numbers = list(form_data["shown_job_numbers"])
#     shown_purchase_orders = list(form_data["shown_purchase_orders"])
#     jn_active = bool(form_data["job_numbers_active"])
#     po_active = bool(form_data["purchase_orders_active"])

#     if part_index_str == "":
#         part_index = -1
#     else:
#         part_index = int(part_index_str)

#     return func_inspections_get_filtered_inspections(inspection_record_id, started_after, finished_before, employee, part_index, revision, inspection_type, disposition_type, shown_job_numbers, shown_purchase_orders, jn_active, po_active)

# @app.route("/inspection_records/inspections/get_job_numbers/", methods = ["POST"])
# def inspection_records_inspections_get_job_numbers():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(job_numbers.id, job_numbers.name)\
#             .join(inspections_job_numbers, (inspections_job_numbers.job_number_id == job_numbers.id))\
#             .join(inspections, (inspections.id == inspections_job_numbers.inspection_id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .order_by(job_numbers.name.asc()).distinct(job_numbers.name).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "is_active": False
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/inspections/get_purchase_orders/", methods = ["POST"])
# def inspection_records_inpsections_get_purchase_orders():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(purchase_orders.id, purchase_orders.name)\
#             .join(inspections_purchase_orders, (inspections_purchase_orders.purchase_order_id == purchase_orders.id))\
#             .join(inspection_records, (inspection_records.id == inspections_purchase_orders.inspection_record_id))\
#             .filter(inspection_records.id == inspection_record_id)\
#             .order_by(purchase_orders.name.asc()).distinct(purchase_orders.name).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "is_active": False
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# # recycled methods

# def func_inspections_add_inspection(part_id:int, inspection_record_id:int, schema_id:int, employee_id:int, part_index:int):

#     # schema detail columns
#     schema_details_columns = [
#         inspection_schema_details.name,
#         inspection_schema_details.nominal,
#         inspection_schema_details.usl,
#         inspection_schema_details.lsl,
#         inspection_schema_details.precision,
#         inspection_schema_details.specification_type_id,
#         inspection_schema_details.dimension_type_id,
#         inspection_schema_details.frequency_type_id,
#         inspection_schema_details.gauge_type_id
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the inspection schema details
#         schema_query = session.query(parts.item, parts.drawing)\
#             .join(inspection_schemas, (inspection_schemas.part_id == parts.id))\
#             .filter(inspection_schemas.id == schema_id).first()
#         if schema_query is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspection_schemas])
#             }
#         schema_item, schema_drawing = schema_query

#         # get the part details
#         part_query = session.query(parts.item, parts.drawing)\
#             .filter(parts.id == part_id).first()
#         if part_query is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts])
#             }
#         part_item, part_drawing = part_query

#         # measurement_set if schema and part match up
#         if schema_item != part_item and schema_drawing != part_drawing:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.generic, err_msg = "the provided schema does not match the provided part")
#             }

#         # create a new inspection report if the provided inspection_id is -1, or ensure it exists
#         if inspection_record_id == -1:

#             # create a new inspection report
#             inspection_record_query = inspection_records(
#                 material_type_id = 0,
#                 employee_id = employee_id,
#             )
#             session.add(inspection_record_query)
#             session.commit()
#             inspection_record_id = inspection_record_query.id
#         else:

#             # get the inspection record details
#             inspection_record_query = session.query(parts.item, parts.drawing)\
#                 .join(inspections, (inspections.part_id == parts.id))\
#                 .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#                 .filter(inspection_records.id == inspection_record_id).first()
#             if inspection_record_query is None:
#                 session.close()
#                 return {
#                     "status": "alert",
#                     "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspections, inspection_records])
#                 }
#             inspection_item, inspection_drawing = inspection_record_query

#             # make sure the inspection record and part match up
#             if inspection_item != part_item and inspection_drawing != part_drawing:
#                 session.close()
#                 return {
#                     "status": "alert",
#                     "response": text_response(stack()[0][3], message_type.generic, err_msg = "the provided inspection report does not match the provided part")
#                 }

#         # calculate the part index if the provided part_index is -1
#         if part_index == -1:
#             part_index_arr = session.query(inspections.part_index)\
#                 .filter(inspections.part_id == part_id)\
#                 .filter(inspections.inspection_record_id == inspection_record_id).all()
#             part_index = max([x[0] for x in part_index_arr]) + 1

#         # create a new inspection
#         inspection_query = inspections(
#             part_index = part_index,
#             datetime_measured = datetime.datetime.now(),
#             inspection_record_id = inspection_record_id,
#             part_id = part_id,
#             employee_id = employee_id,
#             inspection_type_id = 0,
#             disposition_id = 2
#         )
#         session.add(inspection_query)
#         session.commit()
#         inspection_id = inspection_query.id

#         # get the schema details
#         schema_details = session.query(*schema_details_columns)\
#             .filter(inspection_schema_details.schema_id == schema_id)\
#             .order_by(inspection_schema_details.name.asc()).all()
#         schema_details_list = []
#         for name, nominal, usl, lsl, precision, spectype, dimetype, freqtype, gauge_type_id in schema_details:

#             gauge_id = session.query(gauges.id)\
#                 .filter(gauges.gauge_type_id == gauge_type_id)\
#                 .order_by(gauges.name.asc())\
#                 .first()[0]

#             schema_details_list.append({
#                 "name": name,
#                 "nominal": nominal,
#                 "usl": usl,
#                 "lsl": lsl,
#                 "precision": precision,
#                 "specification_type_id": spectype,
#                 "dimension_type_id": dimetype,
#                 "frequency_type_id": freqtype,
#                 "gauge_id": gauge_id
#             })

#         # create the feature records
#         for obj in schema_details_list:
#             features_query = features(
#                 name = obj["name"],
#                 nominal = obj["nominal"],
#                 usl = obj["usl"],
#                 lsl = obj["lsl"],
#                 precision = obj["precision"],
#                 inspection_id = inspection_id,
#                 specification_type_id = obj["specification_type_id"],
#                 dimension_type_id = obj["dimension_type_id"],
#                 frequency_type_id = obj["frequency_type_id"],
#                 gauge_id = obj["gauge_id"]
#             )
#             session.add(features_query)
#         session.commit()

#         # close the database session
#         session.close()

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# def func_inspections_get_filtered_inspections(inspection_record_id:int, started_after:datetime, finished_before:datetime, employee_filter:str, part_index_filter:int, revision_filter:str, inspection_type_filter:int, disposition_type_filter:int, shown_job_numbers:list, shown_purchase_orders:list, jn_active:bool, po_active:bool):

#     # define the output columns
#     columns = [
#         inspections.id,
#         inspections.datetime_measured,
#         inspections.part_index,
#         inspections.employee_id,
#         inspections.inspection_record_id,
#         inspections.inspection_type_id,
#         inspections.disposition_id,
#         parts.item,
#         parts.drawing,
#         parts.revision
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the list of matching inspection schemas
#         inspections_query = session.query(*columns)\
#             .join(parts, (parts.id == inspections.part_id))\
#             .join(employees, (employees.id == inspections.employee_id))\
#             .outerjoin(inspections_job_numbers, (inspections_job_numbers.inspection_id == inspections.id))\
#             .outerjoin(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .outerjoin(inspections_purchase_orders, (inspections_purchase_orders.inspection_record_id == inspection_records.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(parts.revision.ilike(f"%{revision_filter}%"))\
#             .filter(and_(inspections.datetime_measured >= started_after, inspections.datetime_measured <= finished_before))\
#             .filter(or_(employees.first_name.ilike(f"%{employee_filter}%"), employees.last_name.ilike(f"%{employee_filter}%")))

#         # additional filters
#         if jn_active:
#             inspections_query = inspections_query.filter(inspections_job_numbers.job_number_id.in_(shown_job_numbers))
#         if po_active:
#             inspections_query = inspections_query.filter(inspections_purchase_orders.purchase_order_id.in_(shown_purchase_orders))
#         if part_index_filter > -1:
#             inspections_query = inspections_query.filter(inspections.part_index == part_index_filter)
#         if inspection_type_filter > -1:
#             inspections_query = inspections_query.filter(inspections.inspection_type_id == inspection_type_filter)
#         if disposition_type_filter > -1:
#             inspections_query = inspections_query.filter(inspections.disposition_id == disposition_type_filter)

#         # convert to list of tuples
#         inspections_query = inspections_query\
#             .order_by(inspections.part_index.asc(), parts.revision.asc(), inspections.datetime_measured.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(inspections_query) > 0:
#             output_arr = []
#             for inspection_id, timestamp, part_index, employee_id, current_inspection_record_id, inspection_type_id, disposition_type_id, item, drawing, revision in inspections_query:
#                 output_arr.append({
#                     "inspection_id": inspection_id,
#                     "timestamp": timestamp.strftime("%Y-%m-%dT%H:%M"),
#                     "part_index": part_index,
#                     "employee_id": employee_id,
#                     "inspection_record_id": current_inspection_record_id,
#                     "inspection_type_id": inspection_type_id,
#                     "disposition_type_id": disposition_type_id,
#                     "item": item,
#                     "drawing": drawing,
#                     "revision": revision.upper()
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection records - features

# @app.route("/inspection_records/features/get_filtered_features/", methods = ["POST"])
# def inspection_records_features_get_filtered_features():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["identity"]["inspection_record_id"])
#     item = str(form_data["identity"]["item"])
#     drawing = str(form_data["identity"]["drawing"])
#     name = str(form_data["content"]["name"])
#     frequency_type_id = int(form_data["content"]["frequency_type_id"])
#     has_deviations = int(form_data["content"]["has_deviations"])
#     inspector_id = int(form_data["content"]["employee_id"])
#     gauge_id = int(form_data["content"]["gauge_id"])
#     gauge_type_id = int(form_data["content"]["gauge_type_id"])
#     specification_type_id = int(form_data["content"]["specification_type_id"])
#     dimension_type_id = int(form_data["content"]["dimension_type_id"])
#     inspection_type_id = int(form_data["content"]["inspection_type_id"])
#     revision = str(form_data["content"]["revision"])
#     part_index = int(form_data["content"]["part_index"])

#     # make a list of acceptable inspection ids
#     inspection_ids = []
#     for obj in list(form_data["content"]["inspections"]):
#         if bool(int(obj["display_state"])):
#             inspection_ids.append(int(obj["inspection_id"]))

#     # run the required function
#     return func_features_get_filtered_features(
#         inspection_record_id,
#         item,
#         drawing,
#         frequency_type_id,
#         name,
#         has_deviations,
#         inspector_id,
#         gauge_id,
#         gauge_type_id,
#         specification_type_id,
#         dimension_type_id,
#         inspection_type_id,
#         revision,
#         part_index,
#         inspection_ids
#     )

# @app.route("/inspection_records/features/save_features/", methods = ["POST"])
# def inspection_records_features_save_features():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # extract the required information
#     data = form_data["data"]

#     try:

#         if len(data) > 0:

#             # convert the raw data
#             my_data = {}
#             for row in data:

#                 # make sure we're only pulling info we want
#                 if row["column"]["key"] == "measured" or row["column"]["key"] == "gauge_id":

#                     # get the feature id
#                     meas_id = row["row"]["feature_id"]

#                     # check if the feature id already exists in dictionary
#                     if meas_id in my_data:
#                         my_data[meas_id].append({ row["column"]["key"]: row["row"]["value"] })
#                     else:
#                         my_data[meas_id] = [{ row["column"]["key"]: row["row"]["value"] }]

#             # open the database session
#             session = Session(engine)

#             # iterate through the data object
#             rows_affected = 0
#             for k, v in my_data.items():
#                 results = session.query(features).filter(features.id == k)
#                 is_affected = 0
#                 for obj in v:
#                     is_affected = results.update(obj)
#                 if is_affected > 0:
#                     rows_affected += 1

#             # commit the changes
#             session.commit()

#             # close the database session
#             session.close()

#             # return the results
#             if rows_affected > 0:
#                 return {
#                     "status": "alert",
#                     "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [features])
#                 }
#             else:
#                 return {
#                     "status": "alert",
#                     "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [features])
#                 }
#         else:
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.generic, err_msg = "no data passed to flask server")
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/features/tunnel_to_physical_part/", methods = ["POST"])
# def inspection_records_features_tunnel_to_physical_part():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     part_id = int(form_data["part_id"])
#     part_index = int(form_data["part_index"])

#     try:

#         # open the database connection
#         session = Session(engine)

#         # get the part details
#         part_query = session.query(parts.item, parts.drawing, parts.revision)\
#             .filter(parts.id == part_id).first()
#         if part_query is None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts])
#             }
#         item, drawing, revision = part_query

#         # close the database connection
#         session.close()

#         # run the required function
#         return func_features_get_filtered_features(
#             inspection_record_id,
#             item,
#             drawing,
#             -1,
#             "",
#             -1,
#             -1,
#             -1,
#             -1,
#             -1,
#             -1,
#             -1,
#             revision,
#             part_index,
#             None
#         )

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_reports/features/get_filter_parameters/", methods = ["POST"])
# def inspection_records_features_get_filter_parameters():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])
#     item = form_data["item"]
#     drawing = form_data["drawing"]

#     try:

#         # open the database session
#         session = Session(engine)

#         # measurement type list
#         inspection_types_query = session.query(inspection_types.id, inspection_types.name)\
#             .join(inspections, (inspections.inspection_type_id == inspection_types.id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"%{drawing}")))\
#             .order_by(inspection_types.name.asc()).distinct(inspection_types.name).all()

#         # part index list
#         part_index_query = session.query(inspections.part_index)\
#             .join(parts, (parts.id == inspections.part_id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"%{drawing}%")))\
#             .order_by(inspections.part_index.asc()).distinct(inspections.part_index).all()

#         # revisions list
#         revisions_query = session.query(parts.revision)\
#             .join(inspections, (parts.id == inspections.part_id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"%{drawing}%")))\
#             .order_by(parts.revision.asc()).distinct(parts.revision).all()

#         # frequency type list
#         frequency_types_query = session.query(frequency_types.id, frequency_types.name)\
#             .join(features, (features.frequency_type_id == frequency_types.id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"{drawing}%")))\
#             .order_by(frequency_types.name.asc()).distinct(frequency_types.name).all()

#         # inspectors list
#         inspectors_query = session.query(employees.id, employees.first_name, employees.last_name)\
#             .join(inspections, (inspections.employee_id == employees.id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"{drawing}%")))\
#             .order_by(employees.last_name.asc(), employees.first_name.asc()).distinct(employees.first_name, employees.last_name).all()

#         # gauges list
#         gauges_query = session.query(gauges.id, gauges.name)\
#             .join(features, (features.gauge_id == gauges.id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"{drawing}%")))\
#             .order_by(gauges.name.asc()).distinct(gauges.name).all()

#         # gauges types list
#         gauge_types_query = session.query(gauge_types.id, gauge_types.name)\
#             .join(gauges, (gauges.gauge_type_id == gauge_types.id))\
#             .join(features, (features.gauge_id == gauges.id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"{drawing}%")))\
#             .order_by(gauge_types.name.asc()).distinct(gauge_types.name).all()

#         # specification types list
#         specification_types_query = session.query(specification_types.id, specification_types.name)\
#             .join(features, (features.specification_type_id == specification_types.id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"{drawing}%")))\
#             .order_by(specification_types.name.asc()).distinct(specification_types.name).all()

#         # dimension types list
#         dimension_types_query = session.query(dimension_types.id, dimension_types.name)\
#             .join(features, (features.dimension_type_id == dimension_types.id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .filter(inspections.inspection_record_id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"{drawing}%")))\
#             .order_by(dimension_types.name.asc()).distinct(dimension_types.name).all()

#         # close the database session
#         session.close()

#         # return the results
#         inspection_types_list = []
#         for id, name in inspection_types_query:
#             inspection_types_list.append({
#                 "id": id,
#                 "name": name
#             })

#         part_index_list = []
#         for index in part_index_query:
#             part_index_list.append({
#                 "id": index[0],
#                 "name": index[0]
#             })

#         revisions_list = []
#         for revision in revisions_query:
#             revisions_list.append({
#                 "id": revision[0],
#                 "name": revision[0].upper()
#             })

#         frequency_types_list = []
#         for id, name in frequency_types_query:
#             frequency_types_list.append({
#                 "id": id,
#                 "name": name
#             })

#         inspectors_list = []
#         for id, first_name, last_name in inspectors_query:
#             inspectors_list.append({
#                 "id": id,
#                 "name": f"{last_name}, {first_name}"
#             })

#         gauges_list = []
#         for id, name in gauges_query:
#             gauges_list.append({
#                 "id": id,
#                 "name": name
#             })

#         gauge_types_list = []
#         for id, name in gauge_types_query:
#             gauge_types_list.append({
#                 "id": id,
#                 "name": name
#             })

#         specification_types_list = []
#         for id, name in specification_types_query:
#             specification_types_list.append({
#                 "id": id,
#                 "name": name
#             })

#         dimension_types_list = []
#         for id, name in dimension_types_query:
#             dimension_types_list.append({
#                 "id": id,
#                 "name": name
#             })

#         # return the data object
#         return {
#             "status": "ok",
#             "response": {
#                 "inspection_types": inspection_types_list,
#                 "frequency_types": frequency_types_list,
#                 "inspectors": inspectors_list,
#                 "gauges": gauges_list,
#                 "gauge_types": gauge_types_list,
#                 "specification_types": specification_types_list,
#                 "dimension_types": dimension_types_list,
#                 "revisions": revisions_list,
#                 "part_indices": part_index_list
#             }
#         }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# # recycled methods

# def func_features_get_filtered_features(inspection_record_id:int, item:str, drawing:str, frequency_type_id:int, name:str, has_deviations:int, inspector_id:int, gauge_id:int, gauge_type_id:int, specification_type_id:int, dimension_type_id:int, inspection_type_id:int, revision:str, part_index:int, inspection_ids:list):

#     # define the columns
#     columns = [
#         inspections.id,
#         inspections.part_index,
#         inspections.datetime_measured,
#         inspections.employee_id,
#         parts.id,
#         parts.revision,
#         features.id,
#         features.name,
#         features.nominal,
#         features.usl,
#         features.lsl,
#         features.measured,
#         features.precision,
#         gauges.id,
#         gauge_types.id,
#         gauge_types.name,
#         specification_types.name,
#         dimension_types.name,
#         frequency_types.name,
#         inspection_types.name
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(*columns)\
#             .join(gauges, (features.gauge_id == gauges.id))\
#             .join(gauge_types, (gauges.gauge_type_id == gauge_types.id))\
#             .join(inspections, (inspections.id == features.inspection_id))\
#             .join(parts, (inspections.part_id == parts.id))\
#             .join(inspection_records, (inspections.inspection_record_id == inspection_records.id))\
#             .join(specification_types, (features.specification_type_id == specification_types.id))\
#             .join(dimension_types, (features.dimension_type_id == dimension_types.id))\
#             .join(frequency_types, (features.frequency_type_id == frequency_types.id))\
#             .join(inspection_types, (inspections.inspection_type_id == inspection_types.id))\
#             .filter(inspection_records.id == inspection_record_id)\
#             .filter(and_(parts.item.ilike(f"%{item}%"), parts.drawing.ilike(f"%{drawing}%")))\
#             .filter(features.name.ilike(f"%{name}%"))\
#             .filter(parts.revision.ilike(f"%{revision}%"))

#         if part_index > -1:
#             results = results.filter(inspections.part_index == part_index)
#         if dimension_type_id > -1:
#             results = results.filter(dimension_types.id == dimension_type_id)
#         if frequency_type_id > -1:
#             results = results.filter(features.frequency_type_id == frequency_type_id)
#         if has_deviations == 0:
#             results = results.filter(features.id.notin_(session.query(deviations.feature_id)))
#         elif has_deviations == 1:
#             results = results.filter(features.id.in_(session.query(deviations.feature_id)))
#         if inspector_id > -1:
#             results = results.filter(inspections.employee_id == inspector_id)
#         if gauge_id > -1:
#             results = results.filter(features.gauge_id == gauge_id)
#         if gauge_type_id > -1:
#             results = results.filter(gauges.gauge_type_id == gauge_type_id)
#         if specification_type_id > -1:
#             results = results.filter(features.specification_type_id == specification_type_id)
#         if inspection_type_id > -1:
#             results = results.filter(inspections.inspection_type_id == inspection_type_id)
#         if inspection_ids is not None:
#             results = results.filter(features.inspection_id.in_(inspection_ids))

#         # convert to a list
#         measurement_list = results\
#             .order_by(inspections.id.asc(), inspections.part_id.asc(), parts.revision.asc(), features.id.asc()).all()

#         # get the list of features that have deviations
#         deviations_list = [{ "feature_id": x[0], "deviation_type_id": x[1] } for x in session.query(deviations.feature_id, deviations.deviation_type_id).all()]
#         temp_deviations = [x["feature_id"] for x in deviations_list if x["deviation_type_id"] == 0]
#         perm_deviations = [x["feature_id"] for x in deviations_list if x["deviation_type_id"] == 1]

#         # close the database session
#         session.close()

#         # return the results
#         if len(measurement_list) > 0:

#             # assemble measurements output
#             output_arr = []
#             for inspection_id, part_index, timestamp, employee_id, part_id, revision, feature_id, name, nominal, usl, lsl, measured, precision, gauge_id, gauge_type_id, gauge_type, specification_type, dimension_type, frequency_type, inspection_type in measurement_list:

#                 # determine data input type (numerical or boolean)
#                 input_type = "numerical"
#                 if gauge_type_id == 6 or gauge_type_id == 10 or gauge_type_id == 11 or gauge_type_id == 13:
#                     input_type = "boolean"

#                 # parse to floats
#                 nominal_flt = float(nominal)
#                 usl_flt = float(usl)
#                 lsl_flt = float(lsl)

#                 # evaluate measurements
#                 state = "incomplete"
#                 measured_flt = 0
#                 if measured is None:
#                     measured_flt = None
#                 else:
#                     measured_flt = float(measured)
#                     if usl_flt >= measured_flt and lsl_flt <= measured_flt:
#                         state = "pass"
#                     else:
#                         state = "fail"

#                 # inspection for deviation flag
#                 if (feature_id in temp_deviations) and (feature_id not in perm_deviations):
#                     deviation_code = "*"
#                 elif (feature_id not in temp_deviations) and (feature_id in perm_deviations):
#                     deviation_code = "**"
#                 elif (feature_id in temp_deviations) and (feature_id in perm_deviations):
#                     deviation_code = "***"
#                 else:
#                     deviation_code = ""

#                 output_arr.append({
#                     "inspection_record_id": inspection_record_id,
#                     "part_id": part_id,
#                     "item": item,
#                     "drawing": drawing,
#                     "timestamp": timestamp.strftime("%Y-%m-%d, %H:%M"),
#                     "deviation_code": deviation_code,
#                     "feature_id": feature_id,
#                     "inspection_id": inspection_id,
#                     "part_index": part_index,
#                     "revision": revision.upper(),
#                     "name": name,
#                     "nominal": nominal_flt,
#                     "usl": usl_flt,
#                     "lsl": lsl_flt,
#                     "measured": measured_flt,
#                     "precision": precision,
#                     "employee_id": employee_id,
#                     "gauge_id": gauge_id,
#                     "gauge_type_id": gauge_type_id,
#                     "gauge_type": gauge_type,
#                     "specification_type": specification_type,
#                     "dimension_type": dimension_type,
#                     "inspection_type": inspection_type,
#                     "frequency_type": frequency_type,
#                     "state": state,
#                     "input_type": input_type
#                 })

#             # return the data object
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection records - manufactured

# # routes

# @app.route("/inspection_records/manufactured/add_associated_job_number/", methods = ["POST"])
# def inspection_records_manufactured_add_associated_job_number():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])
#     inspection_id = int(form_data["inspection_id"])
#     job_number_id = int(form_data["job_number_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # check if the association already exists
#         exists = session.query(inspections_job_numbers.id)\
#             .filter(inspections_job_numbers.inspection_id == inspection_id)\
#             .filter(inspections_job_numbers.job_number_id == job_number_id)\
#             .first()
#         if exists is not None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [inspections_job_numbers])
#             }

#         # add the association
#         session.add(inspections_job_numbers(**{
#             "inspection_id": inspection_id,
#             "job_number_id": job_number_id,
#         }))

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # reacquire the association list
#         return func_manufactured_get_associated_job_numbers(inspection_record_id, search_term)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/manufactured/save_associated_job_numbers/", methods = ["POST"])
# def inspection_records_manufactured_save_associated_job_numbers():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     raw_data = list(form_data["data"])

#     # convert the raw data
#     clean_data = []
#     for x in raw_data:
#         clean_data.append({
#             "id": int(x["id"]),
#             "data": {
#                 "full_inspect_interval": int(x["full_inspect_interval"]),
#                 "released_qty": int(x["released_qty"]),
#                 "completed_qty": int(x["completed_qty"])
#             }
#         })

#     try:

#         # open the database session
#         session = Session(engine)

#         # update the database
#         rows_affected = 0
#         for x in clean_data:
#             results = session.query(job_numbers).filter(job_numbers.id == int(x["id"]))
#             is_affected = 0
#             for k, v in x["data"].items():
#                 is_affected += results.update({ k: v })
#             if is_affected > 0:
#                 rows_affected += 1

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return the results
#         if rows_affected > 0:
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [job_numbers])
#             }
#         else:
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [job_numbers])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/manufactured/delete_associated_job_number/", methods = ["POST"])
# def inspection_records_manufactured_delete_associated_job_number():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])
#     inspection_id = int(form_data["inspection_id"])
#     job_number_id = int(form_data["job_number_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # delete the associated record
#         rows_deleted = session.query(inspections_job_numbers)\
#             .filter(inspections_job_numbers.inspection_id == inspection_id)\
#             .filter(inspections_job_numbers.job_number_id == job_number_id)\
#             .delete()

#         # commit the changes
#         if rows_deleted > 0:
#             session.commit()

#         # close the database session
#         session.close()

#         # reacquire the association list
#         return func_manufactured_get_associated_job_numbers(inspection_record_id, search_term)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/manufactured/get_associated_job_numbers/", methods = ["POST"])
# def inspection_records_manufactured_get_associated_job_numbers():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])

#     # return the results
#     return func_manufactured_get_associated_job_numbers(inspection_record_id, search_term)

# @app.route("/inspection_records/manufactured/get_filtered_job_numbers/", methods = ["POST"])
# def inspection_records_manufactured_get_filtered_job_numbers():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(job_numbers.id, job_numbers.name)\
#             .filter(job_numbers.name.ilike(f"%{search_term}%"))\
#             .order_by(job_numbers.name.asc()).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [job_numbers])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/manufactured/get_filtered_inspections/", methods = ["POST"])
# def inspection_records_manufactured_get_filtered_inspections():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     inspection_record_id = int(form_data["inspection_record_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(inspections.id, inspections.part_index, parts.revision, inspections.datetime_measured)\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .join(parts, (parts.id == inspections.part_id))\
#             .filter(inspection_records.id == inspection_record_id)\
#             .order_by(inspections.datetime_measured.asc(), inspections.part_index.asc(), parts.revision.asc(), inspections.id.asc())\
#             .all()
#         # results = session.query(parts.id, parts.item, parts.drawing, parts.revision)\
#         #     .join(inspections, (inspections.part_id == parts.id))\
#         #     .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#         #     .filter(or_(parts.item.ilike(f"%{search_term}%"), parts.drawing.ilike(f"%{search_term}%"), parts.revision.ilike(f"%{search_term}%")))\
#         #     .filter(inspection_records.id == inspection_record_id)\
#         #     .order_by(parts.drawing.asc(), parts.revision.asc(), parts.item.asc())\
#         #     .distinct(parts.drawing, parts.revision, parts.item).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, part_index, revision, datetime_measured in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": f"{part_index}, {revision.upper()}, {datetime_measured.strftime('%Y-%m-%d, %H:%M')}"
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_found, tables = [parts, inspections, inspection_records])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# # recycled methods

# def func_manufactured_get_associated_job_numbers(inspection_record_id:int, search_term:str):

#     # define the output columns
#     columns = [
#         job_numbers.id,
#         job_numbers.name,
#         job_numbers.full_inspect_interval,
#         job_numbers.released_qty,
#         job_numbers.completed_qty,
#         parts.revision,
#         parts.id,
#         inspections.id
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(*columns)\
#             .join(inspections_job_numbers, (inspections_job_numbers.job_number_id == job_numbers.id))\
#             .join(inspections, (inspections.id == inspections_job_numbers.inspection_id))\
#             .join(inspection_records, (inspection_records.id == inspections.inspection_record_id))\
#             .join(parts, (parts.id == inspections.part_id))\
#             .filter(inspection_records.id == inspection_record_id)\
#             .filter(or_(parts.revision.ilike(f"%{search_term}%"), job_numbers.name.ilike(f"%{search_term}%")))\
#             .order_by(job_numbers.name.asc())\
#             .distinct(job_numbers.name).all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name, full_inspect_interval, released_qty, completed_qty, revision, part_id, inspection_id in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "full_inspect_interval": full_inspect_interval,
#                     "released_qty": released_qty,
#                     "completed_qty": completed_qty,
#                     "revision": revision.upper(),
#                     "part_id": part_id,
#                     "inspection_id": inspection_id
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection records - received

# # routes

# @app.route("/inspection_records/received/assign_purchase_order_association/", methods = ["POST"])
# def inspection_records_received_assign_purchase_order_asociation():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])
#     purchase_order_id = int(form_data["purchase_order_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # check if the association already exists
#         exists = session.query(inspections_purchase_orders.id)\
#             .filter(inspections_purchase_orders.inspection_record_id == inspection_record_id)\
#             .filter(inspections_purchase_orders.purchase_order_id == purchase_order_id)\
#             .first()
#         if exists is not None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [inspections_purchase_orders])
#             }
        
#         # add the association
#         session.add(inspections_purchase_orders(**{
#             "inspection_record_id": inspection_record_id,
#             "purchase_order_id": purchase_order_id
#         }))

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return an updated association list
#         return func_received_get_filtered_purchase_order_associations(search_term, inspection_record_id)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/assign_receiver_number_association/", methods = ["POST"])
# def inspection_records_received_assign_receiver_number_association():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     receiver_number_id = int(form_data["receiver_number_id"])
#     purchase_order_id = int(form_data["purchase_order_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # check if the association already exists
#         exists = session.query(receiver_numbers.id)\
#             .filter(receiver_numbers.id == receiver_number_id)\
#             .filter(receiver_numbers.purchase_order_id == purchase_order_id)\
#             .first()
#         if exists is not None:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [receiver_numbers])
#             }

#         # add the association
#         is_affected = session.query(receiver_numbers)\
#             .filter(receiver_numbers.id == receiver_number_id)\
#             .update({
#                 "purchase_order_id": purchase_order_id
#             })
#         if is_affected == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [receiver_numbers])
#             }

#         # commit the changes
#         session.commit()

#         # close the database session
#         session.close()

#         # return an updated association list
#         return func_received_get_filtered_child_associations(search_term, purchase_order_id)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/set_associated_supplier/", methods = ["POST"])
# def inspection_records_received_set_associated_supplier():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     purchase_order_id = int(form_data["purchase_order_id"])
#     supplier_id = int(form_data["supplier_id"])

#     # account for null condition
#     if supplier_id == -1:
#         supplier_id = None

#     try:

#         # open the database session
#         session = Session(engine)

#         # assign the association
#         is_affected = session.query(purchase_orders)\
#             .filter(purchase_orders.id == purchase_order_id)\
#             .update({
#                 "supplier_id": supplier_id
#             })
#         if is_affected == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [purchase_orders])
#             }
#         else:
#             session.commit()
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = is_affected, tables = [purchase_orders])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/save_received_quantities/", methods = ["POST"])
# def inspection_records_received_save_received_quantities():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     raw_data = list(form_data["data"])

#     # convert the raw data
#     clean_data = []
#     for x in raw_data:
#         clean_data.append({
#             "id": int(x["id"]),
#             "data": {
#                 "received_qty": int(x["received_qty"])
#             }
#         })

#     try:

#         # open the database session
#         session = Session(engine)

#         # update the database
#         rows_affected = 0
#         for x in clean_data:
#             results = session.query(receiver_numbers).filter(receiver_numbers.id == int(x["id"]))
#             is_affected = 0
#             for k, v in x["data"].items():
#                 is_affected += results.update({ k: v })
#             if is_affected > 0:
#                 rows_affected += 1
#         if rows_affected == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [receiver_numbers])
#             }
#         else:
#             session.commit()
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [receiver_numbers])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/remove_purchase_order_association/", methods = ["POST"])
# def inspection_records_received_remove_purchase_order_association():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])
#     purchase_order_id = int(form_data["purchase_order_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # delete the records
#         rows_affected = session.query(inspections_purchase_orders)\
#             .filter(inspections_purchase_orders.inspection_record_id == inspection_record_id)\
#             .filter(inspections_purchase_orders.purchase_order_id == purchase_order_id)\
#             .delete()
#         if rows_affected == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_deleted, tables = [inspections_purchase_orders])
#             }
#         else:
#             session.commit()
#             session.close()

#         # return the updated list
#         return func_received_get_filtered_purchase_order_associations(search_term, inspection_record_id)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/remove_receiver_number_association/", methods = ["POST"])
# def inspection_records_received_remove_receiver_number_association():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     purchase_order_id = int(form_data["purchase_order_id"])
#     receiver_number_id = int(form_data["receiver_number_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # delete the records
#         rows_affected = session.query(receiver_numbers)\
#             .filter(receiver_numbers.id == receiver_number_id)\
#             .update({
#                 "purchase_order_id": None
#             })
#         if rows_affected == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [inspections_purchase_orders])
#             }
#         else:
#             session.commit()
#             session.close()

#         # return the updated list
#         return func_received_get_filtered_child_associations(search_term, purchase_order_id)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/get_filtered_purchase_order_associations/", methods = ["POST"])
# def inspection_records_received_get_filtered_purchase_order_associations():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])

#     # return the list of associations
#     return func_received_get_filtered_purchase_order_associations(search_term, inspection_record_id)

# @app.route("/inspection_records/received/get_filtered_child_associations/", methods = ["POST"])
# def inspection_records_received_get_filtered_child_associations():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     purchase_order_id = int(form_data["purchase_order_id"])

#     # return the found associations
#     return func_received_get_filtered_child_associations(search_term, purchase_order_id)

# @app.route("/inspection_records/received/get_filtered_purchase_order_options/", methods = ["POST"])
# def inspection_records_received_get_filtered_purchase_order_options():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(purchase_orders.id, purchase_orders.name, purchase_orders.supplier_id)\
#             .filter(purchase_orders.name.ilike(f"%{search_term}%"))\
#             .order_by(purchase_orders.name.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name, supplier_id in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "supplier_id": supplier_id
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/get_filtered_receiver_number_options/", methods = ["POST"])
# def inspection_records_received_get_filtered_receiver_number_options():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(receiver_numbers.id, receiver_numbers.name, receiver_numbers.received_qty, receiver_numbers.purchase_order_id)\
#             .filter(receiver_numbers.name.ilike(f"%{search_term}%"))\
#             .order_by(receiver_numbers.name.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name, received_qty, current_purchase_order_id in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "received_qty": received_qty,
#                     "purchase_order_id": current_purchase_order_id
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_records/received/get_filtered_supplier_options/", methods = ["POST"])
# def inspection_records_received_get_filtered_supplier_options():

#     # get the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(suppliers.id, suppliers.name)\
#             .filter(suppliers.name.ilike(f"%{search_term}%"))\
#             .order_by(suppliers.name.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# # recycled methods

# def func_received_get_filtered_purchase_order_associations(search_term:str, inspection_record_id:int):

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(purchase_orders.id, purchase_orders.name, inspections_purchase_orders.inspection_record_id)\
#             .join(purchase_orders, (purchase_orders.id == inspections_purchase_orders.purchase_order_id))\
#             .filter(purchase_orders.name.ilike(f"%{search_term}%"))\
#             .filter(inspections_purchase_orders.inspection_record_id == inspection_record_id)\
#             .order_by(purchase_orders.name.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for current_purchase_order_id, name, current_inspection_record_id in results:
#                 output_arr.append({
#                     "purchase_order_id": current_purchase_order_id,
#                     "name": name,
#                     "inspection_record_id": current_inspection_record_id
#                 })
#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# def func_received_get_filtered_child_associations(search_term:str, purchase_order_id:int):

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the supplier id
#         supplier_id = session.query(purchase_orders.supplier_id)\
#             .filter(purchase_orders.id == purchase_order_id)\
#             .first()[0]
        
#         # get the receiver numbers
#         receiver_numbers_query = session.query(receiver_numbers.id, receiver_numbers.name, receiver_numbers.received_qty, receiver_numbers.purchase_order_id)\
#             .filter(receiver_numbers.purchase_order_id == purchase_order_id)\
#             .filter(receiver_numbers.name.ilike(f"%{search_term}%"))\
#             .order_by(receiver_numbers.name.asc())\
#             .all()

#         # close the database session
#         session.close()

#         # return the results
#         if len(receiver_numbers_query) > 0:
#             output_arr = []
#             for id, name, received_qty, current_purchase_order_id in receiver_numbers_query:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "received_qty": received_qty,
#                     "purchase_order_id": current_purchase_order_id
#                 })
#             return {
#                 "status": "ok",
#                 "response": {
#                     "supplier_id": supplier_id,
#                     "receiver_numbers": output_arr
#                 }
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": {
#                     "supplier_id": supplier_id,
#                     "receiver_numbers": None
#                 }
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection records - lot numbers

# # routes

# @app.route("/inspection_records/lot_numbers/assign_lot_number/", methods = ["POST"])
# def inspection_records_lot_numbers_assign_lot_number():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])
#     lot_number_id = int(form_data["lot_number_id"])

#     # run the targeted method
#     return func_lot_numbers_assign(
#         search_term,
#         inspection_record_id,
#         lot_number_id
#     )

# @app.route("/inspection_records/lot_numbers/unassign_lot_number/", methods = ["POST"])
# def inspection_records_lot_numbers_unassign_lot_number():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])
#     lot_number_id = int(form_data["lot_number_id"])

#     # run the targeted method
#     return func_lot_numbers_association_remove(
#         search_term,
#         inspection_record_id,
#         lot_number_id
#     )

# @app.route("/inspection_records/lot_numbers/get_filtered_options/", methods = ["POST"])
# def inspection_records_lot_numbers_get_filtered_options():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])

#     return func_lot_numbers_get_filtered_options(
#         search_term
#     )

# @app.route("/inspection_records/lot_numbers/get_filtered_associations/", methods = ["POST"])
# def inspection_records_lot_numbers_get_filtered_associations():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     search_term = str(form_data["search_term"])
#     inspection_record_id = int(form_data["inspection_record_id"])

#     return func_lot_numbers_get_filtered_associations(
#         search_term,
#         inspection_record_id
#     )

# # recycled methods

# def func_lot_numbers_assign(search_term:str, inspection_record_id:int, lot_number_id:int):

#     try:

#         # open the database session
#         session = Session(engine)

#         # measurement_set if the association already exists
#         results = session.query(inspections_lot_numbers.inspection_record_id)\
#             .filter(inspections_lot_numbers.inspection_record_id == inspection_record_id)\
#             .filter(inspections_lot_numbers.lot_number_id == lot_number_id).all()

#         # logic gate
#         if len(results) > 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.record_already_exists, tables = [inspections_lot_numbers])
#             }

#         # add the new association
#         session.add(inspections_lot_numbers(**{
#             "inspection_record_id": inspection_record_id,
#             "lot_number_id": lot_number_id
#         }))
#         session.commit()

#         # close the session
#         session.close()

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

#     return func_lot_numbers_get_filtered_associations(
#         search_term,
#         inspection_record_id
#     )

# def func_lot_numbers_association_remove(search_term:str, inspection_record_id:int, lot_number_id:int):

#     try:

#         # open the database session
#         session = Session(engine)

#         # delete the record that matches the provided criteria
#         deleted_count = session.query(inspections_lot_numbers)\
#             .filter(inspections_lot_numbers.inspection_record_id == inspection_record_id)\
#             .filter(inspections_lot_numbers.lot_number_id == lot_number_id)\
#             .delete()

#         # logic gate
#         if deleted_count == 0:
#             session.close()
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_not_deleted, tables = [inspections_lot_numbers])
#             }
#         else:
#             session.commit()

#         # close the session
#         session.close()

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

#     return func_lot_numbers_get_filtered_associations(search_term, inspection_record_id)

# def func_lot_numbers_get_filtered_associations(search_term:str, inspection_record_id:int):

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(lot_numbers.id, lot_numbers.name)\
#             .join(inspections_lot_numbers, (lot_numbers.id == inspections_lot_numbers.lot_number_id))\
#             .join(inspection_records, (inspection_records.id == inspections_lot_numbers.inspection_record_id))\
#             .filter(inspection_records.id == inspection_record_id)\
#             .filter(lot_numbers.name.ilike(f"%{search_term}%"))\
#             .order_by(lot_numbers.name.asc()).all()

#         # close the session
#         session.close()

#         # return the results
#         arr_size = len(results)
#         if arr_size > 0:
#             output_arr = []
#             for id, name in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name,
#                     "inspection_record_id": inspection_record_id
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# def func_lot_numbers_get_filtered_options(search_term:str):

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(lot_numbers.id, lot_numbers.name)\
#             .filter(lot_numbers.name.ilike(f"%{search_term}%"))\
#             .order_by(lot_numbers.name.asc())\
#             .all()

#         # close the session
#         session.close()

#         # return the results
#         if len(results) > 0:
#             output_arr = []
#             for id, name in results:
#                 output_arr.append({
#                     "id": id,
#                     "name": name
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# #region inspection records - deviations

# # routes

# @app.route("/inspection_records/deviations/save_deviations/", methods = ["POST"])
# def inspection_records_deviations_save_deviations():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # extract the required information
#     feature_id = int(form_data["feature_id"])
#     data = list(form_data["data"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         rows_affected = 0
#         for row in data:
#             deviation_id = int(row["id"])
#             results = session.query(deviations).filter(deviations.id == deviation_id)

#             is_affected = results.update({
#                 "nominal": float(row["nominal"]),
#                 "usl": float(row["usl"]),
#                 "lsl": float(row["lsl"]),
#                 "precision": int(row["precision"]),
#                 "date_implemented": datetime.datetime.strptime(str(row["date_implemented"]), "%Y-%m-%d"),
#                 "notes": str(row["notes"]),
#                 "deviation_type_id": int(row["deviation_type_id"]),
#                 "employee_id": int(row["employee_id"]),
#                 "feature_id": feature_id
#             })

#             if is_affected > 0:
#                 rows_affected += 1

#         # commit the changes
#         session.commit()

#         # close the session
#         session.close()

#         # return the results
#         if rows_affected > 0:
#             return {
#                 "status": "alert",
#                 "response": text_response(stack()[0][3], message_type.records_updated, qty = rows_affected, tables = [deviations])
#             }
#         else:
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_updated, tables = [deviations])
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_reports/deviations/add_deviation/", methods = ["POST"])
# def inspection_reports_deviations_add_deviation():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     feature_id = int(form_data["feature_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # get the inspection's employee
#         employee_id = session.query(inspections.employee_id)\
#             .join(features, (features.inspection_id == inspections.id))\
#             .filter(features.id == feature_id)\
#             .first()[0]

#         # add the placeholder data to the database
#         new_record = deviations(
#             nominal = 1,
#             usl = 1.1,
#             lsl = 0.9,
#             precision = 1,
#             date_implemented = datetime.datetime.now(),
#             notes = "none",
#             deviation_type_id = 0,
#             employee_id = employee_id,
#             feature_id = feature_id
#         )
#         session.add(new_record)
#         session.commit()

#         # close the session
#         session.close()

#         # return the new deviation data
#         return func_deviations_get_feature_deviations(feature_id)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_reports/deviations/delete_deviation/", methods = ["POST"])
# def inspection_reports_deviations_delete_deviation():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     feature_id = int(form_data["feature_id"])
#     deviation_id = int(form_data["deviation_id"])

#     try:

#         # open the database session
#         session = Session(engine)

#         # remove the deviation
#         rows_deleted = session.query(deviations)\
#             .filter(deviations.id == deviation_id)\
#             .delete()
#         if rows_deleted == 0:
#             session.close()
#             return {
#                 "status": "log",
#                 "response": text_response(stack()[0][3], message_type.records_not_deleted, tables = [deviations])
#             }

#         # commit the changes
#         session.commit()

#         # close the session
#         session.close()

#         # return the updated deviations
#         return func_deviations_get_feature_deviations(feature_id)

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# @app.route("/inspection_reports/deviations/get_feature_deviations/", methods = ["POST"])
# def inspection_reports_deviations_get_feature_deviations():

#     # interpret the posted data
#     form_data = json.loads(request.data)

#     # get the required parameters
#     feature_id = int(form_data["feature_id"])

#     # return the results
#     return func_deviations_get_feature_deviations(feature_id)

# # recycled methods

# def func_deviations_get_feature_deviations(feature_id:int):

#     # define the columns
#     columns = [
#         deviations.id,
#         deviations.nominal,
#         deviations.usl,
#         deviations.lsl,
#         deviations.precision,
#         deviations.date_implemented,
#         deviations.notes,
#         deviation_types.id,
#         employees.id
#     ]

#     try:

#         # open the database session
#         session = Session(engine)

#         # query the database
#         results = session.query(*columns)\
#             .join(employees, (employees.id == deviations.employee_id))\
#             .join(deviation_types, (deviation_types.id == deviations.deviation_type_id))\
#             .filter(deviations.feature_id == feature_id)\
#             .order_by(deviations.id.asc())\
#             .distinct(deviations.id).all()

#         # close the session
#         session.close()

#         # return the results
#         if len(results) > 0:

#             output_arr = []
#             for id, nominal, usl, lsl, precision, date_implemented, notes, deviation_type_id, employee_id in results:

#                 # parse decimal to float
#                 nominal_flt = round(float(nominal), precision)
#                 usl_flt = round(float(usl), precision)
#                 lsl_flt = round(float(lsl), precision)

#                 # parse date to string
#                 date_implemented_str = date_implemented.strftime("%Y-%m-%d")

#                 output_arr.append({
#                     "id": id,
#                     "nominal": nominal_flt,
#                     "usl": usl_flt,
#                     "lsl": lsl_flt,
#                     "precision": precision,
#                     "date_implemented": date_implemented_str,
#                     "notes": notes,
#                     "deviation_type_id": deviation_type_id,
#                     "employee_id": employee_id,
#                     "feature_id": feature_id
#                 })

#             return {
#                 "status": "ok",
#                 "response": output_arr
#             }
#         else:
#             return {
#                 "status": "ok",
#                 "response": None
#             }

#     except SQLAlchemyError as e:
#         return {
#             "status": "log",
#             "response": text_response(stack()[0][3], message_type.sql_exception, error = e)
#         }

# #endregion

# --------------------------------------------------

# run the flask server
if __name__ == "__main__":
    app.run(host = "0.0.0.0", debug = True, port = 8000)