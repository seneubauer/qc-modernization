# import dependencies
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from xlrd import xldate_as_datetime
from os import listdir
from os.path import isfile, join
from pickle import dump, load
from random import sample
from numpy import NaN
import pandas as pd
import csv

# bring in config values
from sys import path
path.insert(0, "..")
from reference_values import alias_dict

# define a standard delimitor character
standard_delimitor = "%"
pretty_delimitor = "|"

# clean the metadata item number column values
def clean_item_number(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase and remove whitespace characters
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return a delimitable string or the item itself
        if standard_delimitor in item_str:
            return pretty_delimitor.join([x for x in item_str.split(standard_delimitor) if x is not None and len(x) > 0])
        else:
            return item_str
    else:
        return None

# clean the metadata drawing column values
def clean_drawing(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase and remove whitespace characters
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        return item_str
    else:
        return None

# clean the metadata revision column values
def clean_revision(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        return item_str
    else:
        return None

# clean the metadata inspection date column values
def clean_inspection_date(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the date object if it can be converted
        if "/" in item_str:
            return datetime.strptime(item_str, "%Y/%m/%d").date()
        elif "-" in item_str:
            return datetime.strptime(item_str, "%Y-%m-%d").date()
        else:
            return None
    else:
        return None

# clean the metadata inspector & operator column values
def clean_inspector_operator(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return a splitable string
        if standard_delimitor in item_str:
            arr = [str(x) for x in item_str.split(standard_delimitor) if x is not None and len(x) > 0]
            if not any([x.isnumeric() for x in arr]):
                return pretty_delimitor.join(arr)
            else:
                return None
        elif len(item_str) > 0:
            return item_str
        else:
            return None
    else:
        return None

# clean the metadata disposition column values
def clean_disposition(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        return item_str
    else:
        return None

# clean the metadata supplier column values
def clean_supplier(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        return item_str
    else:
        return None

# clean the metadata receiver number column values
def clean_receiver_number(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        if standard_delimitor in item_str:
            arr = [x for x in item_str.split(standard_delimitor) if x is not None and len(x) > 0]
            i = 1
            while i < len(arr):

                # make sure both items are numeric
                if arr[i - 1].isnumeric() and arr[i].isnumeric():

                    # make sure the preceding number contains more digits than the following number
                    if len(arr[i - 1]) > len(arr[i]):
                        arr[i] = f"{arr[i - 1][:len(arr[i - 1]) - len(arr[i])]}{arr[i]}"
                i += 1

            return pretty_delimitor.join(arr)
        else:
            return item_str
    else:
        return None

# clean the metadata purchase order column values
def clean_purchase_order(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # only permit values with 3 letters at the beginning
        if item_str[:3].isalpha():
            if standard_delimitor in item_str:
                arr = [x for x in item_str.split(standard_delimitor) if x is not None and len(x) > 1]
                i = 1
                while i < len(arr):
                    if arr[i - 1][-3:].isnumeric() and arr[i].isnumeric():
                        arr[i] = f"{arr[i - 1][:3]}{arr[i]}"
                    else:
                        return None
                    i += 1
                return "|".join(arr)
            else:
                return item_str
    else:
        return None

# clean the metadata job order column values
def clean_job_order(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        return item_str
    else:
        return None

# clean the metadata full inspect quantity column values
def clean_full_inspect_qty(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return NaN

        # reclassify characteristics as np.NaN if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return NaN

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # search for specific content
        if any(x in item_str for x in ["1st", "first", "f", "middle", "last", "l"]):

            # references 'first'
            f = 0b000
            if any(x in item_str for x in ["1st", "first", "f"]):
                f = 0b100

            # references 'middle'
            m = 0b000
            if any(x in item_str for x in ["middle"]):
                m = 0b010

            # references 'last'
            l = 0b000
            if any(x in item_str for x in ["last", "l"]):
                l = 0b001

            return -int(f | m | l)
        else:
            if item_str.isnumeric():
                return item_str
            else:
                return NaN
    else:
        return NaN

# clean the metadata received quantity column values
def clean_received_qty(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return NaN

        # reclassify characteristics as np.NaN if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return NaN

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        if item_str.isnumeric():
            return item_str
        else:
            return NaN
    else:
        return NaN

# clean the metadata completed quantity column values
def clean_completed_qty(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to np.NaN if it is NaN
        if item_str == "nan":
            return NaN

        # reclassify characteristics as np.NaN if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return NaN

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # return the remaining string
        if item_str.isnumeric():
            return item_str
        else:
            return NaN
    else:
        return NaN

# clean the measurement feature id column values
def clean_feature_id(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)
        
        # return the item if it is numeric
        if len(item_str) > 0:
            return item_str
        else:
            return None
    else:
        return None

# clean the measurement usl/lsl column values
def clean_speclimits(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to None if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)
        
        # return the item if it is numeric
        if item_str.replace(".", "").isnumeric():
            return item_str
        else:
            return None
    else:
        return None

# clean the measurement gauge column values
def clean_gauge(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to np.NaN if it is NaN
        if item_str == "nan":
            return NaN

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return NaN

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        if item_str.isnumeric():
            return NaN
        else:
            output = 0b000000000000000000000
            for k in alias_dict["gauges"]:
                if any(x in item_str for x in alias_dict["gauges"][k]["keys"]):

                    current_alias = alias_dict["gauges"][k]["alias"]
                    cal_alias = alias_dict["gauges"]["caliper"]["alias"]

                    # flip the appropriate bit
                    output |= current_alias

                    # make sure there is no caliper/idcaliper overlap
                    if (output & cal_alias == cal_alias) and (k == "id_caliper"):
                        output ^= cal_alias

            if output != 0b000000000000000000000:
                return int(output)
            else:
                return NaN
    else:
        return NaN

# clean the measurement data type column values
def clean_data_type(row: pd.Series, arg_dict: dict) -> pd.Series:
    if row is not None:

        # retrieve the argument lists
        none_if_contains = arg_dict["none_if_contains"]
        remove_substrings = arg_dict["remove_substrings"]
        replace_delimitors = arg_dict["replace_delimitors"]

        # enforce lowercase
        item_str = str(row).lower()

        # convert to np.NaN if it is NaN
        if item_str == "nan":
            return None

        # reclassify characteristics as None if they contain a certain substring
        if len(none_if_contains) > 0:
            if any(x in item_str for x in none_if_contains):
                return None

        # remove certain substrings from the characteristic
        if len(remove_substrings) > 0:
            for x in remove_substrings:
                if x in item_str:
                    item_str = item_str.replace(x, "")

        # standardize delimitor characters
        if len(replace_delimitors) > 0:
            for x in replace_delimitors:
                if x in item_str:
                    item_str = item_str.replace(x, standard_delimitor)

        # convert values
        if item_str == "str":
            return "pass_fail"
        elif item_str == "float" or item_str == "int":
            return "measured"
    else:
        return None

# extract the contents of one electronic inspection record into a list of dictionaries
def scrape_one(qc_folder: str, anchor_search_term: str, workbook_name: str, metadata_index: int, worksheet_names: list = []) -> tuple:

    # define the file path
    file_path = join(qc_folder, workbook_name)

    # open the workbook
    wb = load_workbook(filename=file_path, read_only = True, data_only = True)

    # initialize the output DataFrames
    metadata_df = pd.DataFrame()
    measurements_df = pd.DataFrame()

    # check if worksheet names were supplied
    if len(worksheet_names) == 0:
        worksheet_names = wb.worksheets

    # iterate through the specified worksheets
    for ws in worksheet_names:

        # find the data anchor
        anchor_column = 0
        anchor_row = 1
        anchor_found = False
        for i in range(1, 21):
            if ws[f"{get_column_letter(i)}1"].value == anchor_search_term:
                anchor_column = i
                anchor_found = True
                break

        # continue if this is the correct worksheet
        if anchor_found:

            # define search parameters
            search_column = get_column_letter(anchor_column + 2)
            search_row = anchor_row + 17
            index_limit = 200
            initial_i = anchor_row + 17
            initial_j = anchor_column + 2

            # define iterator parameters
            cell_value = "initial"
            i = initial_i

            # get the item count
            while cell_value is not None:

                # exit the iterator
                if cell_value is None:
                    break

                # exit condition
                cell_value = ws[f"{search_column}{i}"].value

                # increment the iterator
                i += 1

                # limit the while loop
                if i >= index_limit:
                    print(f"Row search has exceeded index limit ({index_limit})")
                    break

            # define the item count
            item_count = i - 1 - initial_i

            # continue if there are items in the worksheet
            if item_count > 0:

                # define iterator parameters
                feature_index = 12
                usl_index = 14
                lsl_index = 15
                gauge_index = 16
                cell_value = "initial"
                j = initial_j

                # initialize the storage lists
                data_types = []
                features = []
                usls = []
                lsls = []
                gauges = []
                measurements = []

                # search column by column
                while cell_value is not None:

                    # define the current column
                    column_index = get_column_letter(j)

                    # define the exit condition value
                    cell_value = ws[f"{column_index}{search_row}"].value

                    # exit the iterator
                    if cell_value is None:
                        break

                    # define metadata
                    data_type = str(type(cell_value)).replace("<class ", "").replace(">", "").replace("'", "").strip()
                    feature = ws[f"{column_index}{feature_index}"].value
                    usl = ws[f"{column_index}{usl_index}"].value
                    lsl = ws[f"{column_index}{lsl_index}"].value
                    gauge = ws[f"{column_index}{gauge_index}"].value

                    # only add data when the feature number is defined
                    if feature != 0:

                        data_types.append(data_type)
                        features.append(feature)
                        usls.append(usl)
                        lsls.append(lsl)
                        gauges.append(gauge)

                        for x in range(item_count):
                            measurements.append(ws[f"{column_index}{x + initial_i}"].value)

                    # increment the iterator
                    j += 1

                    # limit the while loop
                    if j >= index_limit:
                        print(f"Column search has exceeded index limit ({index_limit})")
                        break

                # convert pass/fail to numerical data
                for x in range(len(measurements)):
                    if type(measurements[x]) == str:
                        if measurements[x].lower() in alias_dict.keys():
                            measurements[x] = alias_dict[measurements[x].lower()]
                        else:
                            measurements[x] = alias_dict["empty"]

                # slice the raw measurement list into constituent parts
                meas_lists = []
                for x in range(item_count):
                    temp_list = []
                    for y in range(0, len(measurements), item_count):
                        temp_list.append(measurements[x + y])
                    meas_lists.append(temp_list)

                # section 1 metadata
                section_1 = []
                for x in range(21):
                    find_column = get_column_letter(anchor_column + x)
                    data_column = get_column_letter(anchor_column + x + 2)
                    if str(ws[f"{find_column}{anchor_row + 5}"].value).lower() == "item number":
                        # item number
                        value0 = ws[f"{data_column}{anchor_row + 5}"].value
                        # drawing number
                        value1 = ws[f"{data_column}{anchor_row + 6}"].value
                        # revision
                        value2 = ws[f"{data_column}{anchor_row + 7}"].value
                        if value0 is not None:
                            section_1.append(value0)
                        else:
                            section_1.append(alias_dict["empty"])
                        if value1 is not None:
                            section_1.append(value1)
                        else:
                            section_1.append(alias_dict["empty"])
                        if value2 is not None:
                            section_1.append(value2)
                        else:
                            section_1.append(alias_dict["empty"])
                        break

                # section 2 metadata
                section_2 = []
                for x in range(21):
                    find_column = get_column_letter(anchor_column + x)
                    data_column = get_column_letter(anchor_column + x + 1)
                    if str(ws[f"{find_column}{anchor_row + 5}"].value).lower() == "date":
                        # date
                        value0 = ws[f"{data_column}{anchor_row + 5}"].value
                        # inspector
                        value1 = ws[f"{data_column}{anchor_row + 6}"].value
                        # disposition
                        value2 = ws[f"{data_column}{anchor_row + 7}"].value
                        # supplier
                        value3 = ws[f"{data_column}{anchor_row + 8}"].value
                        if value0 is not None:
                            try:
                                if type(value0) is int:
                                    value0 = xldate_as_datetime(value0, 0)
                                section_2.append(date(value0.year, value0.month, value0.day))
                            except AttributeError:
                                print(f"AttributeError in {ws.title} of {workbook_name}; {value0}")
                        else:
                            section_2.append(alias_dict["empty"])
                        if value1 is not None:
                            section_2.append(value1)
                        else:
                            section_2.append(alias_dict["empty"])
                        if value2 is not None:
                            section_2.append(value2)
                        else:
                            section_2.append(alias_dict["empty"])
                        if value3 is not None:
                            section_2.append(value3)
                        else:
                            section_2.append(alias_dict["empty"])
                        break

                # section 3 metadata
                section_3 = []
                for x in range(21):
                    find_column = get_column_letter(anchor_column + x)
                    data_column = get_column_letter(anchor_column + x + 1)
                    if str(ws[f"{find_column}{anchor_row + 5}"].value).lower() == "recv. #":
                        # receiver number
                        value0 = ws[f"{data_column}{anchor_row + 5}"].value
                        # purchase order
                        value1 = ws[f"{data_column}{anchor_row + 6}"].value
                        # job number
                        value2 = ws[f"{data_column}{anchor_row + 7}"].value
                        # operator
                        value3 = ws[f"{data_column}{anchor_row + 8}"].value
                        if value0 is not None:
                            section_3.append(value0)
                        else:
                            section_3.append(alias_dict["empty"])
                        if value1 is not None:
                            section_3.append(value1)
                        else:
                            section_3.append(alias_dict["empty"])
                        if value2 is not None:
                            section_3.append(value2)
                        else:
                            section_3.append(alias_dict["empty"])
                        if value3 is not None:
                            section_3.append(value3)
                        else:
                            section_3.append(alias_dict["empty"])
                        break

                # section 4 metadata
                section_4 = []
                for x in range(21):
                    find_column = get_column_letter(anchor_column + x)
                    data_column = get_column_letter(anchor_column + x + 2)
                    if str(ws[f"{find_column}{anchor_row + 6}"].value).lower() == "qc full inspect interval":
                        # qc full inspect quantity
                        value0 = ws[f"{data_column}{anchor_row + 6}"].value
                        # released quantity
                        value1 = ws[f"{data_column}{anchor_row + 7}"].value
                        # completed quantity
                        value2 = ws[f"{data_column}{anchor_row + 8}"].value
                        if value0 is not None:
                            section_4.append(value0)
                        else:
                            section_4.append(alias_dict["empty"])
                        if value1 is not None:
                            section_4.append(value1)
                        else:
                            section_4.append(alias_dict["empty"])
                        if value2 is not None:
                            section_4.append(value2)
                        else:
                            section_4.append(alias_dict["empty"])
                        break

                # ensure the section lists have contents
                if len(section_1) == 0:
                    section_1 = [alias_dict["empty"],
                                 alias_dict["empty"], alias_dict["empty"]]
                if len(section_2) == 0:
                    section_2 = [alias_dict["empty"], alias_dict["empty"],
                                 alias_dict["empty"], alias_dict["empty"]]
                if len(section_3) == 0:
                    section_3 = [alias_dict["empty"], alias_dict["empty"],
                                 alias_dict["empty"], alias_dict["empty"]]
                if len(section_4) == 0:
                    section_4 = [alias_dict["empty"],
                                 alias_dict["empty"], alias_dict["empty"]]

                try:
                    # store metadata in a DataFrame
                    current_metadata_df = pd.DataFrame({
                        "id": [metadata_index],
                        "item_number": section_1[0],
                        "drawing": section_1[1],
                        "revision": section_1[2],
                        "inspection_date": section_2[0],
                        "inspector": section_2[1],
                        "disposition": section_2[2],
                        "supplier": section_2[3],
                        "receiver_number": section_3[0],
                        "purchase_order": section_3[1],
                        "job_order": section_3[2],
                        "operator": section_3[3],
                        "full_inspect_qty": section_4[0],
                        "received_qty": section_4[1],
                        "completed_qty": section_4[2]
                    })

                    # store measurement data in a DataFrame
                    current_measurements_df = pd.DataFrame({
                        "metadata_id": [metadata_index for i in range(len(features))],
                        "feature_id": features,
                        "usl": usls,
                        "lsl": lsls,
                        "gauge": gauges,
                        "data_type": data_types
                    })

                    # advance the index
                    metadata_index += 1

                    # add measurement data to measurements_df
                    x = 0
                    for meas_list in meas_lists:
                        current_measurements_df[f"part_{x}"] = meas_list
                        x += 1

                    metadata_df = pd.concat(
                        [metadata_df, current_metadata_df], axis=0, ignore_index=True)
                    measurements_df = pd.concat(
                        [measurements_df, current_measurements_df], axis=0, ignore_index=True)

                except IndexError:
                    print(f"IndexError in {ws.title} of {workbook_name}")

    # close the workbook
    wb.close()

    # return the results
    return (metadata_df, measurements_df)

# clean the metadata dataframe
def clean_metadata(raw_df: pd.DataFrame) -> pd.DataFrame:

    # create the functional object
    func_obj = {
        "item_number": {
            "func": clean_item_number,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [" "],
                "replace_delimitors": ["\\", "/", "(", ")"]
            },
            "target_data_type": "string"
        },
        "drawing": {
            "func": clean_drawing,
            "args": {
                "none_if_contains": [" ", "."],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "revision": {
            "func": clean_revision,
            "args": {
                "none_if_contains": [" ", "-", "/"],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "inspection_date": {
            "func": clean_inspection_date,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "datetime"
        },
        "inspector": {
            "func": clean_inspector_operator,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [".", "(", ")", "{", "}", "[", "]", "<", ">"],
                "replace_delimitors": ["\\", "/", " ", "-", ","]
            },
            "target_data_type": "string"
        },
        "disposition": {
            "func": clean_disposition,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "supplier": {
            "func": clean_supplier,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "receiver_number": {
            "func": clean_receiver_number,
            "args": {
                "none_if_contains": ["no"],
                "remove_substrings": [" "],
                "replace_delimitors": ["-", "/", ","]
            },
            "target_data_type": "string"
        },
        "purchase_order": {
            "func": clean_purchase_order,
            "args": {
                "none_if_contains": ["no"],
                "remove_substrings": [" "],
                "replace_delimitors": ["-", "/", ","]
            },
            "target_data_type": "string"
        },
        "job_order": {
            "func": clean_job_order,
            "args": {
                "none_if_contains": [".", "-"],
                "remove_substrings": [" "],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "operator": {
            "func": clean_inspector_operator,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [".", "(", ")", "{", "}", "[", "]", "<", ">"],
                "replace_delimitors": ["\\", "/", " ", "-", ","]
            },
            "target_data_type": "string"
        },
        "full_inspect_qty": {
            "func": clean_full_inspect_qty,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [" "],
                "replace_delimitors": []
            },
            "target_data_type": "float"
        },
        "received_qty": {
            "func": clean_received_qty,
            "args": {
                "none_if_contains": [" ", "/"],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "float"
        },
        "completed_qty": {
            "func": clean_completed_qty,
            "args": {
                "none_if_contains": [" ", "/", "=", ".", "-"],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "float"
        }
    }

    # create a deep copy of the raw dataframe
    std_df = raw_df.copy(deep = True)

    # apply the functional object to standardize the 'unwanted' values
    for k in func_obj:

        # reference the object children
        my_func = func_obj[k]["func"]
        my_args = func_obj[k]["args"]

        # apply the function to its column
        if my_func is not None:
            std_df.loc[:, k] = raw_df[k].apply(my_func, args = (my_args,))

    # create a reduced dataframe from the standardized dataframe
    red_df = std_df.loc[
        (std_df["item_number"].isna() == False) & 
        (std_df["drawing"].isna() == False) & 
        (std_df["revision"].isna() == False), :
    ]

    # create a deep copy of the reduced dataframe
    cln_df = red_df.copy(deep = True)

    # fix data types
    for k in func_obj:
        target = func_obj[k]["target_data_type"]
        if target == "datetime":
            cln_df[k] = pd.to_datetime(cln_df[k], format = "%Y-%m-%d")
        else:
            cln_df = cln_df.astype({ k: target })

    return cln_df

# clean the measurements dataframe
def clean_measurements(raw_df: pd.DataFrame) -> pd.DataFrame:
    
    # create the functional object
    func_obj = {
        "feature_id": {
            "func": clean_feature_id,
            "args": {
                "none_if_contains": ["!", "#", "%", "&", "(", ")", "-", "/"],
                "remove_substrings": [" ", "."],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "usl": {
            "func": clean_speclimits,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [" "],
                "replace_delimitors": []
            },
            "target_data_type": "float"
        },
        "lsl": {
            "func": clean_speclimits,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [" "],
                "replace_delimitors": []
            },
            "target_data_type": "float"
        },
        "gauge": {
            "func": clean_gauge,
            "args": {
                "none_if_contains": ["!", "#"],
                "remove_substrings": [" ", ".", "/", "-"],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        },
        "data_type": {
            "func": clean_data_type,
            "args": {
                "none_if_contains": [],
                "remove_substrings": [],
                "replace_delimitors": []
            },
            "target_data_type": "string"
        }
    }

    # create a deep copy of the raw dataframe
    std_df = raw_df.copy(deep = True)

    # apply the functional object to standardize the 'unwanted' values
    for k in func_obj:

        # reference the object children
        my_func = func_obj[k]["func"]
        my_args = func_obj[k]["args"]

        # apply the function to its column
        if my_func is not None:
            std_df.loc[:, k] = raw_df[k].apply(my_func, args = (my_args,))

    # create a reduced dataframe from the standardized dataframe
    red_df = std_df.loc[
        (std_df["gauge"].isna() == False), :
    ]

    # create a deep copy of the reduced dataframe
    cln_df = red_df.copy(deep = True)

    # fix data types
    for k in func_obj:
        target = func_obj[k]["target_data_type"]
        if target == "datetime":
            cln_df[k] = pd.to_datetime(cln_df[k], format = "%Y-%m-%d")
        else:
            cln_df = cln_df.astype({ k: target })

    return cln_df

# extract the contents of multiple electronic inspection records into lists of dictionaries
def scrape_all(qc_folder: str, anchor_search_term: str, file_extension: str, qty_limit: int = 0, is_random: bool = False, workbooks: list = []) -> tuple:

    # retrieve a list of the folder contents
    folder_contents = listdir(qc_folder)

    # filter the folder contents
    filtered_contents = list(filter(lambda item:
                                    isfile(join(qc_folder, item))
                                    and item[:1].lower() != "~"
                                    and item[:1].lower() != "_"
                                    and item[-len(file_extension):].lower() == file_extension
                                    and "template" not in item.lower()
                                    and "example" not in item.lower()
                                    and "mi" not in item.lower()
                                    and "qa1" not in item.lower(), folder_contents))

    # initialize the list of files
    files = []
    if is_random:
        files = sample(filtered_contents, qty_limit)
    else:
        files = filtered_contents

    # overwrite the list of files if requested
    if len(workbooks) > 0:
        files = workbooks

    # initialize the storage lists
    metadata_list = []
    measurements_list = []

    # initialize the counter
    iterator_count = 1

    # initialize the metadata index
    metadata_index = 0

    # iterate through the directory contents
    file_index = 1
    arr_size = len(files)
    for item in files:

        # conditional break
        if qty_limit > 0:

            # exit the iterator
            if iterator_count > qty_limit:
                break

            # increment the iterator count
            iterator_count += 1

        # interpret the current workbook
        print(f"{file_index}/{arr_size}: {item}")
        metadata_df, measurements_df = scrape_one(qc_folder, anchor_search_term, item, metadata_index)

        # advance the metadata index
        metadata_index += len(metadata_df)

        # append results to the DataFrame lists
        if metadata_df is not None:
            metadata_list.append(metadata_df)
        if measurements_df is not None:
            measurements_list.append(measurements_df)
        
        # advance the file index
        file_index += 1

    # concatenate the DataFrame lists
    if (len(metadata_list) > 0) and (len(measurements_list) > 0):
        raw_metadata_df = pd.concat(metadata_list, axis = 0, ignore_index = True)
        raw_measurement_df = pd.concat(measurements_list, axis = 0, ignore_index = True)

        # return the results
        return (raw_metadata_df, raw_measurement_df)
    else:
        return (None, None)

# save results to a binary file
def to_binary(destination_folder: str, file_name: str, data_object: tuple) -> None:

    # create, populate, then close the binary file
    my_file = open(join(destination_folder, file_name), "wb")
    dump(data_object, my_file)
    my_file.close()

# save results to csv files
def to_csv(destination_folder: str, raw_data_object: tuple = None, cln_data_object: tuple = None) -> None:

    # save raw results
    if raw_data_object is not None:
        raw_metadata_df, raw_measurements_df = raw_data_object
        raw_metadata_df.to_csv(join(destination_folder, "raw_metadata.csv"), index = False)
        raw_measurements_df.to_csv(join(destination_folder, "raw_measurements.csv"), index = False)

    # save clean results
    if cln_data_object is not None:
        cln_metadata_df, cln_measurements_df = cln_data_object
        cln_metadata_df.to_csv(join(destination_folder, "cln_metadata.csv"), index = False)
        cln_measurements_df.to_csv(join(destination_folder, "cln_measurements.csv"), index = False)

# convert binary file to multiple csv files
def binary_to_csvs(binary_folder: str, destination_folder: str, raw_binary_name: str = None, cln_binary_name: str = None) -> None:

    # open, read, then close the raw binary file
    raw_data = None
    if raw_binary_name is not None:
        raw_bin_file = open(join(binary_folder, raw_binary_name), "rb")
        raw_data = load(raw_bin_file)
        raw_bin_file.close()

    # open, read, then close the clean binary file
    cln_data = None
    if cln_binary_name is not None:
        cln_bin_file = open(join(binary_folder, cln_binary_name), "rb")
        cln_data = load(cln_bin_file)
        cln_bin_file.close()

    # save the file contents to csv files
    to_csv(destination_folder, raw_data, cln_data)

# clean csvs
def clean_csvs(destination_folder: str) -> None:

    # read raw csvs into dataframes
    raw_metadata_df = pd.read_csv(join(destination_folder, "raw_metadata.csv"))
    raw_measurements_df = pd.read_csv(join(destination_folder, "raw_measurements.csv"))

    # clean the raw dataframes
    clean_metadata_df = clean_metadata(raw_metadata_df)
    clean_measurements_df = clean_measurements(raw_measurements_df)

    # save the clean dataframes as csvs
    clean_metadata_df.to_csv(join(destination_folder, "cln_metadata.csv"), index = False)
    clean_measurements_df.to_csv(join(destination_folder, "cln_measurements.csv"), index = False)

    # saves as individual files
    to_individuals(clean_metadata_df, clean_measurements_df, join(destination_folder, "_individual_files"))

def make_uid(row, args):
    id = ""
    if row["feature_id"] == "":
        id = ""
    else:
        id = f".{row['feature_id']}"
    return f"{row.name}{id}.{args['item']}.{args['drawing']}.{args['revision']}"

# convert bulk dataframes into individual files
def to_individuals(cln_metadata_df:pd.DataFrame, cln_measurements_df:pd.DataFrame, output_dir:str) -> None:
    
    # add nominals to the measurements dataframe
    cln_measurements_df["nominal"] = (cln_measurements_df.usl + cln_measurements_df.lsl) / 2

    # define the header columns
    columns = [
        "feature_id",
        "gauge",
        "data_type",
        "nominal",
        "usl",
        "lsl"
    ]
    columns.extend([x for x in cln_measurements_df.columns.tolist() if "part" in str(x)])

    # replace nan values
    cln_metadata_df = cln_metadata_df.fillna("")
    cln_measurements_df = cln_measurements_df.fillna("")

    # iterate through the metadata dataframe
    for index, row in cln_metadata_df.iterrows():

        # get the corresponding slice from measurements
        temp_df = cln_measurements_df.loc[cln_measurements_df["metadata_id"] == row["id"], columns]

        args = {
            "item": row["item_number"],
            "drawing": row["drawing"],
            "revision": row["revision"]
        }

        temp_df["feature_id"] = temp_df.apply(make_uid, axis = 1, args = (args,))

        # create the individual csv file
        with open(join(output_dir, f"dataset_{row['id']}.csv"), "w", newline = "") as file:

            # insantiate the csv writer object
            csv_writer = csv.writer(file)

            # write the metadata rows
            csv_writer.writerow(["item_number", row["item_number"]])
            csv_writer.writerow(["drawing", row["drawing"]])
            csv_writer.writerow(["revision", row["revision"]])
            csv_writer.writerow(["inspection_date", row["inspection_date"]])
            csv_writer.writerow(["inspector", row["inspector"]])
            csv_writer.writerow(["disposition", row["disposition"]])
            csv_writer.writerow(["supplier", row["supplier"]])
            csv_writer.writerow(["receiver_number", row["receiver_number"]])
            csv_writer.writerow(["purchase_order", row["purchase_order"]])
            csv_writer.writerow(["job_order", row["job_order"]])
            csv_writer.writerow(["operator", row["operator"]])
            csv_writer.writerow(["full_inspect_qty", row["full_inspect_qty"]])
            csv_writer.writerow(["received_qty", row["received_qty"]])
            csv_writer.writerow(["completed_qty", row["completed_qty"]])

            # write the separating row
            csv_writer.writerow([""])

            # write the measurement header row
            csv_writer.writerow(columns)

            # write the measurements
            for i, r in temp_df.iterrows():
                csv_writer.writerow(r.tolist())
